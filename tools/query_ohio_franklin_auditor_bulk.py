#!/usr/bin/env python3
"""Discover, inspect, download, and stream Franklin County Auditor releases.

The Auditor publishes several independently meaningful file families through
an anonymous IIS directory tree.  This adapter preserves those release and
artifact boundaries while reusing the platform's resumable bulk transport.

Usage:
    uv run python tools/query_ohio_franklin_auditor_bulk.py source
    uv run python tools/query_ohio_franklin_auditor_bulk.py releases appraisal
    uv run python tools/query_ohio_franklin_auditor_bulk.py artifacts \
        tax-accounting --release current
    uv run python tools/query_ohio_franklin_auditor_bulk.py artifact-probe \
        daily-conveyances DailyConveyances_20260723.xlsx
    uv run python tools/query_ohio_franklin_auditor_bulk.py download \
        tax-accounting Payment2025.xlsx --destination /tmp/Payment2025.xlsx
    uv run python tools/query_ohio_franklin_auditor_bulk.py inspect-local \
        /tmp/Payment2025.xlsx --record-family payment
    uv run python tools/query_ohio_franklin_auditor_bulk.py rows \
        /tmp/Payment2025.xlsx --record-family payment \
        --release-id tax-accounting-2026-07-15 --parcel 010-000001-00
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import html
import json
import re
import shutil
import sys
import tempfile
import time
import zipfile
from collections.abc import Callable, Iterator, Mapping, Sequence
from contextlib import contextmanager
from dataclasses import dataclass, replace
from datetime import date, datetime, time as datetime_time
from pathlib import Path, PurePosixPath
from typing import Any
from urllib.parse import unquote, urljoin, urlsplit
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

try:
    from tools.lead_tracker import log_search
    from tools.output_util import add_output_args, write_output
    from tools.public_records_bulk import (
        BulkArtifact,
        BulkSourceError,
        BulkTransferClient,
        file_sha256,
    )
    from tools.public_records_contract import (
        JurisdictionMetadata,
        PublicRecordsError,
        PublicRecordsQuery,
        PublicRecordsResult,
        QueryMetadata,
        ResultStatus,
        SourceMetadata,
        canonical_json,
        sha256_fingerprint,
    )
except ImportError:
    from lead_tracker import log_search
    from output_util import add_output_args, write_output
    from public_records_bulk import (
        BulkArtifact,
        BulkSourceError,
        BulkTransferClient,
        file_sha256,
    )
    from public_records_contract import (
        JurisdictionMetadata,
        PublicRecordsError,
        PublicRecordsQuery,
        PublicRecordsResult,
        QueryMetadata,
        ResultStatus,
        SourceMetadata,
        canonical_json,
        sha256_fingerprint,
    )


SOURCE_ID = "us-oh-franklin-county-auditor-bulk"
COUNTY_GEOID = "39049"
DATA_LANDING_URL = "https://auditor.franklincountyohio.gov/Auditor/FTP"
DIRECTORY_ROOT = "https://apps.franklincountyauditor.com/"
DIRECTORY_HOST = "apps.franklincountyauditor.com"
OUTSIDE_USER_ROOT = f"{DIRECTORY_ROOT}Outside_User_Files/"
DAILY_ROOT = f"{DIRECTORY_ROOT}Daily_Conveyances/"
GIS_ROOT = f"{DIRECTORY_ROOT}GIS_Shapefiles/"
GIS_CURRENT_ROOT = f"{GIS_ROOT}CurrentExtracts/"
PARCEL_CSV_ROOT = f"{DIRECTORY_ROOT}Parcel_CSV/"
DEFAULT_TIMEOUT = 60.0
DEFAULT_RETRY_ATTEMPTS = 3
DEFAULT_MINIMUM_INTERVAL = 0.2
DEFAULT_SAMPLE_BYTES = 64
MAX_DIRECTORY_BYTES = 4 * 1024 * 1024
CSV_FIELD_SIZE_LIMIT = 16 * 1024 * 1024
CURSOR_VERSION = 1
LOCAL_CURSOR_KIND = "franklin-auditor-local-rows"
LIST_CURSOR_KIND = "franklin-auditor-list"

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

SOURCE_WARNINGS = (
    "Release-directory dates, artifact modified times, and row event dates are preserved as separate source fields.",
    "Appraisal, tax-accounting, conveyance, parcel-CSV, and GIS files remain distinct source components even when parcel identifiers overlap.",
    "Owner, payment, sale, and transfer fields are publisher observations in their native component and are not merged into a title or current-owner conclusion by this adapter.",
)

SOURCE_METADATA = SourceMetadata(
    source_id=SOURCE_ID,
    name="Franklin County Auditor Bulk Data Library",
    source_role=(
        "assessment_tax_transfer_sales_parcel_and_gis_bulk_releases"
    ),
    base_url=DATA_LANDING_URL,
    dataset_id="franklin-county-auditor-data-library",
    metadata={
        "authority": "Franklin County Auditor",
        "operator": "Franklin County Auditor",
        "directory_root": DIRECTORY_ROOT,
        "same_authority_interactive_source_id": (
            "us-oh-franklin-county-auditor-property"
        ),
        "statewide_parcel_complement": "us-oh-ogrip-statewide-parcels",
    },
)
JURISDICTION = JurisdictionMetadata(
    jurisdiction_id=COUNTY_GEOID,
    name="Franklin County, Ohio",
    state_code="OH",
    county_fips=COUNTY_GEOID,
    locality="Franklin County",
)


@dataclass(frozen=True)
class FamilyContract:
    key: str
    name: str
    root_url: str
    record_roles: tuple[str, ...]
    cadence: str
    archive_scope: str

    def to_record(self) -> dict[str, Any]:
        ref = f"BULK:{SOURCE_ID}/family/{self.key}"
        return {
            "source_id": SOURCE_ID,
            "record_kind": "bulk_dataset_family",
            "canonical_ref": ref,
            "evidence_ref": ref,
            "native_document_id": self.key,
            "family": self.key,
            "name": self.name,
            "root_url": self.root_url,
            "record_roles": list(self.record_roles),
            "cadence": self.cadence,
            "archive_scope": self.archive_scope,
            "same_authority_lineage": (
                "us-oh-franklin-county-auditor-property"
            ),
        }


FAMILY_CONTRACTS = {
    item.key: item
    for item in (
        FamilyContract(
            key="appraisal",
            name="Outside User Appraisal Files",
            root_url=OUTSIDE_USER_ROOT,
            record_roles=(
                "parcel",
                "land",
                "building",
                "dwelling",
                "improvement",
                "permit",
                "assessor_sale",
            ),
            cadence="monthly source releases",
            archive_scope="year directories beginning in 2015",
        ),
        FamilyContract(
            key="tax-accounting",
            name="Outside User Tax Accounting Files",
            root_url=OUTSIDE_USER_ROOT,
            record_roles=(
                "tax_parcel",
                "payment",
                "rental_contact",
                "special_assessment",
                "tax_detail",
                "tax_distribution",
                "transfer",
                "value",
            ),
            cadence="monthly source releases",
            archive_scope="year directories beginning in 2015",
        ),
        FamilyContract(
            key="daily-conveyances",
            name="Daily Conveyance Workbooks",
            root_url=DAILY_ROOT,
            record_roles=("daily_conveyance", "assessor_sale"),
            cadence="business-day source workbooks plus consolidated ranges",
            archive_scope="flat source-managed workbook series",
        ),
        FamilyContract(
            key="gis-shapefiles",
            name="GIS Shapefile Extracts",
            root_url=GIS_ROOT,
            record_roles=("parcel_geometry", "county_gis_context"),
            cadence="weeknight current extracts and monthly archives",
            archive_scope="year/month directories beginning in 2002",
        ),
        FamilyContract(
            key="parcel-csv",
            name="Parcel CSV",
            root_url=PARCEL_CSV_ROOT,
            record_roles=(
                "assessment_owner_observation",
                "parcel",
                "address",
                "value",
                "last_transfer",
                "building_characteristics",
            ),
            cadence="source-managed current file with year/month archives",
            archive_scope="year/month directories beginning in 1997",
        ),
    )
}
FAMILY_CHOICES = tuple(FAMILY_CONTRACTS)
RECORD_FAMILY_CHOICES = (
    "parcel",
    "value",
    "payment",
    "transfer",
    "sales",
    "daily-conveyance",
)


class FranklinBulkError(RuntimeError):
    status = ResultStatus.UNAVAILABLE
    code = "franklin_auditor_bulk_error"
    category = "source_access"
    retryable = False

    def __init__(
        self,
        message: str,
        *,
        details: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.details = dict(details or {})

    def to_contract_error(self) -> PublicRecordsError:
        return PublicRecordsError(
            code=self.code,
            message=str(self),
            category=self.category,
            retryable=self.retryable,
            details=self.details,
        )


class FranklinSelectionError(FranklinBulkError):
    code = "franklin_auditor_selection_invalid"
    category = "query_selection"


class FranklinSourceChanged(FranklinBulkError):
    status = ResultStatus.SOURCE_CHANGED
    code = "franklin_auditor_source_changed"
    category = "source_schema"


class FranklinUnavailable(FranklinBulkError):
    status = ResultStatus.UNAVAILABLE
    code = "franklin_auditor_unavailable"
    category = "source_transport"
    retryable = True


@dataclass(frozen=True)
class DirectoryEntry:
    name: str
    url: str
    relative_path: str
    is_directory: bool
    size: int | None
    modified_raw: str
    modified_at: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "url": self.url,
            "relative_path": self.relative_path,
            "is_directory": self.is_directory,
            "size": self.size,
            "modified_raw": self.modified_raw,
            "modified_at": self.modified_at,
        }


@dataclass(frozen=True)
class DirectoryListing:
    url: str
    path: str
    entries: tuple[DirectoryEntry, ...]
    fingerprint: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "url": self.url,
            "path": self.path,
            "entry_count": len(self.entries),
            "fingerprint": self.fingerprint,
        }


@dataclass(frozen=True)
class Release:
    family: str
    release_id: str
    directory_url: str
    release_date: str | None
    release_date_basis: str | None
    path_period: str | None
    directory_modified_raw: str | None
    directory_modified_at: str | None
    inline_entries: tuple[DirectoryEntry, ...] = ()

    def to_record(self) -> dict[str, Any]:
        ref = f"BULK:{SOURCE_ID}/release/{self.release_id}"
        return {
            "source_id": SOURCE_ID,
            "record_kind": "bulk_dataset_release",
            "canonical_ref": ref,
            "evidence_ref": ref,
            "native_document_id": self.release_id,
            "family": self.family,
            "release_id": self.release_id,
            "directory_url": self.directory_url,
            "release_date": self.release_date,
            "release_date_basis": self.release_date_basis,
            "path_period": self.path_period,
            "directory_modified_raw": self.directory_modified_raw,
            "directory_modified_at": self.directory_modified_at,
            "inline_artifact_count": len(self.inline_entries),
            "same_authority_lineage": (
                "us-oh-franklin-county-auditor-property"
            ),
        }


_IIS_ENTRY_RE = re.compile(
    r"(?P<date>\d{1,2}/\d{1,2}/\d{4})\s+"
    r"(?P<time>\d{1,2}:\d{2}\s+(?:AM|PM))\s+"
    r"(?P<size><dir>|\d+)\s+"
    r"<a\s+href=[\"'](?P<href>[^\"']+)[\"'][^>]*>"
    r"(?P<name>[^<]+)</a>",
    re.IGNORECASE,
)
_TITLE_RE = re.compile(r"<title>(?P<title>.*?)</title>", re.I | re.S)
_OUTSIDE_RELEASE_RE = re.compile(
    r"^(?P<date>20\d{2}-\d{2}-\d{2})\s+"
    r"(?P<family>Appraisal|Tax Accounting)$",
    re.I,
)
_DAILY_RE = re.compile(
    r"^DailyConveyances_(?P<start>20\d{6})"
    r"(?:_(?P<end>20\d{6}))?\.xlsx$",
    re.I,
)
_GIS_PREFIX_RE = re.compile(r"^(?P<date>20\d{6})_.+\.zip$", re.I)


def _official_directory_url(url: str) -> str:
    parsed = urlsplit(url)
    if parsed.scheme != "https" or parsed.hostname != DIRECTORY_HOST:
        raise FranklinSourceChanged(
            "Franklin Auditor directory URL left the verified official host",
            details={"url": url},
        )
    return url


def _directory_timestamp(date_text: str, time_text: str) -> tuple[str, str]:
    raw = f"{date_text} {time_text.upper()}"
    try:
        parsed = datetime.strptime(raw, "%m/%d/%Y %I:%M %p").replace(
            tzinfo=ZoneInfo("America/New_York")
        )
    except ValueError as error:
        raise FranklinSourceChanged(
            "Franklin Auditor directory timestamp changed format",
            details={"timestamp": raw},
        ) from error
    return raw, parsed.isoformat(timespec="minutes")


def parse_iis_listing(html_text: str, *, source_url: str) -> DirectoryListing:
    """Parse and identify one verified Franklin Auditor IIS listing."""

    _official_directory_url(source_url)
    decoded = html.unescape(html_text)
    title_match = _TITLE_RE.search(decoded)
    source_path = unquote(urlsplit(source_url).path)
    if not source_path.endswith("/"):
        source_path += "/"
    expected_title = f"{DIRECTORY_HOST} - {source_path}"
    observed_title = (
        " ".join(title_match.group("title").split())
        if title_match
        else None
    )
    if observed_title != expected_title:
        raise FranklinSourceChanged(
            "Franklin Auditor IIS listing title no longer matches its URL",
            details={
                "expected_title": expected_title,
                "observed_title": observed_title,
            },
        )

    entries: list[DirectoryEntry] = []
    for match in _IIS_ENTRY_RE.finditer(decoded):
        name = html.unescape(match.group("name")).strip()
        href = html.unescape(match.group("href")).strip()
        absolute_url = _official_directory_url(urljoin(source_url, href))
        absolute_path = unquote(urlsplit(absolute_url).path)
        if not absolute_path.startswith(source_path):
            raise FranklinSourceChanged(
                "Franklin Auditor listing entry escaped its directory",
                details={"source_path": source_path, "entry": href},
            )
        raw_size = match.group("size")
        is_directory = raw_size.casefold() == "<dir>"
        if is_directory and not absolute_path.endswith("/"):
            raise FranklinSourceChanged(
                "Franklin Auditor directory entry lost its trailing slash",
                details={"entry": href},
            )
        modified_raw, modified_at = _directory_timestamp(
            match.group("date"), match.group("time")
        )
        entries.append(
            DirectoryEntry(
                name=name,
                url=absolute_url,
                relative_path=absolute_path.lstrip("/"),
                is_directory=is_directory,
                size=None if is_directory else int(raw_size),
                modified_raw=modified_raw,
                modified_at=modified_at,
            )
        )
    if not entries:
        raise FranklinSourceChanged(
            "Franklin Auditor IIS listing contains no dated entries",
            details={"url": source_url},
        )
    entries.sort(key=lambda item: (item.name.casefold(), item.relative_path))
    fingerprint = sha256_fingerprint([item.to_dict() for item in entries])
    return DirectoryListing(
        url=source_url,
        path=source_path,
        entries=tuple(entries),
        fingerprint=fingerprint,
    )


class FranklinDirectoryClient:
    """Paced client for the anonymous Auditor directory tree."""

    def __init__(
        self,
        *,
        timeout: float = DEFAULT_TIMEOUT,
        minimum_interval: float = DEFAULT_MINIMUM_INTERVAL,
        retry_attempts: int = DEFAULT_RETRY_ATTEMPTS,
        session: requests.Session | Any | None = None,
        sleeper: Callable[[float], None] = time.sleep,
        monotonic: Callable[[], float] = time.monotonic,
    ) -> None:
        if timeout <= 0:
            raise ValueError("timeout must be positive")
        if minimum_interval < 0:
            raise ValueError("minimum_interval must not be negative")
        if retry_attempts <= 0:
            raise ValueError("retry_attempts must be positive")
        self.timeout = timeout
        self.minimum_interval = minimum_interval
        self.retry_attempts = retry_attempts
        self.session = session or requests.Session()
        self._sleeper = sleeper
        self._monotonic = monotonic
        self._last_started: float | None = None
        self._cache: dict[str, DirectoryListing] = {}
        self.request_count = 0

    def _pace(self) -> None:
        now = self._monotonic()
        if self._last_started is not None:
            remaining = self.minimum_interval - (now - self._last_started)
            if remaining > 0:
                self._sleeper(remaining)
                now = self._monotonic()
        self._last_started = now

    def listing(self, url: str, *, refresh: bool = False) -> DirectoryListing:
        url = _official_directory_url(url)
        if not refresh and url in self._cache:
            return self._cache[url]
        last_error: Exception | None = None
        for attempt in range(1, self.retry_attempts + 1):
            self._pace()
            self.request_count += 1
            try:
                response = self.session.get(
                    url,
                    timeout=self.timeout,
                    headers={
                        "Accept": "text/html,application/xhtml+xml",
                        "User-Agent": "Ithildin-Public-Records/1.0",
                    },
                )
            except requests.RequestException as error:
                last_error = error
                if attempt < self.retry_attempts:
                    self._sleeper(0.25 * (2 ** (attempt - 1)))
                    continue
                raise FranklinUnavailable(
                    f"Franklin Auditor directory request failed: {error}",
                    details={"url": url, "attempts": attempt},
                ) from error
            final_url = str(getattr(response, "url", url))
            _official_directory_url(final_url)
            status = int(getattr(response, "status_code", 0))
            if status in {429, 500, 502, 503, 504} and attempt < self.retry_attempts:
                self._sleeper(0.25 * (2 ** (attempt - 1)))
                continue
            if status < 200 or status >= 300:
                raise FranklinUnavailable(
                    f"Franklin Auditor directory returned HTTP {status}",
                    details={"url": url, "status": status},
                )
            content = bytes(getattr(response, "content", b""))
            if len(content) > MAX_DIRECTORY_BYTES:
                raise FranklinSourceChanged(
                    "Franklin Auditor directory exceeded its structural parser bound",
                    details={"url": url, "bytes": len(content)},
                )
            content_type = str(response.headers.get("Content-Type") or "")
            if "html" not in content_type.casefold():
                raise FranklinSourceChanged(
                    "Franklin Auditor directory is no longer HTML",
                    details={"url": url, "content_type": content_type},
                )
            listing = parse_iis_listing(
                content.decode("utf-8", errors="replace"),
                source_url=url,
            )
            self._cache[url] = listing
            return listing
        raise FranklinUnavailable(
            f"Franklin Auditor directory request failed: {last_error}",
            details={"url": url},
        )


def _year_directories(listing: DirectoryListing) -> list[DirectoryEntry]:
    result = [
        entry
        for entry in listing.entries
        if entry.is_directory and re.fullmatch(r"(?:19|20)\d{2}", entry.name)
    ]
    result.sort(key=lambda item: int(item.name))
    return result


def _month_directories(listing: DirectoryListing) -> list[DirectoryEntry]:
    result = [
        entry
        for entry in listing.entries
        if entry.is_directory and re.fullmatch(r"(?:0[1-9]|1[0-2])", entry.name)
    ]
    result.sort(key=lambda item: int(item.name))
    return result


def _compact_date(value: str) -> str:
    try:
        return datetime.strptime(value, "%Y%m%d").date().isoformat()
    except ValueError as error:
        raise FranklinSourceChanged(
            "Franklin Auditor filename contains an invalid calendar date",
            details={"filename_date": value},
        ) from error


def _outside_releases_for_year(
    client: FranklinDirectoryClient | Any,
    family: str,
    year_entry: DirectoryEntry,
) -> list[Release]:
    listing = client.listing(year_entry.url)
    expected_label = "appraisal" if family == "appraisal" else "tax accounting"
    releases: list[Release] = []
    for entry in listing.entries:
        if not entry.is_directory:
            continue
        match = _OUTSIDE_RELEASE_RE.fullmatch(entry.name)
        if not match or match.group("family").casefold() != expected_label:
            continue
        try:
            release_date = date.fromisoformat(match.group("date")).isoformat()
        except ValueError as error:
            raise FranklinSourceChanged(
                "Franklin Auditor release directory contains an invalid calendar date",
                details={"directory_name": entry.name},
            ) from error
        releases.append(
            Release(
                family=family,
                release_id=f"{family}-{release_date}",
                directory_url=entry.url,
                release_date=release_date,
                release_date_basis="source_directory_name",
                path_period=year_entry.name,
                directory_modified_raw=entry.modified_raw,
                directory_modified_at=entry.modified_at,
            )
        )
    releases.sort(key=lambda item: (item.release_date or "", item.release_id))
    return releases


def _discover_outside_releases(
    client: FranklinDirectoryClient | Any,
    family: str,
    *,
    year: int | None,
    all_releases: bool,
) -> list[Release]:
    root = client.listing(OUTSIDE_USER_ROOT)
    years = _year_directories(root)
    if not years:
        raise FranklinSourceChanged(
            "Franklin Auditor Outside User Files has no year directories"
        )
    if year is not None:
        selected = [entry for entry in years if int(entry.name) == year]
        if not selected:
            raise FranklinSelectionError(
                f"Outside User Files does not publish year {year}",
                details={"available_years": [int(item.name) for item in years]},
            )
    elif all_releases:
        selected = years
    else:
        selected = [years[-1]]
    releases = [
        release
        for year_entry in selected
        for release in _outside_releases_for_year(client, family, year_entry)
    ]
    if not releases:
        raise FranklinSourceChanged(
            f"Franklin Auditor publishes no {family} release directories",
            details={"selected_years": [entry.name for entry in selected]},
        )
    if year is None and not all_releases:
        return [max(releases, key=lambda item: item.release_date or "")]
    return sorted(releases, key=lambda item: item.release_id)


def _discover_daily_releases(
    client: FranklinDirectoryClient | Any,
    *,
    year: int | None,
    all_releases: bool,
) -> list[Release]:
    listing = client.listing(DAILY_ROOT)
    releases: list[Release] = []
    for entry in listing.entries:
        if entry.is_directory:
            continue
        match = _DAILY_RE.fullmatch(entry.name)
        if not match:
            continue
        start_date = _compact_date(match.group("start"))
        end_date = _compact_date(match.group("end") or match.group("start"))
        if year is not None and date.fromisoformat(end_date).year != year:
            continue
        release_id = (
            f"daily-conveyance-{start_date}"
            if start_date == end_date
            else f"daily-conveyance-{start_date}-through-{end_date}"
        )
        releases.append(
            Release(
                family="daily-conveyances",
                release_id=release_id,
                directory_url=DAILY_ROOT,
                release_date=end_date,
                release_date_basis="artifact_filename_end_date",
                path_period=(
                    start_date if start_date == end_date else f"{start_date}/{end_date}"
                ),
                directory_modified_raw=entry.modified_raw,
                directory_modified_at=entry.modified_at,
                inline_entries=(entry,),
            )
        )
    if not releases:
        raise FranklinSourceChanged(
            "Franklin Auditor daily-conveyance listing has no recognized workbooks"
        )
    releases.sort(key=lambda item: (item.release_date or "", item.release_id))
    if year is None and not all_releases:
        return [releases[-1]]
    return releases


def _current_gis_release(client: FranklinDirectoryClient | Any) -> Release:
    listing = client.listing(GIS_CURRENT_ROOT)
    file_entries = [entry for entry in listing.entries if not entry.is_directory]
    dates = {
        _compact_date(match.group("date"))
        for entry in file_entries
        if (match := _GIS_PREFIX_RE.fullmatch(entry.name))
    }
    if len(dates) != 1:
        raise FranklinSourceChanged(
            "Franklin Auditor current GIS artifacts do not share one dated prefix",
            details={"dates": sorted(dates)},
        )
    release_date = next(iter(dates))
    return Release(
        family="gis-shapefiles",
        release_id=f"gis-current-{release_date}",
        directory_url=GIS_CURRENT_ROOT,
        release_date=release_date,
        release_date_basis="artifact_filename_prefix",
        path_period="CurrentExtracts",
        directory_modified_raw=None,
        directory_modified_at=None,
        inline_entries=tuple(file_entries),
    )


def _discover_nested_periods(
    client: FranklinDirectoryClient | Any,
    *,
    root_url: str,
    family: str,
    year: int | None,
    all_releases: bool,
) -> list[Release]:
    root = client.listing(root_url)
    years = _year_directories(root)
    if not years:
        raise FranklinSourceChanged(
            f"Franklin Auditor {family} tree has no year directories"
        )
    if year is not None:
        selected = [entry for entry in years if int(entry.name) == year]
        if not selected:
            raise FranklinSelectionError(
                f"{family} does not publish year {year}",
                details={"available_years": [int(item.name) for item in years]},
            )
    elif all_releases:
        selected = years
    else:
        selected = [years[-1]]
    releases: list[Release] = []
    for year_entry in selected:
        months = _month_directories(client.listing(year_entry.url))
        for month_entry in months:
            period = f"{year_entry.name}-{month_entry.name}"
            prefix = "gis-archive" if family == "gis-shapefiles" else "parcel-csv"
            releases.append(
                Release(
                    family=family,
                    release_id=f"{prefix}-{period}",
                    directory_url=month_entry.url,
                    release_date=None,
                    release_date_basis=None,
                    path_period=period,
                    directory_modified_raw=month_entry.modified_raw,
                    directory_modified_at=month_entry.modified_at,
                )
            )
    if not releases:
        raise FranklinSourceChanged(
            f"Franklin Auditor {family} tree has no month directories",
            details={"selected_years": [entry.name for entry in selected]},
        )
    releases.sort(key=lambda item: item.path_period or "")
    if year is None and not all_releases:
        return [releases[-1]]
    return releases


def discover_releases(
    client: FranklinDirectoryClient | Any,
    family: str,
    *,
    year: int | None = None,
    all_releases: bool = False,
) -> list[Release]:
    if family not in FAMILY_CONTRACTS:
        raise FranklinSelectionError(f"unknown Franklin bulk family: {family}")
    if family in {"appraisal", "tax-accounting"}:
        return _discover_outside_releases(
            client,
            family,
            year=year,
            all_releases=all_releases,
        )
    if family == "daily-conveyances":
        return _discover_daily_releases(
            client,
            year=year,
            all_releases=all_releases,
        )
    if family == "gis-shapefiles" and year is None and not all_releases:
        return [_current_gis_release(client)]
    if family == "gis-shapefiles":
        releases = _discover_nested_periods(
            client,
            root_url=GIS_ROOT,
            family=family,
            year=year,
            all_releases=all_releases,
        )
        if all_releases:
            releases.append(_current_gis_release(client))
        return releases
    return _discover_nested_periods(
        client,
        root_url=PARCEL_CSV_ROOT,
        family=family,
        year=year,
        all_releases=all_releases,
    )


def resolve_release(
    client: FranklinDirectoryClient | Any,
    family: str,
    selector: str,
) -> Release:
    if selector == "current":
        return discover_releases(client, family)[0]
    if family == "gis-shapefiles" and selector.startswith("gis-current-"):
        current = _current_gis_release(client)
        if current.release_id == selector:
            return current
        raise FranklinSelectionError(
            f"release {selector!r} is not the published current GIS release",
            details={"current_release_id": current.release_id},
        )
    year_match = re.search(r"(?<!\d)((?:19|20)\d{2})(?!\d)", selector)
    year = int(year_match.group(1)) if year_match else None
    candidates = discover_releases(
        client,
        family,
        year=year,
        all_releases=year is None,
    )
    exact = [item for item in candidates if item.release_id == selector]
    if len(exact) != 1:
        raise FranklinSelectionError(
            f"release {selector!r} is not uniquely published for {family}",
            details={"candidate_release_ids": [item.release_id for item in candidates]},
        )
    return exact[0]


def _artifact_format(filename: str) -> tuple[str, str | None, str | None]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".xlsx":
        return "xlsx", XLSX_MEDIA_TYPE, "xlsx"
    if suffix == ".zip":
        return "zip", "application/zip", "zip"
    if suffix == ".csv":
        return "csv", "text/csv", None
    return suffix.lstrip(".") or "unknown", None, None


def _allowed_artifact(entry: DirectoryEntry, family: str) -> bool:
    if entry.is_directory:
        return False
    suffix = Path(entry.name).suffix.casefold()
    if family in {"appraisal", "tax-accounting"}:
        return suffix in {".xlsx", ".zip"}
    if family == "daily-conveyances":
        return _DAILY_RE.fullmatch(entry.name) is not None
    if family == "gis-shapefiles":
        return suffix == ".zip"
    if family == "parcel-csv":
        return suffix == ".csv"
    return False


def artifact_record(release: Release, entry: DirectoryEntry) -> dict[str, Any]:
    relative_path = entry.relative_path
    artifact_id = hashlib.sha256(relative_path.encode("utf-8")).hexdigest()[:24]
    file_format, media_type, archive_format = _artifact_format(entry.name)
    ref = f"BULK:{SOURCE_ID}/artifact/{artifact_id}"
    return {
        "source_id": SOURCE_ID,
        "record_kind": "bulk_dataset_artifact",
        "canonical_ref": ref,
        "evidence_ref": ref,
        "native_document_id": relative_path,
        "artifact_id": artifact_id,
        "family": release.family,
        "release_id": release.release_id,
        "release_date": release.release_date,
        "release_date_basis": release.release_date_basis,
        "path_period": release.path_period,
        "filename": entry.name,
        "relative_path": relative_path,
        "artifact_url": entry.url,
        "format": file_format,
        "media_type": media_type,
        "archive_format": archive_format,
        "directory_size": entry.size,
        "directory_modified_raw": entry.modified_raw,
        "directory_modified_at": entry.modified_at,
        "record_family_hint": infer_record_family(release.family, entry.name),
        "same_authority_lineage": (
            "us-oh-franklin-county-auditor-property"
        ),
    }


def artifacts_for_release(
    client: FranklinDirectoryClient | Any,
    release: Release,
) -> list[dict[str, Any]]:
    entries = (
        list(release.inline_entries)
        if release.inline_entries
        else list(client.listing(release.directory_url).entries)
    )
    records = [
        artifact_record(release, entry)
        for entry in entries
        if _allowed_artifact(entry, release.family)
    ]
    if not records:
        raise FranklinSourceChanged(
            "Franklin Auditor release contains no recognized artifacts",
            details={
                "family": release.family,
                "release_id": release.release_id,
                "directory_url": release.directory_url,
            },
        )
    records.sort(key=lambda item: (str(item["filename"]).casefold(), item["artifact_id"]))
    return records


def _bulk_artifact(record: Mapping[str, Any]) -> BulkArtifact:
    return BulkArtifact(
        artifact_id=str(record["artifact_id"]),
        url=str(record["artifact_url"]),
        filename=str(record["filename"]),
        media_type=(str(record["media_type"]) if record.get("media_type") else None),
        archive_format=(
            str(record["archive_format"])
            if record.get("archive_format")
            else None
        ),
        expected_size=(
            int(record["directory_size"])
            if record.get("directory_size") is not None
            else None
        ),
        last_modified=(
            str(record["directory_modified_at"])
            if record.get("directory_modified_at")
            else None
        ),
        metadata={
            "family": record["family"],
            "release_id": record["release_id"],
            "relative_path": record["relative_path"],
        },
    )


def resolve_artifact(
    client: FranklinDirectoryClient | Any,
    family: str,
    release_selector: str,
    artifact_selector: str,
) -> tuple[Release, dict[str, Any]]:
    release = resolve_release(client, family, release_selector)
    artifacts = artifacts_for_release(client, release)
    matches = [
        item
        for item in artifacts
        if artifact_selector
        in {
            item["filename"],
            item["artifact_id"],
            item["native_document_id"],
        }
    ]
    if len(matches) != 1:
        raise FranklinSelectionError(
            f"artifact {artifact_selector!r} is not uniquely published in {release.release_id}",
            details={"available_artifacts": [item["filename"] for item in artifacts]},
        )
    return release, matches[0]


def validate_artifact_probe(
    record: Mapping[str, Any],
    probe: Mapping[str, Any] | Any,
) -> None:
    """Validate transport metadata against the selected artifact contract."""

    def probe_value(name: str) -> Any:
        if isinstance(probe, Mapping):
            return probe.get(name)
        return getattr(probe, name, None)

    content_length = probe_value("content_length")
    if (
        record.get("directory_size") is not None
        and content_length is not None
        and int(record["directory_size"]) != int(content_length)
    ):
        raise FranklinSourceChanged(
            "artifact size differs from its official directory entry",
            details={
                "directory_size": record["directory_size"],
                "probe_size": content_length,
            },
        )
    if record.get("format") != "xlsx":
        return
    signature = probe_value("signature_hex")
    format_value = probe_value("format_hint")
    sample_size = probe_value("sample_size")
    signature_hex = str(signature or "").casefold()
    format_hint = str(format_value or "").casefold()
    if (
        int(sample_size or 0) <= 0
        or not signature_hex.startswith("504b0304")
        or format_hint not in {"zip", "xlsx"}
    ):
        raise FranklinSourceChanged(
            "XLSX artifact no longer has a ZIP container signature",
            details={
                "filename": record.get("filename"),
                "signature_hex": signature,
                "format_hint": format_value,
                "sample_size": sample_size,
            },
        )


def _encode_cursor(payload: Mapping[str, Any]) -> str:
    raw = canonical_json(payload).encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")


def _decode_cursor(cursor: str) -> Mapping[str, Any]:
    try:
        raw = base64.urlsafe_b64decode(cursor + "=" * (-len(cursor) % 4))
        value = json.loads(raw)
    except (ValueError, TypeError, json.JSONDecodeError) as error:
        raise FranklinSelectionError("invalid continuation cursor") from error
    if not isinstance(value, Mapping):
        raise FranklinSelectionError("invalid continuation cursor payload")
    return value


def paginate_records(
    records: Sequence[Mapping[str, Any]],
    *,
    selection: Mapping[str, Any],
    limit: int | None,
    cursor: str | None,
) -> tuple[list[dict[str, Any]], str | None]:
    if limit is not None and limit <= 0:
        raise FranklinSelectionError("limit must be positive")
    detached = [json.loads(canonical_json(record)) for record in records]
    selection_fingerprint = sha256_fingerprint(selection)
    listing_fingerprint = sha256_fingerprint(detached)
    start = 0
    if cursor:
        payload = _decode_cursor(cursor)
        if payload.get("version") != CURSOR_VERSION or payload.get("kind") != LIST_CURSOR_KIND:
            raise FranklinSelectionError("continuation cursor has another contract")
        if payload.get("selection") != selection_fingerprint:
            raise FranklinSelectionError(
                "continuation cursor belongs to another release selection"
            )
        if payload.get("listing") != listing_fingerprint:
            raise FranklinSourceChanged(
                "Franklin Auditor listing changed since the continuation was issued"
            )
        try:
            start = int(payload["next_index"])
        except (KeyError, TypeError, ValueError) as error:
            raise FranklinSelectionError(
                "continuation cursor lacks a valid next index"
            ) from error
        if start < 0 or start > len(detached):
            raise FranklinSelectionError("continuation cursor index is out of range")
    stop = len(detached) if limit is None else min(len(detached), start + limit)
    page = detached[start:stop]
    next_cursor = None
    if stop < len(detached):
        next_cursor = _encode_cursor(
            {
                "version": CURSOR_VERSION,
                "kind": LIST_CURSOR_KIND,
                "selection": selection_fingerprint,
                "listing": listing_fingerprint,
                "next_index": stop,
            }
        )
    return page, next_cursor


_NORMALIZE_HEADER_RE = re.compile(r"[^a-z0-9]+")
_PARCEL_HEADERS = {
    "parcelid",
    "parcelnumber",
    "parcelno",
    "parcel",
    "parid",
    "pid",
}
_DATE_HEADERS = {
    "effectivedate",
    "saledt",
    "saledate",
    "transferdate",
    "trandate",
    "trndt",
}
_PRICE_HEADERS = {
    "amount",
    "salesprice",
    "saleprice",
    "price",
    "consideration",
}


def _header_key(value: Any) -> str:
    return _NORMALIZE_HEADER_RE.sub("", str(value or "").casefold())


def _header_contract(record_family: str, headers: Sequence[Any]) -> bool:
    keys = {_header_key(item) for item in headers if _header_key(item)}
    has_parcel = bool(keys & _PARCEL_HEADERS)
    if record_family == "payment":
        return {
            "parcelid",
            "effectivedate",
            "taxyear",
            "billtype",
            "amount",
        }.issubset(keys)
    if record_family == "daily-conveyance":
        return {
            "conveynumber",
            "parcelnumber",
            "saledate",
            "salesprice",
            "ownername1",
            "priorownername1",
        }.issubset(keys)
    if record_family == "parcel":
        return has_parcel and bool(
            keys
            & {
                "ownername",
                "ownername1",
                "name1",
                "siteaddress",
                "streetaddress",
                "apprtot",
                "annualtax",
                "anntax",
            }
        )
    if record_family == "value":
        return has_parcel and any("value" in key or "appr" in key for key in keys)
    if record_family == "transfer":
        return has_parcel and bool(
            keys
            & {
                "effectivedate",
                "transferdate",
                "trandate",
                "conveynumber",
                "instrument",
                "instrumentnumber",
                "grantor",
                "grantee",
            }
        )
    if record_family == "sales":
        return has_parcel and bool(keys & _DATE_HEADERS) and bool(keys & _PRICE_HEADERS)
    raise FranklinSelectionError(f"unknown row record family: {record_family}")


def infer_record_family(release_family: str, filename: str) -> str | None:
    stem = Path(filename).stem.casefold()
    if release_family == "daily-conveyances":
        return "daily-conveyance"
    if release_family == "parcel-csv":
        return "parcel"
    if release_family == "appraisal":
        if stem.startswith("sales"):
            return "sales"
        if stem == "parcel":
            return "parcel"
    if release_family == "tax-accounting":
        if stem.startswith("payment"):
            return "payment"
        if stem.startswith("transfer"):
            return "transfer"
        if stem == "value":
            return "value"
        if stem == "parcel":
            return "parcel"
    return None


def _json_cell(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (date, datetime, datetime_time)):
        return value.isoformat()
    return str(value)


def _find_header(
    rows: Iterator[tuple[int, Sequence[Any]]],
    *,
    record_family: str,
    header_scan_rows: int,
) -> tuple[int, list[str], list[tuple[int, list[Any]]]]:
    buffered: list[tuple[int, list[Any]]] = []
    for row_number, raw_values in rows:
        values = [_json_cell(value) for value in raw_values]
        if row_number <= header_scan_rows and _header_contract(record_family, values):
            headers = [str(value).strip() if value is not None else "" for value in values]
            if any(not header for header in headers):
                raise FranklinSourceChanged(
                    "Franklin Auditor header row contains a blank column name",
                    details={"row_number": row_number, "headers": headers},
                )
            return row_number, headers, buffered
        buffered.append((row_number, values))
        if row_number >= header_scan_rows:
            break
    raise FranklinSourceChanged(
        "Franklin Auditor artifact lacks the selected row-family header contract",
        details={
            "record_family": record_family,
            "header_scan_rows": header_scan_rows,
        },
    )


def _unique_field_names(headers: Sequence[str]) -> list[str]:
    counts: dict[str, int] = {}
    output: list[str] = []
    for header in headers:
        counts[header] = counts.get(header, 0) + 1
        output.append(header if counts[header] == 1 else f"{header}#{counts[header]}")
    return output


@contextmanager
def _selected_local_artifact(
    artifact_path: Path,
    *,
    member: str | None,
) -> Iterator[tuple[Path, dict[str, Any] | None]]:
    if not artifact_path.is_file():
        raise FranklinSelectionError(
            "local artifact does not exist",
            details={"artifact": str(artifact_path)},
        )
    if artifact_path.suffix.casefold() != ".zip":
        if member is not None:
            raise FranklinSelectionError("--member applies only to ZIP artifacts")
        yield artifact_path, None
        return
    try:
        archive = zipfile.ZipFile(artifact_path)
    except (OSError, zipfile.BadZipFile) as error:
        raise FranklinSourceChanged("Franklin Auditor ZIP is invalid") from error
    with archive:
        members = {item.filename: item for item in archive.infolist()}
        if member is None:
            raise FranklinSelectionError(
                "a ZIP artifact requires an explicit --member",
                details={"members": sorted(members)},
            )
        info = members.get(member)
        if info is None or info.is_dir():
            raise FranklinSelectionError(
                "selected ZIP member is not a published file",
                details={"member": member, "members": sorted(members)},
            )
        pure = PurePosixPath(info.filename)
        if pure.is_absolute() or ".." in pure.parts:
            raise FranklinSourceChanged("Franklin Auditor ZIP member path is unsafe")
        if info.flag_bits & 0x1:
            raise FranklinSourceChanged("Franklin Auditor ZIP member is encrypted")
        suffix = Path(info.filename).suffix
        with tempfile.TemporaryDirectory(prefix="osint-franklin-auditor-", dir="/tmp") as workdir:
            selected = Path(workdir) / f"selected{suffix}"
            with archive.open(info) as source, selected.open("wb") as destination:
                shutil.copyfileobj(source, destination, length=1024 * 1024)
            yield selected, {
                "archive_member": info.filename,
                "member_size": info.file_size,
                "member_compressed_size": info.compress_size,
                "member_crc32": f"{info.CRC:08x}",
                "member_sha256": file_sha256(selected),
            }


def _iter_xlsx_rows(
    path: Path,
    *,
    sheet_name: str | None,
) -> tuple[str, Iterator[tuple[int, Sequence[Any]]], Callable[[], None], list[str]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, InvalidFileException, zipfile.BadZipFile) as error:
        raise FranklinSourceChanged("Franklin Auditor XLSX is invalid") from error
    sheets = list(workbook.sheetnames)
    if not sheets:
        workbook.close()
        raise FranklinSourceChanged("Franklin Auditor workbook has no worksheets")
    selected = sheet_name or sheets[0]
    if selected not in sheets:
        workbook.close()
        raise FranklinSelectionError(
            "selected worksheet is not present",
            details={"sheet": selected, "sheets": sheets},
        )
    worksheet = workbook[selected]

    def iterator() -> Iterator[tuple[int, Sequence[Any]]]:
        for row_number, row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            yield row_number, row

    return selected, iterator(), workbook.close, sheets


def _text_rows(
    path: Path,
    *,
    encoding: str,
) -> tuple[Iterator[tuple[int, Sequence[Any]]], Callable[[], None], str]:
    delimiter = "\t" if path.suffix.casefold() in {".tsv", ".txt"} else ","
    csv.field_size_limit(CSV_FIELD_SIZE_LIMIT)
    handle = path.open("r", encoding=encoding, newline="")
    reader = csv.reader(handle, delimiter=delimiter)

    def iterator() -> Iterator[tuple[int, Sequence[Any]]]:
        for row_number, row in enumerate(reader, start=1):
            yield row_number, row

    return iterator(), handle.close, delimiter


def inspect_local_artifact(
    artifact_path: Path,
    *,
    record_family: str | None,
    member: str | None,
    sheet_name: str | None,
    sample_rows: int,
    header_scan_rows: int,
    encoding: str,
) -> dict[str, Any]:
    if sample_rows < 0:
        raise FranklinSelectionError("sample_rows must not be negative")
    if header_scan_rows <= 0:
        raise FranklinSelectionError("header_scan_rows must be positive")
    if not artifact_path.is_file():
        raise FranklinSelectionError(
            "local artifact does not exist",
            details={"artifact": str(artifact_path)},
        )
    artifact_hash = file_sha256(artifact_path)
    artifact_size = artifact_path.stat().st_size
    if artifact_path.suffix.casefold() == ".zip" and member is None:
        try:
            with zipfile.ZipFile(artifact_path) as archive:
                members = []
                for info in archive.infolist():
                    pure = PurePosixPath(info.filename)
                    if pure.is_absolute() or ".." in pure.parts:
                        raise FranklinSourceChanged(
                            "Franklin Auditor ZIP member path is unsafe"
                        )
                    if info.flag_bits & 0x1:
                        raise FranklinSourceChanged(
                            "Franklin Auditor ZIP member is encrypted"
                        )
                    members.append(
                        {
                            "name": info.filename,
                            "size": info.file_size,
                            "compressed_size": info.compress_size,
                            "crc32": f"{info.CRC:08x}",
                            "is_directory": info.is_dir(),
                        }
                    )
        except (OSError, zipfile.BadZipFile) as error:
            raise FranklinSourceChanged("Franklin Auditor ZIP is invalid") from error
        return {
            "source_id": SOURCE_ID,
            "record_kind": "bulk_artifact_inspection",
            "native_document_id": artifact_hash,
            "artifact_path": str(artifact_path.resolve()),
            "artifact_filename": artifact_path.name,
            "artifact_size": artifact_size,
            "artifact_sha256": artifact_hash,
            "archive_members": members,
            "schema_state": "member_selection_required",
        }
    if record_family is None:
        raise FranklinSelectionError(
            "schema inspection requires --record-family for this artifact"
        )
    with _selected_local_artifact(artifact_path, member=member) as (selected, member_meta):
        suffix = selected.suffix.casefold()
        if suffix == ".xlsx":
            selected_sheet, rows, closer, sheets = _iter_xlsx_rows(
                selected, sheet_name=sheet_name
            )
            delimiter = None
        elif suffix in {".csv", ".tsv", ".txt"}:
            if sheet_name is not None:
                raise FranklinSelectionError("--sheet applies only to XLSX artifacts")
            rows, closer, delimiter = _text_rows(selected, encoding=encoding)
            selected_sheet = None
            sheets = []
        else:
            raise FranklinSelectionError(
                f"unsupported local artifact format: {suffix or 'none'}"
            )
        try:
            header_row, headers, _ = _find_header(
                rows,
                record_family=record_family,
                header_scan_rows=header_scan_rows,
            )
            samples: list[dict[str, Any]] = []
            for row_number, values in rows:
                if len(samples) >= sample_rows:
                    break
                normalized = [_json_cell(value) for value in values]
                samples.append(
                    {
                        "source_row_number": row_number,
                        "raw_values": normalized,
                        "source_fields": _source_field_map(headers, normalized),
                    }
                )
        finally:
            closer()
    return {
        "source_id": SOURCE_ID,
        "record_kind": "bulk_artifact_inspection",
        "native_document_id": artifact_hash,
        "artifact_path": str(artifact_path.resolve()),
        "artifact_filename": artifact_path.name,
        "artifact_size": artifact_size,
        "artifact_sha256": artifact_hash,
        "archive_member": member_meta,
        "record_family": record_family,
        "worksheets": sheets,
        "selected_sheet": selected_sheet,
        "delimiter": delimiter,
        "header_row_number": header_row,
        "raw_headers": headers,
        "header_fingerprint": sha256_fingerprint(headers),
        "sample_rows": samples,
        "row_count_state": "not_scanned",
    }


def _source_field_map(headers: Sequence[str], values: Sequence[Any]) -> dict[str, Any]:
    names = _unique_field_names(headers)
    if len(values) > len(names):
        names.extend(
            f"__extra_column_{index}"
            for index in range(1, len(values) - len(names) + 1)
        )
    padded = list(values) + [None] * max(0, len(names) - len(values))
    return dict(zip(names, padded, strict=True))


def _first_field(fields: Mapping[str, Any], aliases: set[str]) -> Any:
    for key, value in fields.items():
        if _header_key(key.split("#", 1)[0]) in aliases and value not in (None, ""):
            return value
    return None


def _row_date(fields: Mapping[str, Any]) -> str | None:
    value = _first_field(fields, _DATE_HEADERS)
    if value is None:
        return None
    if isinstance(value, str):
        candidate = value.strip()
        try:
            return date.fromisoformat(candidate).isoformat()
        except ValueError:
            pass
        try:
            return datetime.fromisoformat(candidate).date().isoformat()
        except ValueError:
            pass
        for pattern in ("%m/%d/%Y", "%m/%d/%Y %H:%M:%S"):
            try:
                return datetime.strptime(candidate, pattern).date().isoformat()
            except ValueError:
                continue
        return None
    return None


def _normalized_parcel(value: Any) -> str | None:
    if value in (None, ""):
        return None
    normalized = re.sub(r"[^A-Za-z0-9]+", "", str(value)).upper()
    return normalized or None


def _parsed_fields(record_family: str, fields: Mapping[str, Any]) -> dict[str, Any]:
    owners = [
        value
        for key, value in fields.items()
        if _header_key(key.split("#", 1)[0])
        in {
            "ownername",
            "ownername1",
            "ownername2",
            "name1",
            "name2",
        }
        and value not in (None, "")
    ]
    prior_owners = [
        value
        for key, value in fields.items()
        if _header_key(key.split("#", 1)[0])
        in {"priorownername", "priorownername1", "priorownername2"}
        and value not in (None, "")
    ]
    is_exempt = _first_field(
        fields,
        {"isexempt", "exemptsale", "isexemptsale"},
    )
    sale_type = _first_field(fields, {"saletype"})
    sale_validity = _first_field(
        fields,
        {"valid", "validsale", "salevalidity", "validity", "isvalidsale"},
    )
    instrument_number = _first_field(
        fields,
        {"instruno", "instrumentno", "instrumentnumber"},
    )
    conveyance_number = _first_field(
        fields,
        {"conveynumber", "conveyanceno", "conveyancenum", "conveyancenumber"},
    )
    return {
        "record_family": record_family,
        "parcel_id": _first_field(fields, _PARCEL_HEADERS),
        "event_date": _row_date(fields),
        "amount": _first_field(fields, _PRICE_HEADERS),
        "owner_names": owners,
        "prior_owner_names": prior_owners,
        "instrument": _first_field(
            fields,
            {"instrument", "instrumenttype"},
        ),
        "instrument_number": instrument_number,
        "conveyance_number": conveyance_number,
        "tax_year": _first_field(fields, {"taxyear"}),
        "bill_type": _first_field(fields, {"billtype"}),
        "is_exempt": is_exempt,
        "sale_type": sale_type,
        "sale_validity": sale_validity,
        "source_sale_flags": {
            "is_exempt": is_exempt,
            "sale_type": sale_type,
            "sale_validity": sale_validity,
        },
    }


def _row_matches(
    fields: Mapping[str, Any],
    *,
    query: str | None,
    parcel: str | None,
    from_date: str | None,
    to_date: str | None,
) -> bool:
    if query:
        haystack = canonical_json(fields).casefold()
        if query.casefold() not in haystack:
            return False
    if parcel:
        observed = _normalized_parcel(_first_field(fields, _PARCEL_HEADERS))
        if observed != _normalized_parcel(parcel):
            return False
    event_date = _row_date(fields)
    if from_date and (event_date is None or event_date < from_date):
        return False
    if to_date and (event_date is None or event_date > to_date):
        return False
    return True


def _validate_iso_date(value: str | None, field_name: str) -> str | None:
    if value is None:
        return None
    try:
        return date.fromisoformat(value).isoformat()
    except ValueError as error:
        raise FranklinSelectionError(f"{field_name} must be YYYY-MM-DD") from error


def stream_local_rows(
    artifact_path: Path,
    *,
    record_family: str,
    release_id: str,
    release_date: str | None,
    source_url: str | None,
    member: str | None,
    sheet_name: str | None,
    query: str | None,
    parcel: str | None,
    from_date: str | None,
    to_date: str | None,
    limit: int | None,
    cursor: str | None,
    header_scan_rows: int,
    encoding: str,
) -> tuple[list[dict[str, Any]], str | None, dict[str, Any]]:
    if limit is not None and limit <= 0:
        raise FranklinSelectionError("limit must be positive")
    release_date = _validate_iso_date(release_date, "release_date")
    from_date = _validate_iso_date(from_date, "from_date")
    to_date = _validate_iso_date(to_date, "to_date")
    if from_date and to_date and from_date > to_date:
        raise FranklinSelectionError("from_date must not be after to_date")
    artifact_hash = file_sha256(artifact_path)
    criteria = {
        "record_family": record_family,
        "release_id": release_id,
        "release_date": release_date,
        "source_url": source_url,
        "member": member,
        "sheet": sheet_name,
        "query": query,
        "parcel": parcel,
        "from_date": from_date,
        "to_date": to_date,
    }
    criteria_fingerprint = sha256_fingerprint(criteria)
    start_row = 1
    if cursor:
        payload = _decode_cursor(cursor)
        if payload.get("version") != CURSOR_VERSION or payload.get("kind") != LOCAL_CURSOR_KIND:
            raise FranklinSelectionError("row cursor has another contract")
        if payload.get("criteria") != criteria_fingerprint:
            raise FranklinSelectionError("row cursor belongs to another query")
        if payload.get("artifact_sha256") != artifact_hash:
            raise FranklinSourceChanged(
                "local Franklin Auditor artifact changed since the cursor was issued"
            )
        try:
            start_row = int(payload["next_row"])
        except (KeyError, TypeError, ValueError) as error:
            raise FranklinSelectionError("row cursor lacks a valid next row") from error
        if start_row <= 0:
            raise FranklinSelectionError("row cursor next row must be positive")

    output: list[dict[str, Any]] = []
    next_cursor = None
    with _selected_local_artifact(artifact_path, member=member) as (selected, member_meta):
        suffix = selected.suffix.casefold()
        if suffix == ".xlsx":
            selected_sheet, rows, closer, _ = _iter_xlsx_rows(
                selected, sheet_name=sheet_name
            )
        elif suffix in {".csv", ".tsv", ".txt"}:
            if sheet_name is not None:
                raise FranklinSelectionError("--sheet applies only to XLSX artifacts")
            rows, closer, _ = _text_rows(selected, encoding=encoding)
            selected_sheet = None
        else:
            raise FranklinSelectionError(
                f"unsupported local artifact format: {suffix or 'none'}"
            )
        try:
            header_row, headers, _ = _find_header(
                rows,
                record_family=record_family,
                header_scan_rows=header_scan_rows,
            )
            for row_number, raw_values in rows:
                if row_number < start_row:
                    continue
                values = [_json_cell(value) for value in raw_values]
                if not any(value not in (None, "") for value in values):
                    continue
                fields = _source_field_map(headers, values)
                if not _row_matches(
                    fields,
                    query=query,
                    parcel=parcel,
                    from_date=from_date,
                    to_date=to_date,
                ):
                    continue
                if limit is not None and len(output) >= limit:
                    next_cursor = _encode_cursor(
                        {
                            "version": CURSOR_VERSION,
                            "kind": LOCAL_CURSOR_KIND,
                            "criteria": criteria_fingerprint,
                            "artifact_sha256": artifact_hash,
                            "next_row": row_number,
                        }
                    )
                    break
                location = canonical_json(
                    {
                        "archive_member": member,
                        "worksheet": selected_sheet,
                        "file": (
                            artifact_path.name
                            if member is None and selected_sheet is None
                            else None
                        ),
                    }
                )
                occurrence = (
                    f"{release_id}:{artifact_hash}:{location}:row:{row_number}"
                )
                native_id = hashlib.sha256(
                    occurrence.encode("utf-8")
                ).hexdigest()
                ref = f"BULK:{SOURCE_ID}/row/{native_id}"
                parsed = _parsed_fields(record_family, fields)
                output.append(
                    {
                        "source_id": SOURCE_ID,
                        "record_kind": f"{record_family.replace('-', '_')}_row_observation",
                        "canonical_ref": ref,
                        "evidence_ref": ref,
                        "native_document_id": native_id,
                        "native_occurrence": occurrence,
                        "release_id": release_id,
                        "release_date": release_date,
                        "artifact_filename": artifact_path.name,
                        "artifact_sha256": artifact_hash,
                        "artifact_size": artifact_path.stat().st_size,
                        "artifact_source_url": source_url,
                        "archive_member": member_meta,
                        "worksheet": selected_sheet,
                        "source_row_number": row_number,
                        "header_row_number": header_row,
                        "raw_headers": headers,
                        "raw_values": values,
                        "source_fields": fields,
                        "parsed_fields": parsed,
                        "join_candidates": {
                            "county_geoid": COUNTY_GEOID,
                            "parcel_id": parsed["parcel_id"],
                            "normalized_parcel_id": _normalized_parcel(
                                parsed["parcel_id"]
                            ),
                        },
                        "same_authority_lineage": (
                            "us-oh-franklin-county-auditor-property"
                        ),
                    }
                )
        finally:
            closer()
    metadata = {
        "artifact_sha256": artifact_hash,
        "artifact_size": artifact_path.stat().st_size,
        "record_family": record_family,
        "release_id": release_id,
        "release_date": release_date,
        "source_url": source_url,
        "header_scan_rows": header_scan_rows,
        "returned_rows": len(output),
        "continuation_available": next_cursor is not None,
    }
    return output, next_cursor, metadata


def _public_query(
    operation: str,
    parameters: Mapping[str, Any],
    *,
    limit: int | None = None,
    cursor: str | None = None,
) -> PublicRecordsQuery:
    return PublicRecordsQuery(
        source=SOURCE_METADATA,
        jurisdiction=JURISDICTION,
        query=QueryMetadata(
            operation=operation,
            parameters=parameters,
            requested_limit=limit,
            cursor=cursor,
        ),
    )


def build_query(args: argparse.Namespace) -> PublicRecordsQuery:
    parameters = {
        key: str(value) if isinstance(value, Path) else value
        for key, value in vars(args).items()
        if key
        not in {
            "output",
            "json_out",
            "timeout",
            "minimum_interval",
            "retry_attempts",
            "chunk_size",
        }
        and key not in {"limit", "cursor"}
        and value is not None
    }
    return _public_query(
        args.command,
        parameters,
        limit=getattr(args, "limit", None),
        cursor=getattr(args, "cursor", None),
    )


def _directory_client(args: argparse.Namespace) -> FranklinDirectoryClient:
    return FranklinDirectoryClient(
        timeout=args.timeout,
        minimum_interval=args.minimum_interval,
        retry_attempts=args.retry_attempts,
    )


def _bulk_client(args: argparse.Namespace) -> BulkTransferClient:
    return BulkTransferClient(
        timeout=args.timeout,
        max_attempts=args.retry_attempts,
        chunk_size=args.chunk_size,
    )


def _failure(query: PublicRecordsQuery, error: Exception) -> PublicRecordsResult:
    if isinstance(error, FranklinBulkError):
        return PublicRecordsResult.failure(
            query,
            error.status,
            [error.to_contract_error()],
            warnings=SOURCE_WARNINGS,
        )
    if isinstance(error, BulkSourceError):
        return PublicRecordsResult.failure(
            query,
            error.result_status,
            [error.to_contract_error()],
            warnings=SOURCE_WARNINGS,
        )
    return PublicRecordsResult.failure(
        query,
        ResultStatus.UNAVAILABLE,
        [
            PublicRecordsError(
                code="franklin_auditor_local_io_failed",
                message=str(error),
                category="local_io",
            )
        ],
        warnings=SOURCE_WARNINGS,
    )


def _log(query: PublicRecordsQuery, count: int | None) -> None:
    try:
        log_search(canonical_json(query.to_dict()), SOURCE_ID, count)
    except Exception as error:
        print(f"Warning: could not log search: {error}", file=sys.stderr)


def execute(
    args: argparse.Namespace,
    *,
    directory_client: FranklinDirectoryClient | Any | None = None,
    bulk_client: BulkTransferClient | Any | None = None,
    log_results: bool = True,
) -> PublicRecordsResult:
    query = build_query(args)
    try:
        if args.command == "source":
            record = {
                "source_id": SOURCE_ID,
                "record_kind": "bulk_source_contract",
                "canonical_ref": f"BULK:{SOURCE_ID}/source",
                "native_document_id": SOURCE_ID,
                "official_data_landing": DATA_LANDING_URL,
                "anonymous_directory_root": DIRECTORY_ROOT,
                "families": [
                    FAMILY_CONTRACTS[key].to_record() for key in FAMILY_CHOICES
                ],
                "identity_contract": {
                    "release": "family plus source-published release directory or filename date",
                    "artifact": "official relative path",
                    "artifact_version": "size, modified timestamp, ETag when present, and computed SHA-256 after download",
                    "row_occurrence": "release, artifact SHA-256, archive member and worksheet when applicable, and physical source row",
                    "parcel_join": "source-published parcel identifier retained as a join candidate",
                },
            }
            result = PublicRecordsResult.success(query, [record], warnings=SOURCE_WARNINGS)
        elif args.command == "families":
            result = PublicRecordsResult.success(
                query,
                [FAMILY_CONTRACTS[key].to_record() for key in FAMILY_CHOICES],
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "releases":
            client = directory_client or _directory_client(args)
            releases = discover_releases(
                client,
                args.family,
                year=args.year,
                all_releases=args.all_releases,
            )
            records, next_cursor = paginate_records(
                [item.to_record() for item in releases],
                selection={
                    "operation": "releases",
                    "family": args.family,
                    "year": args.year,
                    "all_releases": args.all_releases,
                },
                limit=args.limit,
                cursor=args.cursor,
            )
            result = PublicRecordsResult(
                query=query,
                status=ResultStatus.PARTIAL if next_cursor else ResultStatus.OK,
                records=records,
                next_cursor=next_cursor,
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "artifacts":
            client = directory_client or _directory_client(args)
            release = resolve_release(client, args.family, args.release)
            records, next_cursor = paginate_records(
                artifacts_for_release(client, release),
                selection={
                    "operation": "artifacts",
                    "family": args.family,
                    "release": args.release,
                    "resolved_release_id": release.release_id,
                },
                limit=args.limit,
                cursor=args.cursor,
            )
            result = PublicRecordsResult(
                query=query,
                status=ResultStatus.PARTIAL if next_cursor else ResultStatus.OK,
                records=records,
                next_cursor=next_cursor,
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "artifact-probe":
            client = directory_client or _directory_client(args)
            release, record = resolve_artifact(
                client, args.family, args.release, args.artifact
            )
            probe = (bulk_client or _bulk_client(args)).probe(
                _bulk_artifact(record), sample_bytes=args.sample_bytes
            )
            validate_artifact_probe(record, probe)
            output = dict(record)
            output["release"] = release.to_record()
            output["artifact_probe"] = probe.to_dict()
            result = PublicRecordsResult.success(query, [output], warnings=SOURCE_WARNINGS)
        elif args.command == "download":
            client = directory_client or _directory_client(args)
            release, record = resolve_artifact(
                client, args.family, args.release, args.artifact
            )
            artifact = _bulk_artifact(record)
            if args.expected_sha256:
                artifact = replace(artifact, expected_sha256=args.expected_sha256)
            destination = Path(args.destination)
            if destination.exists() and not args.overwrite:
                raise FranklinSelectionError(
                    "download destination exists; use --overwrite to replace it",
                    details={"destination": str(destination)},
                )
            receipt = (bulk_client or _bulk_client(args)).download(
                artifact,
                destination,
                resume=args.resume,
                max_bytes=args.max_download_bytes,
            )
            output = dict(record)
            output["release"] = release.to_record()
            output["download"] = receipt.to_dict()
            result = PublicRecordsResult.success(
                query,
                [output],
                raw_artifact_refs=(receipt.path,),
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "inspect-local":
            record = inspect_local_artifact(
                Path(args.artifact),
                record_family=args.record_family,
                member=args.member,
                sheet_name=args.sheet,
                sample_rows=args.sample_rows,
                header_scan_rows=args.header_scan_rows,
                encoding=args.encoding,
            )
            result = PublicRecordsResult.success(
                query,
                [record],
                raw_artifact_refs=(str(Path(args.artifact).resolve()),),
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "rows":
            records, next_cursor, metadata = stream_local_rows(
                Path(args.artifact),
                record_family=args.record_family,
                release_id=args.release_id,
                release_date=args.release_date,
                source_url=args.source_url,
                member=args.member,
                sheet_name=args.sheet,
                query=args.query,
                parcel=args.parcel,
                from_date=args.from_date,
                to_date=args.to_date,
                limit=args.limit,
                cursor=args.cursor,
                header_scan_rows=args.header_scan_rows,
                encoding=args.encoding,
            )
            for record in records:
                record["retrieval_snapshot"] = metadata
            if next_cursor:
                result = PublicRecordsResult(
                    query=query,
                    status=ResultStatus.PARTIAL,
                    records=records,
                    next_cursor=next_cursor,
                    raw_artifact_refs=(str(Path(args.artifact).resolve()),),
                    warnings=SOURCE_WARNINGS,
                )
            else:
                result = PublicRecordsResult.success(
                    query,
                    records,
                    raw_artifact_refs=(str(Path(args.artifact).resolve()),),
                    warnings=SOURCE_WARNINGS,
                )
        elif args.command == "probe":
            client = directory_client or _directory_client(args)
            root = client.listing(DIRECTORY_ROOT)
            required = {
                "Daily_Conveyances",
                "GIS_Shapefiles",
                "Outside_User_Files",
                "Parcel_CSV",
            }
            observed = {item.name for item in root.entries if item.is_directory}
            if not required.issubset(observed):
                raise FranklinSourceChanged(
                    "Franklin Auditor root listing lost required data families",
                    details={"missing": sorted(required - observed)},
                )
            current = {
                family: discover_releases(client, family)[0]
                for family in FAMILY_CHOICES
            }
            daily_artifact = artifacts_for_release(
                client, current["daily-conveyances"]
            )[0]
            artifact_probe = (bulk_client or _bulk_client(args)).probe(
                _bulk_artifact(daily_artifact),
                sample_bytes=args.sample_bytes,
            )
            validate_artifact_probe(daily_artifact, artifact_probe)
            output = {
                "source_id": SOURCE_ID,
                "record_kind": "source_health_check",
                "canonical_ref": f"BULK:{SOURCE_ID}/source-health/bounded-live-probe",
                "native_document_id": "bounded-live-probe",
                "status": "ok",
                "official_data_landing": DATA_LANDING_URL,
                "anonymous_directory_root": DIRECTORY_ROOT,
                "root_listing": root.to_dict(),
                "current_releases": {
                    key: value.to_record() for key, value in current.items()
                },
                "sampled_artifact": daily_artifact,
                "artifact_probe": artifact_probe.to_dict(),
                "large_artifacts_downloaded": False,
                "directory_request_count": getattr(client, "request_count", None),
                "sample_bytes_requested": args.sample_bytes,
            }
            result = PublicRecordsResult.success(query, [output], warnings=SOURCE_WARNINGS)
        else:
            raise FranklinSelectionError(f"unsupported command: {args.command}")
    except (
        FranklinBulkError,
        BulkSourceError,
        OSError,
        ValueError,
        csv.Error,
    ) as error:
        result = _failure(query, error)

    if log_results and args.command not in {"source", "families", "probe"}:
        count = (
            len(result.records)
            if result.status
            in {ResultStatus.OK, ResultStatus.NO_RESULTS, ResultStatus.PARTIAL}
            else None
        )
        _log(query, count)
    return result


def _emit(result: PublicRecordsResult, args: argparse.Namespace) -> None:
    payload = result.to_dict()
    if write_output(
        payload,
        args,
        summary=f"Franklin Auditor bulk {args.command} ({result.status.value})",
    ):
        return
    if args.json_out:
        print(json.dumps(payload, indent=2, sort_keys=True))
        return
    print(
        f"Franklin Auditor bulk {args.command}: {result.status.value} "
        f"({len(result.records)} records)"
    )
    if result.next_cursor:
        print(f"Next cursor: {result.next_cursor}")
    for error in result.errors:
        print(f"ERROR [{error.code}]: {error.message}", file=sys.stderr)


def _add_output(parser: argparse.ArgumentParser) -> None:
    add_output_args(parser)


def _add_runtime(parser: argparse.ArgumentParser, *, transfer: bool = False) -> None:
    parser.add_argument("--timeout", type=float, default=DEFAULT_TIMEOUT)
    parser.add_argument(
        "--minimum-interval",
        type=float,
        default=DEFAULT_MINIMUM_INTERVAL,
    )
    parser.add_argument(
        "--retry-attempts", type=int, default=DEFAULT_RETRY_ATTEMPTS
    )
    parser.add_argument("--chunk-size", type=int, default=1024 * 1024)
    if transfer:
        parser.add_argument(
            "--max-download-bytes",
            type=int,
            help="Optional caller-selected maximum artifact size",
        )
    _add_output(parser)


def _add_listing_window(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--limit", type=int)
    parser.add_argument("--cursor")


def _add_local_schema(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--member")
    parser.add_argument("--sheet")
    parser.add_argument("--header-scan-rows", type=int, default=25)
    parser.add_argument("--encoding", default="utf-8-sig")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Query Franklin County Auditor official bulk releases"
    )
    sub = parser.add_subparsers(dest="command", required=True)

    source = sub.add_parser("source", help="Describe the source contract")
    _add_output(source)

    families = sub.add_parser("families", help="List official file families")
    _add_output(families)

    releases = sub.add_parser("releases", help="Discover current or archived releases")
    releases.add_argument("family", choices=FAMILY_CHOICES)
    selection = releases.add_mutually_exclusive_group()
    selection.add_argument("--year", type=int)
    selection.add_argument("--all-releases", action="store_true")
    _add_listing_window(releases)
    _add_runtime(releases)

    artifacts = sub.add_parser("artifacts", help="List artifacts in one release")
    artifacts.add_argument("family", choices=FAMILY_CHOICES)
    artifacts.add_argument("--release", default="current")
    _add_listing_window(artifacts)
    _add_runtime(artifacts)

    artifact_probe = sub.add_parser(
        "artifact-probe",
        help="Probe one listed artifact without downloading it",
    )
    artifact_probe.add_argument("family", choices=FAMILY_CHOICES)
    artifact_probe.add_argument("artifact")
    artifact_probe.add_argument("--release", default="current")
    artifact_probe.add_argument("--sample-bytes", type=int, default=DEFAULT_SAMPLE_BYTES)
    _add_runtime(artifact_probe)

    download = sub.add_parser(
        "download",
        help="Download one listed artifact with resumable verification",
    )
    download.add_argument("family", choices=FAMILY_CHOICES)
    download.add_argument("artifact")
    download.add_argument("--release", default="current")
    download.add_argument("--destination", type=Path, required=True)
    download.add_argument("--expected-sha256")
    download.add_argument("--overwrite", action="store_true")
    download.add_argument("--no-resume", action="store_false", dest="resume")
    download.set_defaults(resume=True)
    _add_runtime(download, transfer=True)

    inspect_parser = sub.add_parser(
        "inspect-local",
        help="Inspect headers or ZIP membership of an explicit local artifact",
    )
    inspect_parser.add_argument("artifact", type=Path)
    inspect_parser.add_argument("--record-family", choices=RECORD_FAMILY_CHOICES)
    inspect_parser.add_argument("--sample-rows", type=int, default=3)
    _add_local_schema(inspect_parser)
    _add_output(inspect_parser)

    rows = sub.add_parser(
        "rows",
        help="Stream and optionally filter one explicit local artifact",
    )
    rows.add_argument("artifact", type=Path)
    rows.add_argument("--record-family", required=True, choices=RECORD_FAMILY_CHOICES)
    rows.add_argument("--release-id", required=True)
    rows.add_argument("--release-date")
    rows.add_argument("--source-url")
    rows.add_argument("--query")
    rows.add_argument("--parcel")
    rows.add_argument("--from-date")
    rows.add_argument("--to-date")
    rows.add_argument("--limit", type=int)
    rows.add_argument("--cursor")
    _add_local_schema(rows)
    _add_output(rows)

    probe = sub.add_parser(
        "probe",
        help="Run bounded current-release and small-artifact probes",
    )
    probe.add_argument("--sample-bytes", type=int, default=DEFAULT_SAMPLE_BYTES)
    _add_runtime(probe)

    return parser


def _validate_args(parser: argparse.ArgumentParser, args: argparse.Namespace) -> None:
    for name in ("timeout", "retry_attempts", "chunk_size"):
        if hasattr(args, name) and getattr(args, name) <= 0:
            parser.error(f"--{name.replace('_', '-')} must be positive")
    if hasattr(args, "minimum_interval") and args.minimum_interval < 0:
        parser.error("--minimum-interval must not be negative")
    for name in (
        "limit",
        "sample_bytes",
        "max_download_bytes",
        "header_scan_rows",
    ):
        if hasattr(args, name):
            value = getattr(args, name)
            if value is not None and value <= 0:
                parser.error(f"--{name.replace('_', '-')} must be positive")
    if hasattr(args, "sample_rows") and args.sample_rows < 0:
        parser.error("--sample-rows must not be negative")
    if hasattr(args, "year") and args.year is not None:
        if args.year < 1900 or args.year > 9999:
            parser.error("--year must be a four-digit year")
    if hasattr(args, "release_date") and args.release_date:
        try:
            args.release_date = date.fromisoformat(args.release_date).isoformat()
        except ValueError:
            parser.error("--release-date must be YYYY-MM-DD")


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    _validate_args(parser, args)
    result = execute(args)
    _emit(result, args)
    return (
        0
        if result.status
        in {ResultStatus.OK, ResultStatus.NO_RESULTS, ResultStatus.PARTIAL}
        else 1
    )


if __name__ == "__main__":
    raise SystemExit(main())
