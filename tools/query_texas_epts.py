#!/usr/bin/env python3
"""Plan acquisition and locally search Texas Comptroller EPTS records.

The Texas Comptroller's Property Tax Assistance Division (PTAD) receives
Electronic Property Transaction Submission (EPTS) files from appraisal
districts. The agency describes the holdings and their 52-field layout, but
does not publish a statewide transaction download on its data-submission page.
This adapter therefore keeps acquisition and local parsing separate:

* ``request-plan`` prepares a structured CRRS handoff and does not submit it.
* ``discover`` and ``schema`` expose the verified official source contract.
* ``inspect`` validates a delivered artifact against the September 2025 layout.
* ``parse`` and ``search`` stream delivered CSV, tab-delimited, XLSX, or ZIP
  members while preserving every source row occurrence.

EPTS rows are appraisal-district transaction reports. Deed identifiers are
county-clerk search pivots; they are not recorded-instrument copies, and the
rows are not title determinations.

Examples:
    uv run python tools/query_texas_epts.py request-plan --json
    uv run python tools/query_texas_epts.py schema --output /tmp/epts-schema.json
    uv run python tools/query_texas_epts.py inspect /tmp/epts.zip --json
    uv run python tools/query_texas_epts.py search /tmp/epts.csv "ACME LLC" \
        --field party --limit 100 --output /tmp/epts-acme.json
"""

from __future__ import annotations

import argparse
import base64
import binascii
import csv
import hashlib
import io
import json
import re
import shutil
import sys
import tempfile
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO, Iterator, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

try:
    from tools.output_util import add_output_args, write_output
    from tools.public_records_contract import (
        JurisdictionMetadata,
        PublicRecordsError,
        PublicRecordsQuery,
        PublicRecordsResult,
        QueryMetadata,
        ResultStatus,
        SourceMetadata,
        canonical_json,
        sha256_fingerprint,
    )
except ImportError:
    from output_util import add_output_args, write_output
    from public_records_contract import (
        JurisdictionMetadata,
        PublicRecordsError,
        PublicRecordsQuery,
        PublicRecordsResult,
        QueryMetadata,
        ResultStatus,
        SourceMetadata,
        canonical_json,
        sha256_fingerprint,
    )


SOURCE_ID = "us-tx-comptroller-epts"
STATE_FIPS = "48"
STATE_CODE = "TX"
MANUAL_PUBLICATION = "September 2025"
LAYOUT_VERSION = "texas-epts-2025-09"
CURSOR_PREFIX = "tx-epts:v1:"
CURSOR_VERSION = 1
DEFAULT_LIMIT = 100
CSV_FIELD_SIZE_LIMIT = 1_000_000

LANDING_URL = (
    "https://comptroller.texas.gov/taxes/property-tax/"
    "data-submissions.php"
)
MANUAL_URL = (
    "https://comptroller.texas.gov/taxes/property-tax/docs/96-1208.pdf"
)
PUBLIC_INFORMATION_URL = (
    "https://comptroller.texas.gov/about/policies/open-records/"
    "public-information-act.php"
)
CRRS_URL = "https://crrs.cpa.texas.gov/"
OPEN_RECORDS_EMAIL = "open.records@cpa.texas.gov"


@dataclass(frozen=True)
class FieldSpec:
    number: int
    header: str
    name: str
    field_type: str
    length: int
    decimal_places: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "number": self.number,
            "header": self.header,
            "name": self.name,
            "field_type": self.field_type,
            "length": self.length,
            "decimal_places": self.decimal_places,
        }


EPTS_FIELDS = (
    FieldSpec(1, "CAD_ID", "CAD Code", "numeric", 3, 0),
    FieldSpec(2, "TU_ID", "School District Code", "numeric", 6, 0),
    FieldSpec(3, "PROP_CATG_CD", "Category Code", "alphanumeric", 2),
    FieldSpec(4, "PROP_SHT_ID", "Short Account Number", "alphanumeric", 15),
    FieldSpec(5, "PROP_ID1_TX", "Account Number", "alphanumeric", 25),
    FieldSpec(6, "PROP_ID2_TX", "Legal Description", "alphanumeric", 200),
    FieldSpec(7, "PRCL_AD_TX", "Parcel Address", "alphanumeric", 50),
    FieldSpec(8, "SALE_DT", "Sale Date", "date", 10, 0),
    FieldSpec(9, "PROP_SALE_AM", "Sale Price", "numeric", 12, 0),
    FieldSpec(10, "DEED_DT", "Deed Date", "date", 10, 0),
    FieldSpec(11, "DEED_VOL_NR", "Deed Volume", "alphanumeric", 5),
    FieldSpec(12, "DEED_PAGE_NR", "Deed Page", "alphanumeric", 7),
    FieldSpec(13, "DEED_NR", "Deed Number", "alphanumeric", 20),
    FieldSpec(14, "DEED_TY_CD", "Deed Type", "alphanumeric", 12),
    FieldSpec(15, "MULT_ACCT_CD", "Multiple Account Code", "alphanumeric", 1),
    FieldSpec(
        16,
        "ADDNL_PROPS",
        "Additional Accounts in Transaction",
        "alphanumeric",
        50,
    ),
    FieldSpec(17, "GNTE_FRST_NM", "Grantee First Name", "alphanumeric", 50),
    FieldSpec(
        18,
        "GNTE_LST_BUS_NM",
        "Grantee Last Name or Business Name",
        "alphanumeric",
        50,
    ),
    FieldSpec(
        19,
        "GNTE_LINE_1_AD_TX",
        "Grantee Address Line 1",
        "alphanumeric",
        35,
    ),
    FieldSpec(
        20,
        "GNTE_LINE_2_AD_TX",
        "Grantee Address Line 2",
        "alphanumeric",
        35,
    ),
    FieldSpec(21, "GNTE_CITY_NM", "Grantee City", "alphanumeric", 24),
    FieldSpec(22, "GNTE_ST_CD", "Grantee State", "alphanumeric", 2),
    FieldSpec(23, "GNTE_AD_ZP", "Grantee ZIP", "numeric", 9, 0),
    FieldSpec(24, "GNTR_FRST_NM", "Grantor First Name", "alphanumeric", 50),
    FieldSpec(
        25,
        "GNTR_LST_BUS_NM",
        "Grantor Last Name or Business Name",
        "alphanumeric",
        50,
    ),
    FieldSpec(
        26,
        "GNTR_LINE_1_AD_TX",
        "Grantor Address Line 1",
        "alphanumeric",
        35,
    ),
    FieldSpec(
        27,
        "GNTR_LINE_2_AD_TX",
        "Grantor Address Line 2",
        "alphanumeric",
        35,
    ),
    FieldSpec(28, "GNTR_CITY_NM", "Grantor City", "alphanumeric", 24),
    FieldSpec(29, "GNTR_ST_CD", "Grantor State", "alphanumeric", 2),
    FieldSpec(30, "GNTR_AD_ZP", "Grantor ZIP", "numeric", 9, 0),
    FieldSpec(31, "CAD_SALE_SRC_CD", "CAD Sale Source Code", "alphanumeric", 3),
    FieldSpec(32, "VALD_CD", "Validity Code", "alphanumeric", 1),
    FieldSpec(33, "CNFD_CD", "Confidential Code", "alphanumeric", 1),
    FieldSpec(
        34,
        "FRZN_CHAR_CD",
        "Frozen Characteristics",
        "alphanumeric",
        1,
    ),
    FieldSpec(35, "CERT_VAL_YR", "Most Recent Appraisal Year", "numeric", 4, 0),
    FieldSpec(36, "ARB_VAL_CD", "Value Set by ARB", "alphanumeric", 1),
    FieldSpec(37, "PROP_RPTD_LAND_AM", "CAD Value - Land", "numeric", 12, 0),
    FieldSpec(
        38,
        "PROP_RPTD_IMPV_AM",
        "CAD Value - Improvement",
        "numeric",
        12,
        0,
    ),
    FieldSpec(
        39,
        "PROP_RPTD_PPROP_AM",
        "CAD Value - Personal Property",
        "numeric",
        12,
        0,
    ),
    FieldSpec(
        40,
        "PROP_RPTD_TOTL_AM",
        "Total CAD Value",
        "numeric",
        12,
        0,
    ),
    FieldSpec(41, "PCT_OWNSHP", "Percent Ownership", "numeric", 7, 6),
    FieldSpec(
        42,
        "PCT_COMP",
        "Current Year Percent Complete",
        "numeric",
        5,
        2,
    ),
    FieldSpec(
        43,
        "SQFT_IMPV_QY",
        "Square Footage - Improvement",
        "numeric",
        7,
        0,
    ),
    FieldSpec(44, "BUILT_YR", "Year Built", "numeric", 4, 0),
    FieldSpec(45, "LAND_UNIT_TY_CD", "Land Unit Type", "alphanumeric", 2),
    FieldSpec(46, "LAND_UNIT_QY", "Number of Land Units", "numeric", 14, 4),
    FieldSpec(47, "FNC_CD", "Financing Code", "alphanumeric", 5),
    FieldSpec(48, "DY_ON_MRKT_QY", "Number of Days on Market", "numeric", 3, 0),
    FieldSpec(
        49,
        "PREV_RPTD_LAND_AM",
        "Previous CAD Value - Land",
        "numeric",
        12,
        0,
    ),
    FieldSpec(
        50,
        "PREV_RPTD_IMPV_AM",
        "Previous CAD Value - Improvement",
        "numeric",
        12,
        0,
    ),
    FieldSpec(51, "CAD_LINE_1_CMNT_TX", "CAD Comments Line 1", "alphanumeric", 50),
    FieldSpec(52, "CAD_LINE_2_CMNT_TX", "CAD Comments Line 2", "alphanumeric", 50),
)

EXPECTED_HEADERS = tuple(field.header for field in EPTS_FIELDS)
FIELD_BY_HEADER = {field.header: field for field in EPTS_FIELDS}
FIELD_BY_NUMBER = {field.number: field for field in EPTS_FIELDS}
SCHEMA_FINGERPRINT = sha256_fingerprint(
    {
        "layout_version": LAYOUT_VERSION,
        "manual_publication": MANUAL_PUBLICATION,
        "fields": [field.to_dict() for field in EPTS_FIELDS],
    }
)

SALE_SOURCE_CODES = {
    "APP": "local appraiser",
    "BUY": "property buyer",
    "SEL": "property seller",
    "AGT": "real estate or tax agent",
    "SER": "subscription service data source",
    "INT": "internet source",
}
VALIDITY_CODES = {
    "Y": "appraisal district considers sale valid and arm's-length",
    "N": "appraisal district does not consider sale valid and arm's-length",
    "U": "appraisal district reports validity unknown",
}
CONFIDENTIAL_FIELD_NUMBERS = {
    "A": (7,),
    "B": (9,),
    "C": (19, 20),
    "D": (26, 27),
    "E": (7, 9),
    "F": (7, 19, 20),
    "G": (7, 26, 27),
    "H": (9, 19, 20),
    "I": (9, 26, 27),
    "J": (19, 20, 26, 27),
    "K": (7, 9, 19, 20),
    "L": (7, 9, 26, 27),
    "M": (7, 19, 20, 26, 27),
    "N": (9, 19, 20, 26, 27),
    "O": (7, 9, 19, 20, 26, 27),
    "P": (),
    "Q": (),
}
CONFIDENTIAL_CODE_DESCRIPTIONS = {
    "A": "parcel address",
    "B": "sale price",
    "C": "grantee address",
    "D": "grantor address",
    "E": "parcel address and sale price",
    "F": "parcel address and grantee address",
    "G": "parcel address and grantor address",
    "H": "sale price and grantee address",
    "I": "sale price and grantor address",
    "J": "grantee address and grantor address",
    "K": "parcel address, sale price and grantee address",
    "L": "parcel address, sale price and grantor address",
    "M": "parcel address, grantee address and grantor address",
    "N": "sale price, grantee address and grantor address",
    "O": "parcel address, sale price, grantee address and grantor address",
    "P": "other confidential information",
    "Q": "no confidential information",
}
SENSITIVE_FIELD_NUMBERS = frozenset({7, 9, 19, 20, 26, 27})
REDACTION_MARKERS = frozenset(
    {
        "REDACTED",
        "[REDACTED]",
        "WITHHELD",
        "[WITHHELD]",
        "CONFIDENTIAL",
        "[CONFIDENTIAL]",
        "***",
    }
)

SEARCH_FIELDS = {
    "account": (
        "CAD_ID",
        "PROP_SHT_ID",
        "PROP_ID1_TX",
        "ADDNL_PROPS",
    ),
    "party": (
        "GNTE_FRST_NM",
        "GNTE_LST_BUS_NM",
        "GNTR_FRST_NM",
        "GNTR_LST_BUS_NM",
    ),
    "address": (
        "PRCL_AD_TX",
        "GNTE_LINE_1_AD_TX",
        "GNTE_LINE_2_AD_TX",
        "GNTE_CITY_NM",
        "GNTE_ST_CD",
        "GNTE_AD_ZP",
        "GNTR_LINE_1_AD_TX",
        "GNTR_LINE_2_AD_TX",
        "GNTR_CITY_NM",
        "GNTR_ST_CD",
        "GNTR_AD_ZP",
    ),
    "deed": (
        "DEED_DT",
        "DEED_VOL_NR",
        "DEED_PAGE_NR",
        "DEED_NR",
        "DEED_TY_CD",
    ),
    "legal": ("PROP_ID2_TX",),
    "comment": ("CAD_LINE_1_CMNT_TX", "CAD_LINE_2_CMNT_TX"),
    "cad": ("CAD_ID", "TU_ID", "CAD_SALE_SRC_CD", "VALD_CD"),
    "transaction": (
        "SALE_DT",
        "PROP_SALE_AM",
        "DEED_DT",
        "DEED_VOL_NR",
        "DEED_PAGE_NR",
        "DEED_NR",
        "DEED_TY_CD",
        "FNC_CD",
    ),
}

SOURCE_METADATA = SourceMetadata(
    source_id=SOURCE_ID,
    name="Texas Comptroller Electronic Property Transaction Submissions",
    source_role="statewide_appraisal_district_property_transaction_reporting",
    base_url=LANDING_URL,
    dataset_id="EPTS",
    metadata={
        "publisher": "Texas Comptroller of Public Accounts",
        "division": "Property Tax Assistance Division",
        "record_grain": "appraisal_district_submission_row_occurrence",
        "acquisition": "public_information_request_then_local_artifact",
        "manual_url": MANUAL_URL,
        "manual_publication": MANUAL_PUBLICATION,
        "field_count": 52,
        "recorded_instrument_copy": False,
        "title_determination": False,
    },
)
JURISDICTION = JurisdictionMetadata(
    jurisdiction_id=STATE_FIPS,
    name="Texas",
    state_code=STATE_CODE,
)
SOURCE_WARNINGS = (
    "EPTS records are appraisal-district transaction reports; deed locators "
    "are county-clerk search pivots, not recorded-instrument copies.",
    "A grantor or grantee role reports the submitted transaction role and is "
    "not a current-title or beneficial-ownership determination.",
    "Field 33 confidentiality codes use the September 2025 A-Q mapping; "
    "published blanks and redaction markers remain explicit source states.",
)


class EPTSError(ValueError):
    """Base error for EPTS planning and local artifact processing."""


class EPTSLayoutError(EPTSError):
    """A delivered artifact does not match the official 52-field layout."""


class EPTSCursorError(EPTSError):
    """A cursor does not belong to the current artifact and query."""


@dataclass(frozen=True)
class ArtifactIdentity:
    path: Path
    sha256: str
    byte_length: int
    container_format: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "filename": self.path.name,
            "sha256": self.sha256,
            "byte_length": self.byte_length,
            "container_format": self.container_format,
        }


@dataclass(frozen=True)
class MemberSpec:
    member_id: str
    source_name: str
    file_format: str
    header_row_number: int
    delimiter: str | None = None
    encoding: str | None = None
    archive_member: str | None = None
    archive_member_index: int | None = None
    sheet_name: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "member_id": self.member_id,
            "source_name": self.source_name,
            "file_format": self.file_format,
            "header_row_number": self.header_row_number,
            "delimiter": self.delimiter,
            "encoding": self.encoding,
            "archive_member": self.archive_member,
            "archive_member_index": self.archive_member_index,
            "sheet_name": self.sheet_name,
            "header": list(EXPECTED_HEADERS),
            "header_count": len(EXPECTED_HEADERS),
            "schema_fingerprint": SCHEMA_FINGERPRINT,
        }


@dataclass(frozen=True)
class RowOccurrence:
    member: MemberSpec
    source_row_number: int
    logical_data_row: int
    raw_values: tuple[Any, ...]
    physical_line_start: int | None = None
    physical_line_end: int | None = None
    cell_metadata: Mapping[str, Mapping[str, Any]] | None = None


def source_discovery_record() -> dict[str, Any]:
    """Return the verified official source and acquisition contract."""

    return {
        "canonical_ref": f"TX-EPTS:{STATE_FIPS}/source-discovery",
        "source_id": SOURCE_ID,
        "record_kind": "property_transaction_source_discovery",
        "publisher": "Texas Comptroller of Public Accounts",
        "division": "Property Tax Assistance Division",
        "official_sources": {
            "data_submission_landing": LANDING_URL,
            "current_manual": MANUAL_URL,
            "public_information_instructions": PUBLIC_INFORMATION_URL,
            "request_portal": CRRS_URL,
        },
        "holdings_description": {
            "program": "Electronic Property Transaction Submission",
            "reported_by": "Texas appraisal districts",
            "record_grain": "one appraisal-district account row per transfer",
            "statewide_compilation_described": True,
            "public_statewide_download_found": False,
        },
        "acquisition": {
            "mode": "public_information_request_then_local_artifact",
            "request_plan_command": (
                "uv run python tools/query_texas_epts.py request-plan --json"
            ),
            "submission_performed": False,
        },
        "formats": {
            "manual_submission_formats": [
                "tab-delimited text",
                "comma-delimited text",
            ],
            "local_parser_formats": [
                "tab-delimited text",
                "comma-delimited text",
                "XLSX workbook",
                "ZIP containing supported members",
            ],
            "legacy_xls_parser_available": False,
        },
        "schema": {
            "layout_version": LAYOUT_VERSION,
            "manual_publication": MANUAL_PUBLICATION,
            "field_count": len(EPTS_FIELDS),
            "schema_fingerprint": SCHEMA_FINGERPRINT,
        },
        "specimen_state": {
            "delivered_epts_artifact_reviewed": False,
            "contract_basis": "official September 2025 manual",
            "first_delivery_requires_local_inspection": True,
        },
        "source_semantics": {
            "recorded_instrument_copy": False,
            "title_determination": False,
            "deed_locator_role": "county_clerk_search_pivot",
        },
    }


def schema_record() -> dict[str, Any]:
    """Return the official September 2025 EPTS layout as structured data."""

    return {
        "canonical_ref": f"TX-EPTS:{STATE_FIPS}/schema/{LAYOUT_VERSION}",
        "source_id": SOURCE_ID,
        "record_kind": "property_transaction_schema",
        "source_url": MANUAL_URL,
        "layout_version": LAYOUT_VERSION,
        "manual_publication": MANUAL_PUBLICATION,
        "field_count": len(EPTS_FIELDS),
        "manual_declared_total_length": 1059,
        "schema_fingerprint": SCHEMA_FINGERPRINT,
        "fields": [field.to_dict() for field in EPTS_FIELDS],
        "code_sets": {
            "cad_sale_source": dict(SALE_SOURCE_CODES),
            "validity": dict(VALIDITY_CODES),
            "confidentiality": {
                code: {
                    "description": CONFIDENTIAL_CODE_DESCRIPTIONS[code],
                    "field_numbers": list(CONFIDENTIAL_FIELD_NUMBERS[code]),
                    "headers": [
                        FIELD_BY_NUMBER[number].header
                        for number in CONFIDENTIAL_FIELD_NUMBERS[code]
                    ],
                }
                for code in CONFIDENTIAL_FIELD_NUMBERS
            },
        },
        "identity_contract": {
            "source_occurrence": "artifact SHA-256 + member + row",
            "property_candidate": "CAD_ID + PROP_ID1_TX",
            "transaction_group_candidate": (
                "CAD and deed locator/date/party components"
            ),
            "automatic_deduplication": False,
        },
    }


def _iso_date(value: str | None, field_name: str) -> str | None:
    if value is None:
        return None
    candidate = value.strip()
    if not candidate:
        return None
    try:
        return date.fromisoformat(candidate).isoformat()
    except ValueError as error:
        raise EPTSError(f"{field_name} must be an ISO date") from error


def _cad_ids(values: Sequence[str]) -> list[str]:
    normalized: list[str] = []
    for value in values:
        candidate = value.strip()
        if not re.fullmatch(r"\d{3}", candidate):
            raise EPTSError("CAD IDs must be three digits")
        if candidate not in normalized:
            normalized.append(candidate)
    return normalized


def request_plan_record(
    *,
    cad_ids: Sequence[str] = (),
    start_date: str | None = None,
    end_date: str | None = None,
    preferred_format: str = "native-machine-readable",
) -> dict[str, Any]:
    """Build a request handoff without contacting or submitting to CRRS."""

    normalized_cads = _cad_ids(cad_ids)
    normalized_start = _iso_date(start_date, "start date")
    normalized_end = _iso_date(end_date, "end date")
    if (
        normalized_start is not None
        and normalized_end is not None
        and normalized_start > normalized_end
    ):
        raise EPTSError("start date must not follow end date")

    if normalized_cads:
        cad_scope = "CAD IDs " + ", ".join(normalized_cads)
    else:
        cad_scope = "all appraisal districts represented in the requested files"
    if normalized_start and normalized_end:
        date_scope = f"records dated {normalized_start} through {normalized_end}"
    elif normalized_start:
        date_scope = f"records dated on or after {normalized_start}"
    elif normalized_end:
        date_scope = f"records dated on or before {normalized_end}"
    else:
        date_scope = "all available EPTS submission periods"

    request_text = (
        "Please provide the Electronic Property Transaction Submission "
        "(EPTS) data files held by the Property Tax Assistance Division for "
        f"{date_scope} and {cad_scope}. Please include all 52 columns in the "
        f"{MANUAL_PUBLICATION} EPTS layout, including Field 33 (CNFD_CD), "
        "and retain source filenames or other submission identifiers. "
        "Machine-readable native files, CSV, tab-delimited text, or XLSX are "
        "preferred. If a field is withheld, please preserve the row and column "
        "structure and release the remaining fields."
    )
    scope = {
        "cad_ids": normalized_cads,
        "cad_scope": cad_scope,
        "start_date": normalized_start,
        "end_date": normalized_end,
        "date_scope": date_scope,
        "preferred_format": preferred_format,
        "requested_headers": list(EXPECTED_HEADERS),
    }
    plan_fingerprint = sha256_fingerprint(
        {
            "source_id": SOURCE_ID,
            "scope": scope,
            "request_text": request_text,
        }
    )
    return {
        "canonical_ref": f"TX-EPTS:{STATE_FIPS}/request-plan/{plan_fingerprint}",
        "source_id": SOURCE_ID,
        "record_kind": "public_information_request_plan_handoff",
        "plan_fingerprint": plan_fingerprint,
        "state": "prepared_for_human_review",
        "submission_performed": False,
        "request_text": request_text,
        "scope": scope,
        "routes": [
            {
                "route_kind": "comptroller_record_request_system",
                "information_url": PUBLIC_INFORMATION_URL,
                "request_url": CRRS_URL,
                "submission_performed": False,
            },
            {
                "route_kind": "open_records_email",
                "information_url": PUBLIC_INFORMATION_URL,
                "email": OPEN_RECORDS_EMAIL,
                "submission_performed": False,
            },
        ],
        "handoff": {
            "next_actor": "human_requestor",
            "review_scope_and_request_text": True,
            "requestor_identity_and_contact_supplied_at_submission": True,
            "receipt_or_tracking_id_expected_after_submission": True,
        },
        "receipt_workflow": {
            "inspect_command": (
                "uv run python tools/query_texas_epts.py inspect ARTIFACT --json"
            ),
            "search_command": (
                "uv run python tools/query_texas_epts.py search ARTIFACT "
                "QUERY --output RESULTS.json"
            ),
        },
    }


def _json_value(value: Any) -> Any:
    """Convert spreadsheet cell values to stable JSON values."""

    if value is None or isinstance(value, (bool, str, int)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        if value != value:
            return "NaN"
        if value in {float("inf"), float("-inf")}:
            return str(value)
        return value
    return str(value)


def _raw_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _artifact_format(path: Path) -> str:
    suffix = path.suffix.casefold()
    if suffix == ".xls":
        raise EPTSLayoutError(
            "legacy XLS is not supported; request or export CSV, tab text, or XLSX"
        )
    if suffix == ".xlsx":
        return "xlsx"
    if suffix == ".zip":
        return "zip"
    if suffix in {".csv", ".txt", ".tsv"}:
        return "delimited"

    with path.open("rb") as stream:
        magic = stream.read(4)
    if magic.startswith(b"PK"):
        try:
            with zipfile.ZipFile(path) as archive:
                names = set(archive.namelist())
        except zipfile.BadZipFile as error:
            raise EPTSLayoutError("artifact has an invalid ZIP container") from error
        if "[Content_Types].xml" in names and "xl/workbook.xml" in names:
            return "xlsx"
        return "zip"
    return "delimited"


def artifact_identity(path: str | Path) -> ArtifactIdentity:
    artifact_path = Path(path).expanduser().resolve()
    if not artifact_path.is_file():
        raise OSError(f"artifact is not a readable file: {artifact_path}")

    digest = hashlib.sha256()
    byte_length = 0
    with artifact_path.open("rb") as stream:
        while True:
            chunk = stream.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
            byte_length += len(chunk)
    return ArtifactIdentity(
        path=artifact_path,
        sha256=digest.hexdigest(),
        byte_length=byte_length,
        container_format=_artifact_format(artifact_path),
    )


def _member_id(
    *,
    archive_member: str | None,
    archive_member_index: int | None,
    sheet_name: str | None = None,
) -> str:
    if archive_member is None:
        base = "artifact"
    else:
        base = f"zip[{archive_member_index}]::{archive_member}"
    if sheet_name is not None:
        return f"{base}#sheet={sheet_name}"
    return base


@contextmanager
def _open_member_binary(
    artifact_path: Path,
    *,
    archive_member: str | None,
    archive_member_index: int | None,
) -> Iterator[BinaryIO]:
    if archive_member is None:
        with artifact_path.open("rb") as stream:
            yield stream
        return

    if archive_member_index is None:
        raise EPTSLayoutError("archive member index is missing")
    try:
        with zipfile.ZipFile(artifact_path) as archive:
            members = archive.infolist()
            if archive_member_index >= len(members):
                raise EPTSLayoutError("archive member index is no longer present")
            info = members[archive_member_index]
            if info.filename != archive_member:
                raise EPTSLayoutError("archive member identity changed")
            with archive.open(info) as stream:
                yield stream
    except (zipfile.BadZipFile, RuntimeError) as error:
        raise EPTSLayoutError(
            f"cannot read archive member {archive_member!r}: {error}"
        ) from error


@contextmanager
def _open_member_workbook(
    artifact_path: Path,
    *,
    archive_member: str | None,
    archive_member_index: int | None,
) -> Iterator[Any]:
    workbook = None
    spool = None
    try:
        if archive_member is None:
            workbook = load_workbook(
                artifact_path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        else:
            spool = tempfile.SpooledTemporaryFile(max_size=16 * 1024 * 1024)
            with _open_member_binary(
                artifact_path,
                archive_member=archive_member,
                archive_member_index=archive_member_index,
            ) as stream:
                shutil.copyfileobj(stream, spool, length=1024 * 1024)
            spool.seek(0)
            workbook = load_workbook(
                spool,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        yield workbook
    except (InvalidFileException, zipfile.BadZipFile, KeyError) as error:
        label = archive_member or artifact_path.name
        raise EPTSLayoutError(f"cannot read XLSX member {label!r}: {error}") from error
    finally:
        if workbook is not None:
            workbook.close()
        if spool is not None:
            spool.close()


def _read_delimited_header(
    artifact_path: Path,
    *,
    archive_member: str | None,
    archive_member_index: int | None,
) -> tuple[list[str], str, str]:
    observations: list[tuple[list[str], str, str]] = []
    decode_errors: list[str] = []
    for encoding in ("utf-8-sig", "cp1252"):
        for delimiter in (",", "\t"):
            try:
                with _open_member_binary(
                    artifact_path,
                    archive_member=archive_member,
                    archive_member_index=archive_member_index,
                ) as binary:
                    text = io.TextIOWrapper(binary, encoding=encoding, newline="")
                    header = next(csv.reader(text, delimiter=delimiter), [])
            except UnicodeDecodeError as error:
                decode_errors.append(f"{encoding}: {error}")
                break
            observations.append((header, delimiter, encoding))
            if tuple(header) == EXPECTED_HEADERS:
                return header, delimiter, encoding

    if observations:
        header, delimiter, encoding = max(
            observations,
            key=lambda item: (
                sum(value in FIELD_BY_HEADER for value in item[0]),
                len(item[0]),
            ),
        )
        return header, delimiter, encoding
    raise EPTSLayoutError(
        "could not decode delimited header: " + "; ".join(decode_errors)
    )


def _header_rejection(
    *,
    member_id: str,
    source_name: str,
    file_format: str,
    observed_header: Sequence[Any],
    archive_member: str | None = None,
    archive_member_index: int | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    header = [_json_value(value) for value in observed_header]
    mismatch_at = None
    for index in range(max(len(header), len(EXPECTED_HEADERS))):
        observed = header[index] if index < len(header) else None
        expected = EXPECTED_HEADERS[index] if index < len(EXPECTED_HEADERS) else None
        if observed != expected:
            mismatch_at = {
                "field_number": index + 1,
                "expected": expected,
                "observed": observed,
            }
            break
    return {
        "member_id": member_id,
        "source_name": source_name,
        "file_format": file_format,
        "archive_member": archive_member,
        "archive_member_index": archive_member_index,
        "sheet_name": sheet_name,
        "reason": "header_does_not_match_official_52_field_layout",
        "observed_header_count": len(header),
        "observed_header": header,
        "first_mismatch": mismatch_at,
    }


def _discover_delimited_member(
    artifact_path: Path,
    *,
    source_name: str,
    archive_member: str | None,
    archive_member_index: int | None,
) -> tuple[MemberSpec | None, dict[str, Any] | None]:
    member_id = _member_id(
        archive_member=archive_member,
        archive_member_index=archive_member_index,
    )
    try:
        header, delimiter, encoding = _read_delimited_header(
            artifact_path,
            archive_member=archive_member,
            archive_member_index=archive_member_index,
        )
    except EPTSError as error:
        return None, {
            "member_id": member_id,
            "source_name": source_name,
            "file_format": "delimited",
            "archive_member": archive_member,
            "archive_member_index": archive_member_index,
            "reason": "member_unreadable",
            "detail": str(error),
        }
    if tuple(header) != EXPECTED_HEADERS:
        return None, _header_rejection(
            member_id=member_id,
            source_name=source_name,
            file_format="delimited",
            observed_header=header,
            archive_member=archive_member,
            archive_member_index=archive_member_index,
        )
    return (
        MemberSpec(
            member_id=member_id,
            source_name=source_name,
            file_format="delimited",
            header_row_number=1,
            delimiter=delimiter,
            encoding=encoding,
            archive_member=archive_member,
            archive_member_index=archive_member_index,
        ),
        None,
    )


def _trim_trailing_blank(values: Sequence[Any]) -> list[Any]:
    trimmed = list(values)
    while trimmed and _raw_text(trimmed[-1]).strip() == "":
        trimmed.pop()
    return trimmed


def _discover_xlsx_members(
    artifact_path: Path,
    *,
    source_name: str,
    archive_member: str | None,
    archive_member_index: int | None,
) -> tuple[list[MemberSpec], list[dict[str, Any]]]:
    members: list[MemberSpec] = []
    rejected: list[dict[str, Any]] = []
    with _open_member_workbook(
        artifact_path,
        archive_member=archive_member,
        archive_member_index=archive_member_index,
    ) as workbook:
        for worksheet in workbook.worksheets:
            member_id = _member_id(
                archive_member=archive_member,
                archive_member_index=archive_member_index,
                sheet_name=worksheet.title,
            )
            first_row = next(worksheet.iter_rows(values_only=True), ())
            header = _trim_trailing_blank(first_row)
            if tuple(header) == EXPECTED_HEADERS:
                members.append(
                    MemberSpec(
                        member_id=member_id,
                        source_name=source_name,
                        file_format="xlsx",
                        header_row_number=1,
                        archive_member=archive_member,
                        archive_member_index=archive_member_index,
                        sheet_name=worksheet.title,
                    )
                )
            else:
                rejected.append(
                    _header_rejection(
                        member_id=member_id,
                        source_name=source_name,
                        file_format="xlsx",
                        observed_header=header,
                        archive_member=archive_member,
                        archive_member_index=archive_member_index,
                        sheet_name=worksheet.title,
                    )
                )
    return members, rejected


def discover_artifact_members(
    path: str | Path,
) -> tuple[ArtifactIdentity, list[MemberSpec], list[dict[str, Any]]]:
    """Discover valid EPTS members without accepting a partial schema match."""

    identity = artifact_identity(path)
    members: list[MemberSpec] = []
    rejected: list[dict[str, Any]] = []

    if identity.container_format == "delimited":
        member, rejection = _discover_delimited_member(
            identity.path,
            source_name=identity.path.name,
            archive_member=None,
            archive_member_index=None,
        )
        if member is not None:
            members.append(member)
        if rejection is not None:
            rejected.append(rejection)
    elif identity.container_format == "xlsx":
        found, failures = _discover_xlsx_members(
            identity.path,
            source_name=identity.path.name,
            archive_member=None,
            archive_member_index=None,
        )
        members.extend(found)
        rejected.extend(failures)
    else:
        try:
            with zipfile.ZipFile(identity.path) as archive:
                infos = archive.infolist()
        except zipfile.BadZipFile as error:
            raise EPTSLayoutError("artifact has an invalid ZIP container") from error
        for index, info in enumerate(infos):
            if info.is_dir() or info.filename.startswith("__MACOSX/"):
                continue
            suffix = Path(info.filename).suffix.casefold()
            if suffix in {".csv", ".txt", ".tsv", ""}:
                member, rejection = _discover_delimited_member(
                    identity.path,
                    source_name=info.filename,
                    archive_member=info.filename,
                    archive_member_index=index,
                )
                if member is not None:
                    members.append(member)
                if rejection is not None:
                    rejected.append(rejection)
            elif suffix == ".xlsx":
                try:
                    found, failures = _discover_xlsx_members(
                        identity.path,
                        source_name=info.filename,
                        archive_member=info.filename,
                        archive_member_index=index,
                    )
                except EPTSError as error:
                    rejected.append(
                        {
                            "member_id": _member_id(
                                archive_member=info.filename,
                                archive_member_index=index,
                            ),
                            "source_name": info.filename,
                            "file_format": "xlsx",
                            "archive_member": info.filename,
                            "archive_member_index": index,
                            "reason": "member_unreadable",
                            "detail": str(error),
                        }
                    )
                else:
                    members.extend(found)
                    rejected.extend(failures)
            elif suffix == ".xls":
                rejected.append(
                    {
                        "member_id": _member_id(
                            archive_member=info.filename,
                            archive_member_index=index,
                        ),
                        "source_name": info.filename,
                        "file_format": "xls",
                        "archive_member": info.filename,
                        "archive_member_index": index,
                        "reason": "legacy_xls_not_supported",
                    }
                )
    return identity, members, rejected


def _select_members(
    members: Sequence[MemberSpec],
    selector: str | None,
) -> list[MemberSpec]:
    if not members:
        raise EPTSLayoutError(
            "artifact contains no member with the official 52-field EPTS header"
        )
    if selector is None:
        return list(members)
    candidate = selector.strip()
    selected = [
        member
        for member in members
        if candidate
        in {
            member.member_id,
            member.source_name,
            member.archive_member,
            member.sheet_name,
        }
    ]
    if not selected:
        available = ", ".join(member.member_id for member in members)
        raise EPTSLayoutError(
            f"member {candidate!r} was not found; available members: {available}"
        )
    if len(selected) > 1:
        matches = ", ".join(member.member_id for member in selected)
        raise EPTSLayoutError(
            f"member selector {candidate!r} is ambiguous; matches: {matches}"
        )
    return selected


def _iter_delimited_rows(
    artifact_path: Path,
    member: MemberSpec,
) -> Iterator[RowOccurrence]:
    if member.delimiter is None or member.encoding is None:
        raise EPTSLayoutError("delimited member metadata is incomplete")
    csv.field_size_limit(CSV_FIELD_SIZE_LIMIT)
    with _open_member_binary(
        artifact_path,
        archive_member=member.archive_member,
        archive_member_index=member.archive_member_index,
    ) as binary:
        text = io.TextIOWrapper(binary, encoding=member.encoding, newline="")
        reader = csv.reader(text, delimiter=member.delimiter)
        header = next(reader, [])
        if tuple(header) != EXPECTED_HEADERS:
            raise EPTSLayoutError(
                f"{member.member_id}: header changed since member discovery"
            )
        previous_line = reader.line_num
        for logical_data_row, values in enumerate(reader, start=1):
            physical_start = previous_line + 1
            physical_end = reader.line_num
            previous_line = physical_end
            if not values or all(value == "" for value in values):
                continue
            if len(values) != len(EPTS_FIELDS):
                raise EPTSLayoutError(
                    f"{member.member_id}: data row {logical_data_row} has "
                    f"{len(values)} fields; expected {len(EPTS_FIELDS)}"
                )
            yield RowOccurrence(
                member=member,
                source_row_number=physical_start,
                logical_data_row=logical_data_row,
                raw_values=tuple(values),
                physical_line_start=physical_start,
                physical_line_end=physical_end,
            )


def _iter_xlsx_rows(
    artifact_path: Path,
    member: MemberSpec,
) -> Iterator[RowOccurrence]:
    if member.sheet_name is None:
        raise EPTSLayoutError("XLSX member sheet name is missing")
    with _open_member_workbook(
        artifact_path,
        archive_member=member.archive_member,
        archive_member_index=member.archive_member_index,
    ) as workbook:
        if member.sheet_name not in workbook.sheetnames:
            raise EPTSLayoutError(
                f"{member.member_id}: worksheet changed since member discovery"
            )
        worksheet = workbook[member.sheet_name]
        rows = worksheet.iter_rows(values_only=False)
        header_cells = next(rows, ())
        header = _trim_trailing_blank([cell.value for cell in header_cells])
        if tuple(header) != EXPECTED_HEADERS:
            raise EPTSLayoutError(
                f"{member.member_id}: header changed since member discovery"
            )
        for logical_data_row, cells in enumerate(rows, start=1):
            values = [cell.value for cell in cells]
            if len(values) > len(EPTS_FIELDS):
                extras = values[len(EPTS_FIELDS) :]
                if any(_raw_text(value).strip() for value in extras):
                    raise EPTSLayoutError(
                        f"{member.member_id}: worksheet row "
                        f"{logical_data_row + 1} has nonblank data beyond "
                        f"the official {len(EPTS_FIELDS)} fields"
                    )
                values = values[: len(EPTS_FIELDS)]
                cells = cells[: len(EPTS_FIELDS)]
            if len(values) < len(EPTS_FIELDS):
                values.extend([None] * (len(EPTS_FIELDS) - len(values)))
            if all(_raw_text(value).strip() == "" for value in values):
                continue
            source_row_number = logical_data_row + 1
            cell_metadata: dict[str, dict[str, Any]] = {}
            for index, cell in enumerate(cells[: len(EPTS_FIELDS)]):
                if _raw_text(cell.value).strip() == "":
                    continue
                cell_metadata[EXPECTED_HEADERS[index]] = {
                    "coordinate": getattr(cell, "coordinate", None),
                    "data_type": getattr(cell, "data_type", None),
                    "number_format": getattr(cell, "number_format", None),
                }
            yield RowOccurrence(
                member=member,
                source_row_number=source_row_number,
                logical_data_row=logical_data_row,
                raw_values=tuple(values),
                cell_metadata=cell_metadata,
            )


def _iter_selected_rows(
    artifact_path: Path,
    members: Sequence[MemberSpec],
) -> Iterator[RowOccurrence]:
    for member in members:
        if member.file_format == "delimited":
            yield from _iter_delimited_rows(artifact_path, member)
        elif member.file_format == "xlsx":
            yield from _iter_xlsx_rows(artifact_path, member)
        else:
            raise EPTSLayoutError(
                f"unsupported discovered member format: {member.file_format}"
            )


def iter_artifact_rows(
    path: str | Path,
    *,
    member: str | None = None,
) -> Iterator[RowOccurrence]:
    """Yield validated rows in stable artifact/member/source-row order."""

    identity, members, _ = discover_artifact_members(path)
    selected = _select_members(members, member)
    yield from _iter_selected_rows(identity.path, selected)


def _filename_metadata(name: str) -> dict[str, Any] | None:
    match = re.search(
        r"(?<!\d)(?P<cad>\d{3})EPTS"
        r"(?P<month>\d{2})(?P<day>\d{2})(?P<year>\d{2})(?!\d)",
        Path(name).name,
        flags=re.IGNORECASE,
    )
    if match is None:
        return None
    raw_date = "".join(
        (match.group("month"), match.group("day"), match.group("year"))
    )
    try:
        submission_date = date(
            2000 + int(match.group("year")),
            int(match.group("month")),
            int(match.group("day")),
        ).isoformat()
        state = "parsed"
    except ValueError:
        submission_date = None
        state = "invalid_filename_date"
    return {
        "cad_id": match.group("cad"),
        "submission_date_raw": raw_date,
        "submission_date": submission_date,
        "state": state,
        "pattern": "XXXEPTSMMDDYY",
    }


def inspect_artifact(
    path: str | Path,
    *,
    member: str | None = None,
) -> dict[str, Any]:
    """Validate a delivered artifact and count exact row occurrences."""

    identity, members, rejected = discover_artifact_members(path)
    selected = _select_members(members, member)
    counts = {item.member_id: 0 for item in selected}
    cad_ids: set[str] = set()
    for occurrence in _iter_selected_rows(identity.path, selected):
        counts[occurrence.member.member_id] += 1
        cad = _raw_text(occurrence.raw_values[0]).strip()
        if cad:
            cad_ids.add(cad.zfill(3) if cad.isdigit() and len(cad) <= 3 else cad)

    filename_metadata: list[dict[str, Any]] = []
    artifact_name_metadata = _filename_metadata(identity.path.name)
    if artifact_name_metadata is not None:
        filename_metadata.append(
            {"source_name": identity.path.name, **artifact_name_metadata}
        )
    for item in selected:
        metadata = _filename_metadata(item.source_name)
        if metadata is not None and not any(
            row["source_name"] == item.source_name for row in filename_metadata
        ):
            filename_metadata.append({"source_name": item.source_name, **metadata})

    return {
        "canonical_ref": f"TX-EPTS:{identity.sha256}/inspection",
        "source_id": SOURCE_ID,
        "record_kind": "local_artifact_inspection",
        "artifact": identity.to_dict(),
        "layout_version": LAYOUT_VERSION,
        "schema_fingerprint": SCHEMA_FINGERPRINT,
        "schema_valid": True,
        "selected_members": [item.to_dict() for item in selected],
        "other_valid_members": [
            item.to_dict() for item in members if item not in selected
        ],
        "rejected_members": rejected,
        "member_record_counts": counts,
        "record_count": sum(counts.values()),
        "distinct_cad_ids": sorted(cad_ids),
        "filename_metadata": filename_metadata,
        "row_handling": {
            "source_occurrences_preserved": True,
            "multiple_account_rows_collapsed": False,
            "automatic_deduplication": False,
        },
        "specimen_state": {
            "this_artifact_inspected": True,
            "official_schema_basis": MANUAL_PUBLICATION,
        },
    }


def _identifier_text(value: Any, *, width: int | None = None) -> str | None:
    text = _raw_text(value).strip()
    if not text:
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    if width is not None and text.isdigit() and len(text) <= width:
        return text.zfill(width)
    return text


def _field_state(
    field: FieldSpec,
    value: Any,
    confidentiality_code: str,
) -> dict[str, Any]:
    text = _raw_text(value).strip()
    code = confidentiality_code.upper()
    flagged_numbers = CONFIDENTIAL_FIELD_NUMBERS.get(code, ())
    if text.upper() in REDACTION_MARKERS:
        state = "publisher_redaction_marker"
    elif not text and field.number in flagged_numbers:
        state = "publisher_blank_confidential_field"
    elif not text:
        state = "blank"
    else:
        state = "reported"
    return {
        "state": state,
        "confidentiality_code_flags_field": field.number in flagged_numbers,
        "raw_marker": text if state == "publisher_redaction_marker" else None,
    }


def _date_value(
    field: FieldSpec,
    value: Any,
    state: Mapping[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    raw = _json_value(value)
    result = {
        "raw": raw,
        "iso": None,
        "state": state["state"],
        "source_format": None,
    }
    issues: list[dict[str, Any]] = []
    if state["state"] != "reported":
        return result, issues
    if isinstance(value, datetime):
        result["iso"] = value.date().isoformat()
        result["source_format"] = "xlsx_datetime"
        return result, issues
    if isinstance(value, date):
        result["iso"] = value.isoformat()
        result["source_format"] = "xlsx_date"
        return result, issues

    candidate = _raw_text(value).strip()
    for date_format, label in (
        ("%m/%d/%Y", "MM/DD/YYYY"),
        ("%Y-%m-%d", "ISO_8601"),
    ):
        try:
            result["iso"] = datetime.strptime(candidate, date_format).date().isoformat()
            result["source_format"] = label
            if label != "MM/DD/YYYY":
                issues.append(
                    {
                        "field_number": field.number,
                        "header": field.header,
                        "issue": "non_manual_date_format",
                        "expected": "MM/DD/YYYY",
                        "observed": candidate,
                    }
                )
            return result, issues
        except ValueError:
            continue
    result["state"] = "invalid_date"
    issues.append(
        {
            "field_number": field.number,
            "header": field.header,
            "issue": "invalid_date",
            "expected": "MM/DD/YYYY",
            "observed": candidate,
        }
    )
    return result, issues


def _numeric_value(
    field: FieldSpec,
    value: Any,
    state: Mapping[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    raw = _json_value(value)
    result = {
        "raw": raw,
        "normalized": None,
        "value": None,
        "state": state["state"],
    }
    issues: list[dict[str, Any]] = []
    if state["state"] != "reported":
        return result, issues
    candidate = _raw_text(value).strip()
    try:
        decimal_value = Decimal(candidate)
    except InvalidOperation:
        result["state"] = "invalid_numeric"
        issues.append(
            {
                "field_number": field.number,
                "header": field.header,
                "issue": "invalid_numeric",
                "observed": candidate,
            }
        )
        return result, issues
    if not decimal_value.is_finite():
        result["state"] = "invalid_numeric"
        issues.append(
            {
                "field_number": field.number,
                "header": field.header,
                "issue": "non_finite_numeric",
                "observed": candidate,
            }
        )
        return result, issues

    normalized = format(decimal_value, "f")
    result["normalized"] = normalized
    decimal_places = max(0, -decimal_value.as_tuple().exponent)
    allowed_places = field.decimal_places
    if allowed_places is not None and decimal_places > allowed_places:
        if allowed_places != 0 or decimal_value != decimal_value.to_integral_value():
            issues.append(
                {
                    "field_number": field.number,
                    "header": field.header,
                    "issue": "decimal_places_exceed_layout",
                    "allowed": allowed_places,
                    "observed": decimal_places,
                }
            )
    if decimal_value == decimal_value.to_integral_value():
        result["value"] = int(decimal_value)
    else:
        result["value"] = normalized
    return result, issues


def _candidate(
    candidate_kind: str,
    components: Mapping[str, Any],
    *,
    required: Sequence[str] = (),
) -> dict[str, Any]:
    normalized = {
        key: value
        for key, value in components.items()
        if value is not None and value != ""
    }
    missing = [key for key in required if key not in normalized]
    return {
        "candidate_kind": candidate_kind,
        "candidate_key": sha256_fingerprint(
            {
                "source_id": SOURCE_ID,
                "candidate_kind": candidate_kind,
                "components": normalized,
            }
        ),
        "components": normalized,
        "required_components_present": not missing,
        "missing_required_components": missing,
        "automatic_deduplication": False,
    }


def _party_record(
    role: str,
    values: Mapping[str, Any],
    field_states: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    if role == "grantee":
        prefix = "GNTE"
    else:
        prefix = "GNTR"
    first_name = _identifier_text(values[f"{prefix}_FRST_NM"])
    last_or_business = _identifier_text(values[f"{prefix}_LST_BUS_NM"])
    display_name = " ".join(
        value for value in (first_name, last_or_business) if value
    ) or None
    address_headers = (
        f"{prefix}_LINE_1_AD_TX",
        f"{prefix}_LINE_2_AD_TX",
        f"{prefix}_CITY_NM",
        f"{prefix}_ST_CD",
        f"{prefix}_AD_ZP",
    )
    return {
        "role": role,
        "first_name": first_name,
        "last_name_or_business_name": last_or_business,
        "display_name": display_name,
        "address": {
            "line_1": _identifier_text(values[address_headers[0]]),
            "line_2": _identifier_text(values[address_headers[1]]),
            "city": _identifier_text(values[address_headers[2]]),
            "state": _identifier_text(values[address_headers[3]]),
            "postal_code": _identifier_text(values[address_headers[4]]),
            "field_states": {
                header: dict(field_states[header]) for header in address_headers
            },
        },
        "role_semantics": "submitted_transaction_role",
        "current_title_determination": False,
        "beneficial_ownership_determination": False,
    }


def normalize_occurrence(
    identity: ArtifactIdentity,
    occurrence: RowOccurrence,
    *,
    artifact_occurrence_index: int,
) -> dict[str, Any]:
    """Normalize one row while retaining all source values and identity layers."""

    raw_values = {
        header: _json_value(value)
        for header, value in zip(EXPECTED_HEADERS, occurrence.raw_values, strict=True)
    }
    source_values = {
        header: value
        for header, value in zip(EXPECTED_HEADERS, occurrence.raw_values, strict=True)
    }
    confidentiality_raw = _raw_text(source_values["CNFD_CD"]).strip()
    confidentiality_code = confidentiality_raw.upper()
    field_states = {
        field.header: _field_state(
            field,
            source_values[field.header],
            confidentiality_code,
        )
        for field in EPTS_FIELDS
    }
    validation_issues: list[dict[str, Any]] = []
    for field in EPTS_FIELDS:
        source_value = source_values[field.header]
        if field.field_type == "date" and isinstance(source_value, (date, datetime)):
            raw_text = source_value.strftime("%m/%d/%Y")
        else:
            raw_text = _raw_text(source_value)
        if len(raw_text) > field.length:
            validation_issues.append(
                {
                    "field_number": field.number,
                    "header": field.header,
                    "issue": "value_exceeds_layout_length",
                    "allowed": field.length,
                    "observed": len(raw_text),
                }
            )

    sale_date, issues = _date_value(
        FIELD_BY_HEADER["SALE_DT"],
        source_values["SALE_DT"],
        field_states["SALE_DT"],
    )
    validation_issues.extend(issues)
    deed_date, issues = _date_value(
        FIELD_BY_HEADER["DEED_DT"],
        source_values["DEED_DT"],
        field_states["DEED_DT"],
    )
    validation_issues.extend(issues)

    numeric: dict[str, dict[str, Any]] = {}
    for header in (
        "PROP_SALE_AM",
        "CERT_VAL_YR",
        "PROP_RPTD_LAND_AM",
        "PROP_RPTD_IMPV_AM",
        "PROP_RPTD_PPROP_AM",
        "PROP_RPTD_TOTL_AM",
        "PCT_OWNSHP",
        "PCT_COMP",
        "SQFT_IMPV_QY",
        "BUILT_YR",
        "LAND_UNIT_QY",
        "DY_ON_MRKT_QY",
        "PREV_RPTD_LAND_AM",
        "PREV_RPTD_IMPV_AM",
    ):
        normalized, issues = _numeric_value(
            FIELD_BY_HEADER[header],
            source_values[header],
            field_states[header],
        )
        numeric[header] = normalized
        validation_issues.extend(issues)

    cad_id = _identifier_text(source_values["CAD_ID"], width=3)
    school_district_id = _identifier_text(source_values["TU_ID"], width=6)
    account_number = _identifier_text(source_values["PROP_ID1_TX"])
    short_account_number = _identifier_text(source_values["PROP_SHT_ID"])
    deed_number = _identifier_text(source_values["DEED_NR"])
    deed_volume = _identifier_text(source_values["DEED_VOL_NR"])
    deed_page = _identifier_text(source_values["DEED_PAGE_NR"])
    grantee = _party_record("grantee", source_values, field_states)
    grantor = _party_record("grantor", source_values, field_states)

    row_hash = sha256_fingerprint(
        {
            "schema_fingerprint": SCHEMA_FINGERPRINT,
            "raw_values": raw_values,
        }
    )
    occurrence_components = {
        "artifact_sha256": identity.sha256,
        "member_id": occurrence.member.member_id,
        "source_row_number": occurrence.source_row_number,
        "logical_data_row": occurrence.logical_data_row,
        "row_hash": row_hash,
    }
    occurrence_id = sha256_fingerprint(occurrence_components)

    deed_locator = {
        "deed_number": deed_number,
        "volume": deed_volume,
        "page": deed_page,
        "deed_type_code": _identifier_text(source_values["DEED_TY_CD"]),
        "deed_date": deed_date,
        "locator_available": any((deed_number, deed_volume, deed_page)),
        "locator_role": "county_clerk_search_pivot",
        "recorded_instrument_copy": False,
        "title_determination": False,
    }
    property_candidate = _candidate(
        "cad_property",
        {
            "cad_id": cad_id,
            "account_number": account_number,
            "short_account_number": short_account_number,
        },
        required=("cad_id", "account_number"),
    )
    transaction_group_components = {
        "cad_id": cad_id,
        "sale_date": sale_date["iso"] or sale_date["raw"],
        "deed_date": deed_date["iso"] or deed_date["raw"],
        "deed_number": deed_number,
        "deed_volume": deed_volume,
        "deed_page": deed_page,
        "grantor": grantor["display_name"],
        "grantee": grantee["display_name"],
    }
    transaction_group_candidate = _candidate(
        "reported_transaction_group",
        transaction_group_components,
        required=("cad_id",),
    )
    transaction_account_candidate = _candidate(
        "reported_transaction_account",
        {
            **transaction_group_components,
            "account_number": account_number,
        },
        required=("cad_id", "account_number"),
    )

    sale_source_code = _raw_text(source_values["CAD_SALE_SRC_CD"]).strip().upper()
    validity_code = _raw_text(source_values["VALD_CD"]).strip().upper()
    recognized_confidentiality = (
        confidentiality_code in CONFIDENTIAL_FIELD_NUMBERS
    )
    if sale_source_code and sale_source_code not in SALE_SOURCE_CODES:
        validation_issues.append(
            {
                "field_number": 31,
                "header": "CAD_SALE_SRC_CD",
                "issue": "unrecognized_code",
                "observed": sale_source_code,
            }
        )
    if validity_code and validity_code not in VALIDITY_CODES:
        validation_issues.append(
            {
                "field_number": 32,
                "header": "VALD_CD",
                "issue": "unrecognized_code",
                "observed": validity_code,
            }
        )
    if confidentiality_code and not recognized_confidentiality:
        validation_issues.append(
            {
                "field_number": 33,
                "header": "CNFD_CD",
                "issue": "unrecognized_code",
                "observed": confidentiality_code,
            }
        )

    return {
        "canonical_ref": f"TX-EPTS:{occurrence_id}",
        "source_id": SOURCE_ID,
        "record_kind": "reported_property_transaction",
        "record_id": occurrence_id,
        "source_occurrence": {
            **occurrence_components,
            "artifact_occurrence_index": artifact_occurrence_index,
            "artifact_filename": identity.path.name,
            "artifact_path": str(identity.path),
            "member": occurrence.member.to_dict(),
            "physical_line_start": occurrence.physical_line_start,
            "physical_line_end": occurrence.physical_line_end,
            "cell_metadata": dict(occurrence.cell_metadata or {}),
        },
        "identity_candidates": {
            "source_occurrence": {
                "candidate_kind": "source_occurrence",
                "candidate_key": occurrence_id,
                "components": occurrence_components,
                "automatic_deduplication": False,
            },
            "property": property_candidate,
            "transaction_account": transaction_account_candidate,
            "transaction_group": transaction_group_candidate,
            "multiple_account_rows_must_be_retained": True,
        },
        "jurisdiction": {
            "state_code": STATE_CODE,
            "state_fips": STATE_FIPS,
            "cad_id": cad_id,
            "school_district_code": school_district_id,
        },
        "property": {
            "cad_id": cad_id,
            "school_district_code": school_district_id,
            "category_code": _identifier_text(source_values["PROP_CATG_CD"]),
            "short_account_number": short_account_number,
            "account_number": account_number,
            "legal_description": _identifier_text(source_values["PROP_ID2_TX"]),
            "parcel_address": {
                "value": _identifier_text(source_values["PRCL_AD_TX"]),
                **field_states["PRCL_AD_TX"],
            },
        },
        "transaction": {
            "sale_date": sale_date,
            "consideration": numeric["PROP_SALE_AM"],
            "deed_locator": deed_locator,
            "multiple_account_code": _identifier_text(
                source_values["MULT_ACCT_CD"]
            ),
            "additional_accounts_raw": _identifier_text(
                source_values["ADDNL_PROPS"]
            ),
            "financing_code": _identifier_text(source_values["FNC_CD"]),
            "days_on_market": numeric["DY_ON_MRKT_QY"],
        },
        "parties": [grantee, grantor],
        "appraisal": {
            "frozen_characteristics_code": _identifier_text(
                source_values["FRZN_CHAR_CD"]
            ),
            "most_recent_appraisal_year": numeric["CERT_VAL_YR"],
            "value_set_by_arb_code": _identifier_text(
                source_values["ARB_VAL_CD"]
            ),
            "current_values": {
                "land": numeric["PROP_RPTD_LAND_AM"],
                "improvement": numeric["PROP_RPTD_IMPV_AM"],
                "personal_property": numeric["PROP_RPTD_PPROP_AM"],
                "total": numeric["PROP_RPTD_TOTL_AM"],
            },
            "previous_values": {
                "land": numeric["PREV_RPTD_LAND_AM"],
                "improvement": numeric["PREV_RPTD_IMPV_AM"],
            },
        },
        "ownership_and_improvements": {
            "percent_ownership": numeric["PCT_OWNSHP"],
            "current_year_percent_complete": numeric["PCT_COMP"],
            "improvement_square_feet": numeric["SQFT_IMPV_QY"],
            "year_built": numeric["BUILT_YR"],
            "land_unit_type_code": _identifier_text(
                source_values["LAND_UNIT_TY_CD"]
            ),
            "land_units": numeric["LAND_UNIT_QY"],
        },
        "reporting": {
            "sale_source": {
                "raw": _json_value(source_values["CAD_SALE_SRC_CD"]),
                "code": sale_source_code or None,
                "description": SALE_SOURCE_CODES.get(sale_source_code),
                "recognized": sale_source_code in SALE_SOURCE_CODES,
                "state": (
                    "blank"
                    if not sale_source_code
                    else (
                        "reported"
                        if sale_source_code in SALE_SOURCE_CODES
                        else "unrecognized"
                    )
                ),
            },
            "validity": {
                "raw": _json_value(source_values["VALD_CD"]),
                "code": validity_code or None,
                "description": VALIDITY_CODES.get(validity_code),
                "recognized": validity_code in VALIDITY_CODES,
                "state": (
                    "blank"
                    if not validity_code
                    else (
                        "reported"
                        if validity_code in VALIDITY_CODES
                        else "unrecognized"
                    )
                ),
            },
            "confidentiality": {
                "raw": _json_value(source_values["CNFD_CD"]),
                "code": confidentiality_code or None,
                "description": CONFIDENTIAL_CODE_DESCRIPTIONS.get(
                    confidentiality_code
                ),
                "recognized": recognized_confidentiality,
                "field_numbers": list(
                    CONFIDENTIAL_FIELD_NUMBERS.get(confidentiality_code, ())
                ),
                "headers": [
                    FIELD_BY_NUMBER[number].header
                    for number in CONFIDENTIAL_FIELD_NUMBERS.get(
                        confidentiality_code, ()
                    )
                ],
                "other_confidential_information": confidentiality_code == "P",
                "no_confidential_information": confidentiality_code == "Q",
                "state": (
                    "blank"
                    if not confidentiality_code
                    else (
                        "reported"
                        if recognized_confidentiality
                        else "unrecognized"
                    )
                ),
                "scope": (
                    "other_unspecified"
                    if confidentiality_code == "P"
                    else (
                        "none"
                        if confidentiality_code == "Q"
                        else (
                            "specified_fields"
                            if recognized_confidentiality
                            else "unknown"
                        )
                    )
                ),
                "sensitive_field_states": {
                    FIELD_BY_NUMBER[number].header: dict(
                        field_states[FIELD_BY_NUMBER[number].header]
                    )
                    for number in sorted(SENSITIVE_FIELD_NUMBERS)
                },
            },
            "comments": {
                "line_1": _identifier_text(source_values["CAD_LINE_1_CMNT_TX"]),
                "line_2": _identifier_text(source_values["CAD_LINE_2_CMNT_TX"]),
            },
        },
        "source_semantics": {
            "reporting_system": "appraisal_district_epts",
            "reported_transaction_record": True,
            "recorded_instrument_copy": False,
            "title_determination": False,
            "deed_locator_role": "county_clerk_search_pivot",
        },
        "field_states": field_states,
        "raw_values": raw_values,
        "validation": {
            "layout_field_count": len(occurrence.raw_values),
            "layout_valid": True,
            "value_issue_count": len(validation_issues),
            "value_issues": validation_issues,
        },
    }


def _member_set_fingerprint(members: Sequence[MemberSpec]) -> str:
    return sha256_fingerprint(
        {
            "schema_fingerprint": SCHEMA_FINGERPRINT,
            "members": [
                {
                    "member_id": member.member_id,
                    "file_format": member.file_format,
                    "header_row_number": member.header_row_number,
                    "delimiter": member.delimiter,
                    "encoding": member.encoding,
                    "archive_member": member.archive_member,
                    "archive_member_index": member.archive_member_index,
                    "sheet_name": member.sheet_name,
                }
                for member in members
            ],
        }
    )


def _encode_cursor(
    *,
    identity: ArtifactIdentity,
    criteria_fingerprint: str,
    member_set_fingerprint: str,
    next_occurrence_index: int,
) -> str:
    payload = {
        "version": CURSOR_VERSION,
        "source_id": SOURCE_ID,
        "artifact_sha256": identity.sha256,
        "schema_fingerprint": SCHEMA_FINGERPRINT,
        "criteria_fingerprint": criteria_fingerprint,
        "member_set_fingerprint": member_set_fingerprint,
        "next_occurrence_index": next_occurrence_index,
    }
    envelope = {
        "payload": payload,
        "checksum": sha256_fingerprint(payload),
    }
    encoded = base64.urlsafe_b64encode(
        canonical_json(envelope).encode("utf-8")
    ).decode("ascii")
    return CURSOR_PREFIX + encoded.rstrip("=")


def _decode_cursor(
    cursor: str,
    *,
    identity: ArtifactIdentity,
    criteria_fingerprint: str,
    member_set_fingerprint: str,
) -> int:
    if not cursor.startswith(CURSOR_PREFIX):
        raise EPTSCursorError("cursor prefix does not identify Texas EPTS")
    encoded = cursor[len(CURSOR_PREFIX) :]
    try:
        padding = "=" * (-len(encoded) % 4)
        decoded = base64.urlsafe_b64decode(encoded + padding).decode("utf-8")
        envelope = json.loads(decoded)
    except (
        ValueError,
        UnicodeDecodeError,
        json.JSONDecodeError,
        binascii.Error,
    ) as error:
        raise EPTSCursorError("cursor payload is not valid") from error
    if not isinstance(envelope, dict):
        raise EPTSCursorError("cursor envelope is not an object")
    payload = envelope.get("payload")
    checksum = envelope.get("checksum")
    if not isinstance(payload, dict) or not isinstance(checksum, str):
        raise EPTSCursorError("cursor envelope is incomplete")
    if checksum != sha256_fingerprint(payload):
        raise EPTSCursorError("cursor checksum does not match its payload")

    expected = {
        "version": CURSOR_VERSION,
        "source_id": SOURCE_ID,
        "artifact_sha256": identity.sha256,
        "schema_fingerprint": SCHEMA_FINGERPRINT,
        "criteria_fingerprint": criteria_fingerprint,
        "member_set_fingerprint": member_set_fingerprint,
    }
    for key, value in expected.items():
        if payload.get(key) != value:
            raise EPTSCursorError(f"cursor {key} does not match this query")
    next_index = payload.get("next_occurrence_index")
    if (
        isinstance(next_index, bool)
        or not isinstance(next_index, int)
        or next_index < 0
    ):
        raise EPTSCursorError("cursor occurrence index is invalid")
    if set(payload) != {*expected, "next_occurrence_index"}:
        raise EPTSCursorError("cursor payload contains an unexpected shape")
    return next_index


def _search_headers(field: str) -> tuple[str, ...]:
    if field == "all":
        return EXPECTED_HEADERS
    try:
        return SEARCH_FIELDS[field]
    except KeyError as error:
        choices = ", ".join(("all", *SEARCH_FIELDS))
        raise EPTSError(f"search field must be one of: {choices}") from error


def _occurrence_matches(
    occurrence: RowOccurrence,
    query_text: str,
    headers: Sequence[str],
) -> bool:
    needle = query_text.casefold()
    values = {
        header: value
        for header, value in zip(
            EXPECTED_HEADERS,
            occurrence.raw_values,
            strict=True,
        )
    }
    return any(needle in _raw_text(values[header]).casefold() for header in headers)


def query_artifact(
    path: str | Path,
    *,
    operation: str,
    selector: str | None = None,
    search_field: str = "all",
    limit: int = DEFAULT_LIMIT,
    cursor: str | None = None,
    member: str | None = None,
) -> dict[str, Any]:
    """Page a local parse or substring search without collapsing occurrences."""

    if operation not in {"parse", "search"}:
        raise EPTSError("local artifact operation must be parse or search")
    if isinstance(limit, bool) or not isinstance(limit, int) or limit <= 0:
        raise EPTSError("limit must be a positive integer")
    if operation == "search":
        query_text = (selector or "").strip()
        if not query_text:
            raise EPTSError("search query must not be blank")
        headers = _search_headers(search_field)
    else:
        query_text = ""
        headers = EXPECTED_HEADERS
        search_field = "all"

    identity, members, rejected = discover_artifact_members(path)
    selected_members = _select_members(members, member)
    member_fingerprint = _member_set_fingerprint(selected_members)
    criteria = {
        "operation": operation,
        "query_casefold": query_text.casefold(),
        "search_field": search_field,
        "search_headers": list(headers),
        "selected_member_ids": [
            selected.member_id for selected in selected_members
        ],
    }
    criteria_fingerprint = sha256_fingerprint(criteria)
    start_index = 0
    if cursor is not None:
        start_index = _decode_cursor(
            cursor,
            identity=identity,
            criteria_fingerprint=criteria_fingerprint,
            member_set_fingerprint=member_fingerprint,
        )

    records: list[dict[str, Any]] = []
    next_cursor = None
    visited_occurrences = 0
    matched_occurrences = 0
    for occurrence_index, occurrence in enumerate(
        _iter_selected_rows(identity.path, selected_members)
    ):
        visited_occurrences = occurrence_index + 1
        if occurrence_index < start_index:
            continue
        if operation == "search" and not _occurrence_matches(
            occurrence,
            query_text,
            headers,
        ):
            continue
        matched_occurrences += 1
        if len(records) == limit:
            next_cursor = _encode_cursor(
                identity=identity,
                criteria_fingerprint=criteria_fingerprint,
                member_set_fingerprint=member_fingerprint,
                next_occurrence_index=occurrence_index,
            )
            break
        records.append(
            normalize_occurrence(
                identity,
                occurrence,
                artifact_occurrence_index=occurrence_index,
            )
        )

    return {
        "artifact": identity.to_dict(),
        "layout_version": LAYOUT_VERSION,
        "schema_fingerprint": SCHEMA_FINGERPRINT,
        "criteria": criteria,
        "criteria_fingerprint": criteria_fingerprint,
        "selected_members": [selected.to_dict() for selected in selected_members],
        "rejected_members": rejected,
        "start_occurrence_index": start_index,
        "visited_occurrences": visited_occurrences,
        "matched_occurrences_in_scan": matched_occurrences,
        "records": records,
        "next_cursor": next_cursor,
        "automatic_deduplication": False,
    }


def build_query(args: argparse.Namespace) -> PublicRecordsQuery:
    parameters: dict[str, Any] = {}
    for name in (
        "artifact",
        "query_text",
        "field",
        "member",
        "cad_id",
        "start_date",
        "end_date",
        "preferred_format",
    ):
        value = getattr(args, name, None)
        if value not in (None, [], ()):
            parameters[name] = value
    requested_limit = getattr(args, "limit", None)
    cursor = getattr(args, "cursor", None)
    return PublicRecordsQuery(
        source=SOURCE_METADATA,
        jurisdiction=JURISDICTION,
        query=QueryMetadata(
            operation=args.command,
            parameters=parameters,
            requested_limit=requested_limit,
            cursor=cursor,
            metadata={
                "layout_version": LAYOUT_VERSION,
                "schema_fingerprint": SCHEMA_FINGERPRINT,
                "network_request_performed": False,
            },
        ),
    )


def _failure(
    query: PublicRecordsQuery,
    *,
    status: ResultStatus,
    code: str,
    message: str,
    category: str,
) -> PublicRecordsResult:
    return PublicRecordsResult.failure(
        query,
        status,
        [
            PublicRecordsError(
                code=code,
                message=message,
                category=category,
                retryable=False,
            )
        ],
        warnings=SOURCE_WARNINGS,
    )


def execute(args: argparse.Namespace) -> PublicRecordsResult:
    """Execute a discovery, handoff, inspection, parse, or search operation."""

    query = build_query(args)
    try:
        raw_artifact_refs: list[str] = []
        next_cursor = None
        if args.command == "discover":
            records = [source_discovery_record()]
        elif args.command == "schema":
            records = [schema_record()]
        elif args.command == "request-plan":
            records = [
                request_plan_record(
                    cad_ids=args.cad_id,
                    start_date=args.start_date,
                    end_date=args.end_date,
                    preferred_format=args.preferred_format,
                )
            ]
        elif args.command == "inspect":
            records = [inspect_artifact(args.artifact, member=args.member)]
            raw_artifact_refs = [str(Path(args.artifact).expanduser().resolve())]
        elif args.command in {"parse", "search"}:
            page = query_artifact(
                args.artifact,
                operation=args.command,
                selector=getattr(args, "query_text", None),
                search_field=getattr(args, "field", "all"),
                limit=args.limit,
                cursor=args.cursor,
                member=args.member,
            )
            records = page["records"]
            next_cursor = page["next_cursor"]
            raw_artifact_refs = [str(Path(args.artifact).expanduser().resolve())]
        else:
            raise EPTSError(f"unsupported operation: {args.command}")
        return PublicRecordsResult.success(
            query,
            records,
            next_cursor=next_cursor,
            raw_artifact_refs=raw_artifact_refs,
            warnings=SOURCE_WARNINGS,
        )
    except EPTSCursorError as error:
        return _failure(
            query,
            status=ResultStatus.SOURCE_CHANGED,
            code="epts_cursor_mismatch",
            message=str(error),
            category="query_selection",
        )
    except EPTSLayoutError as error:
        return _failure(
            query,
            status=ResultStatus.SOURCE_CHANGED,
            code="epts_layout_mismatch",
            message=str(error),
            category="source_schema",
        )
    except OSError as error:
        return _failure(
            query,
            status=ResultStatus.UNAVAILABLE,
            code="local_artifact_unavailable",
            message=str(error),
            category="local_io",
        )
    except EPTSError as error:
        return _failure(
            query,
            status=ResultStatus.UNAVAILABLE,
            code="invalid_epts_query",
            message=str(error),
            category="query_selection",
        )


def _emit(result: PublicRecordsResult, args: argparse.Namespace) -> None:
    data = result.to_dict()
    if write_output(
        data,
        args,
        summary=f"Texas EPTS {args.command} ({result.status.value})",
        result_count=len(result.records)
        if result.status
        in {ResultStatus.OK, ResultStatus.NO_RESULTS, ResultStatus.PARTIAL}
        else None,
    ):
        return
    if getattr(args, "json_out", False):
        print(json.dumps(data, indent=2, sort_keys=True))
        return
    print(
        f"Texas EPTS {args.command}: {result.status.value} "
        f"({len(result.records)} records)"
    )
    if result.next_cursor:
        print(f"Next cursor: {result.next_cursor}")
    for error in result.errors:
        print(f"ERROR [{error.code}]: {error.message}", file=sys.stderr)


def _add_artifact_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("artifact", help="Local EPTS artifact path")
    parser.add_argument(
        "--member",
        help="Exact member ID, archive member, or unique workbook sheet",
    )


def _add_page_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--limit", type=int, default=DEFAULT_LIMIT)
    parser.add_argument("--cursor")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Plan acquisition and inspect, parse, or search local Texas "
            "Comptroller EPTS artifacts"
        )
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    discover = subparsers.add_parser(
        "discover",
        help="Describe the official source and acquisition route",
    )
    add_output_args(discover)

    schema = subparsers.add_parser(
        "schema",
        help="Emit the official September 2025 52-field layout",
    )
    add_output_args(schema)

    request_plan = subparsers.add_parser(
        "request-plan",
        help="Prepare a CRRS/email request handoff without submitting it",
    )
    request_plan.add_argument(
        "--cad-id",
        action="append",
        default=[],
        help="Three-digit appraisal district code; repeat for multiple CADs",
    )
    request_plan.add_argument("--start-date", help="ISO date")
    request_plan.add_argument("--end-date", help="ISO date")
    request_plan.add_argument(
        "--preferred-format",
        default="native-machine-readable",
    )
    add_output_args(request_plan)

    inspect_parser = subparsers.add_parser(
        "inspect",
        help="Validate and inventory a delivered artifact",
    )
    _add_artifact_args(inspect_parser)
    add_output_args(inspect_parser)

    parse = subparsers.add_parser(
        "parse",
        help="Stream normalized row occurrences from a delivered artifact",
    )
    _add_artifact_args(parse)
    _add_page_args(parse)
    add_output_args(parse)

    search = subparsers.add_parser(
        "search",
        help="Search selected raw fields in a delivered artifact",
    )
    _add_artifact_args(search)
    search.add_argument("query_text")
    search.add_argument(
        "--field",
        choices=("all", *SEARCH_FIELDS),
        default="all",
    )
    _add_page_args(search)
    add_output_args(search)
    return parser


def _validate_args(parser: argparse.ArgumentParser, args: argparse.Namespace) -> None:
    limit = getattr(args, "limit", None)
    if limit is not None and limit <= 0:
        parser.error("--limit must be positive")


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    _validate_args(parser, args)
    result = execute(args)
    _emit(result, args)
    if result.status in {
        ResultStatus.OK,
        ResultStatus.NO_RESULTS,
        ResultStatus.PARTIAL,
    }:
        return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
