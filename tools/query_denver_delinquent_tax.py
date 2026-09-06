#!/usr/bin/env python3
"""Denver Treasury delinquent-real-property-tax workbook adapter.

The City and County of Denver publishes an annual XLSX used for the real
property tax-lien sale.  The publication page identifies the current release;
the workbook preserves tax-year blocks, parcel identifiers, published owner
names, addresses, valuation, tax, interest, fees, total due, and the source's
tax-sale and partial-payment indicators.

Usage:
    uv run python tools/query_denver_delinquent_tax.py discover
    uv run python tools/query_denver_delinquent_tax.py probe --output /tmp/probe.json
    uv run python tools/query_denver_delinquent_tax.py inspect /tmp/list.xlsx
    uv run python tools/query_denver_delinquent_tax.py download \
        --destination /tmp/denver-delinquent-tax.xlsx
    uv run python tools/query_denver_delinquent_tax.py search \
        --owner "EXAMPLE LLC" --output /tmp/results.json
    uv run python tools/query_denver_delinquent_tax.py search \
        --artifact /tmp/list.xlsx --tax-year 2024
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sys
import tempfile
import time
from collections import Counter
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import unquote, urljoin, urlsplit
from urllib.request import Request, urlopen
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

try:
    from tools.lead_tracker import log_search
    from tools.output_util import add_output_args, write_output
    from tools.public_records_bulk import (
        ArchiveSafetyError,
        ArchiveSafetyPolicy,
        BulkArtifact,
        BulkHTTPStatusError,
        BulkSourceError,
        BulkTransferClient,
        BulkTransportError,
        inspect_zip,
    )
    from tools.public_records_catalog import (
        DEFAULT_DB_PATH as DEFAULT_CATALOG_DB_PATH,
        AcquisitionUnavailableError,
        CatalogError,
        acquisition_result_status,
    )
    from tools.public_records_contract import (
        JurisdictionMetadata,
        PublicRecordsError,
        PublicRecordsQuery,
        PublicRecordsResult,
        QueryMetadata,
        ResultStatus,
        SourceMetadata,
        canonical_json,
        sha256_fingerprint,
    )
    from tools.public_records_store import canonical_property_ref
    from tools.seed_public_records_catalog import (
        DEFAULT_CONFIG_PATH as DEFAULT_CATALOG_CONFIG_PATH,
        ensure_catalog_source,
    )
except ImportError:
    from lead_tracker import log_search
    from output_util import add_output_args, write_output
    from public_records_bulk import (
        ArchiveSafetyError,
        ArchiveSafetyPolicy,
        BulkArtifact,
        BulkHTTPStatusError,
        BulkSourceError,
        BulkTransferClient,
        BulkTransportError,
        inspect_zip,
    )
    from public_records_catalog import (
        DEFAULT_DB_PATH as DEFAULT_CATALOG_DB_PATH,
        AcquisitionUnavailableError,
        CatalogError,
        acquisition_result_status,
    )
    from public_records_contract import (
        JurisdictionMetadata,
        PublicRecordsError,
        PublicRecordsQuery,
        PublicRecordsResult,
        QueryMetadata,
        ResultStatus,
        SourceMetadata,
        canonical_json,
        sha256_fingerprint,
    )
    from public_records_store import canonical_property_ref
    from seed_public_records_catalog import (
        DEFAULT_CONFIG_PATH as DEFAULT_CATALOG_CONFIG_PATH,
        ensure_catalog_source,
    )


SOURCE_ID = "us-co-denver-delinquent-real-property-tax-list"
DENVER_GEOID = "08031"
PUBLICATION_PAGE = (
    "https://www.denvergov.org/Government/Agencies-Departments-Offices/"
    "Agencies-Departments-Offices-Directory/Department-of-Finance/"
    "Our-Divisions/Treasury/Property-Taxes/"
    "Real-Estate-Delinquent-Taxes-and-Tax-Lien-Sale"
)
OFFICIAL_HOST = "www.denvergov.org"
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
USER_AGENT = "Ithildin-Public-Records/1.0"

# These protect the page and OOXML parser, not the number of returned records.
MAX_PUBLICATION_PAGE_BYTES = 5 * 1024 * 1024
DEFAULT_MAX_ARCHIVE_MEMBERS = 1_000
DEFAULT_MAX_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
DEFAULT_MAX_MEMBER_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
DEFAULT_MAX_COMPRESSION_RATIO = 1_000.0
DEFAULT_SAMPLE_BYTES = 4_096

EXPECTED_HEADERS = (
    "Owner Name",
    "Parcel ID",
    "Parcel Valuation",
    "Tax Owed",
    "Interest",
    "Fees And ADV",
    "Total Owed",
    "Tax Sale Indicator",
    "Partially Paid",
    "Additional Owner1",
    "Additional Owner2",
    "Additional Owner3",
    "Parcel Address Line1",
    "Legal Description",
)
YEAR_MARKER_RE = re.compile(r"^\*+\s*(?P<year>\d{4})\s*$")
TAX_HEADING_RE = re.compile(
    r"\b(?P<year>20\d{2})\s+DELINQUENT\s+REAL\s+"
    r"(?:ESTATE|PROPERTY)\s+TAXES\b",
    re.I,
)
RELEASE_DATE_RE = re.compile(
    r"(?<!\d)(?P<month>\d{1,2})[.-](?P<day>\d{1,2})[.-]"
    r"(?P<year>20\d{2})(?!\d)"
)
CURSOR_VERSION = "v1"
CURSOR_RE = re.compile(
    r"^denver-delinquent-tax:v1:"
    r"criteria:(?P<criteria>[0-9a-f]{64}):"
    r"artifact:(?P<artifact>[0-9a-f]{64}):"
    r"row:(?P<row>\d+)$"
)
TRANSIENT_ARTIFACT_PATH_RE = re.compile(
    r"/(?:private/)?tmp/osint-denver-tax-[A-Za-z0-9._-]+"
    r"(?:/[^\s,'\"\]\}]+)?"
)

RELEASE_SCOPE_CATEGORIES = (
    "general_real_estate_tax",
    "local_public_improvement_special_assessment",
    "maintenance_district",
    "improvement_district",
    "sanitary_sewer_charge",
    "storm_drainage_charge",
)

SOURCE_WARNINGS = (
    "The workbook is an annual delinquency publication, not a current-balance response.",
    (
        "The publication covers several delinquency categories, but the workbook "
        "does not identify a category on each row."
    ),
    (
        "Published owner names are source observations associated with the tax "
        "account and are not a title chain."
    ),
    (
        "A blank tax-sale or partial-payment indicator means the workbook did "
        "not mark that condition; it is not an affirmative negative."
    ),
)

SOURCE_METADATA = SourceMetadata(
    source_id=SOURCE_ID,
    name="Denver Delinquent Real Property Tax List",
    source_role="property_tax_delinquency_bulk",
    base_url=PUBLICATION_PAGE,
    dataset_id="denver-annual-delinquent-real-property-tax-list",
    metadata={
        "authority": "City and County of Denver Department of Finance",
        "operator": "Denver Treasury Division",
        "coverage": "City and County of Denver, Colorado",
        "release_scope_categories": RELEASE_SCOPE_CATEGORIES,
        "stable_key_fields": ["tax_year", "parcel_id"],
    },
)
JURISDICTION = JurisdictionMetadata(
    jurisdiction_id=DENVER_GEOID,
    name="City and County of Denver, Colorado",
    state_code="CO",
    county_fips=DENVER_GEOID,
    locality="Denver",
)
ADAPTER_SCHEMA_FINGERPRINT = sha256_fingerprint(
    {
        "source_id": SOURCE_ID,
        "normalization_version": 1,
        "headers": EXPECTED_HEADERS,
        "tax_year_marker": YEAR_MARKER_RE.pattern,
        "stable_key_fields": ["tax_year", "parcel_id"],
    }
)


class DenverTaxSourceChanged(BulkSourceError):
    """The official page, artifact, or workbook no longer matches its contract."""

    result_status = ResultStatus.SOURCE_CHANGED
    code = "denver_tax_source_changed"
    category = "source_schema"


class DenverTaxUnavailable(BulkSourceError):
    """The official source or a requested local artifact is unavailable."""

    result_status = ResultStatus.UNAVAILABLE
    code = "denver_tax_unavailable"
    category = "source_access"
    retryable = True


class DenverTaxQueryError(BulkSourceError):
    """A caller selector or local output choice is invalid."""

    result_status = ResultStatus.UNAVAILABLE
    code = "denver_tax_query_invalid"
    category = "query"

    def __init__(
        self,
        message: str,
        *,
        code: str | None = None,
        details: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(message, details=details)
        if code is not None:
            self.code = code


def _official_url(url: str, *, artifact: bool = False) -> str:
    parsed = urlsplit(url)
    if parsed.scheme != "https" or parsed.hostname != OFFICIAL_HOST:
        raise DenverTaxSourceChanged(
            "Denver source URL is no longer on the official HTTPS host",
            details={"url": url},
        )
    if artifact and (
        not parsed.path.startswith("/files/assets/public/")
        or not parsed.path.casefold().endswith(".xlsx")
    ):
        raise DenverTaxSourceChanged(
            "Denver release link is not an official public XLSX artifact",
            details={"url": url},
        )
    return url


@dataclass(frozen=True)
class ReleaseLink:
    tax_year: int
    url: str
    filename: str
    link_text: str
    page_size_label: str | None
    release_date: str | None

    @property
    def artifact_id(self) -> str:
        digest = hashlib.sha256(self.url.encode("utf-8")).hexdigest()[:16]
        return f"denver-delinquent-tax-{self.tax_year}-{digest}"

    def bulk_artifact(self) -> BulkArtifact:
        return BulkArtifact(
            artifact_id=self.artifact_id,
            url=self.url,
            filename=self.filename,
            media_type=XLSX_MEDIA_TYPE,
            archive_format="xlsx",
            metadata={
                "tax_year": self.tax_year,
                "publication_page": PUBLICATION_PAGE,
                "page_size_label": self.page_size_label,
                "release_date": self.release_date,
                "release_date_basis": (
                    "artifact_filename" if self.release_date else None
                ),
            },
        )

    def to_record(self) -> dict[str, Any]:
        canonical_ref = (
            f"BULK:{SOURCE_ID}/{self.tax_year}/{self.artifact_id}"
        )
        return {
            "source_id": SOURCE_ID,
            "record_kind": "bulk_dataset_artifact",
            "record_scope": "annual_property_tax_delinquency_release",
            "canonical_ref": canonical_ref,
            "evidence_ref": canonical_ref,
            "native_document_id": self.artifact_id,
            "artifact_id": self.artifact_id,
            "tax_year": self.tax_year,
            "release_date": self.release_date,
            "release_date_basis": (
                "artifact_filename" if self.release_date else None
            ),
            "filename": self.filename,
            "format": "xlsx",
            "media_type": XLSX_MEDIA_TYPE,
            "artifact_url": self.url,
            "publication_page": PUBLICATION_PAGE,
            "link_text": self.link_text,
            "page_size_label": self.page_size_label,
            "release_scope_categories": list(RELEASE_SCOPE_CATEGORIES),
            "stable_key_fields": ["tax_year", "parcel_id"],
        }


class _ReleasePageParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.current_tax_year: int | None = None
        self.current_anchor: dict[str, Any] | None = None
        self.anchors: list[dict[str, Any]] = []

    def handle_starttag(
        self,
        tag: str,
        attrs: list[tuple[str, str | None]],
    ) -> None:
        if tag.casefold() != "a":
            return
        values = {key.casefold(): value for key, value in attrs}
        self.current_anchor = {
            "href": values.get("href"),
            "title": values.get("title"),
            "text": [],
            "tax_year": self.current_tax_year,
        }

    def handle_data(self, data: str) -> None:
        match = TAX_HEADING_RE.search(data)
        if match:
            self.current_tax_year = int(match.group("year"))
        if self.current_anchor is not None:
            self.current_anchor["text"].append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.casefold() != "a" or self.current_anchor is None:
            return
        anchor = dict(self.current_anchor)
        anchor["text"] = re.sub(
            r"\s+",
            " ",
            " ".join(anchor["text"]),
        ).strip()
        self.anchors.append(anchor)
        self.current_anchor = None


def _release_date(filename: str) -> str | None:
    match = RELEASE_DATE_RE.search(filename)
    if not match:
        return None
    try:
        return date(
            int(match.group("year")),
            int(match.group("month")),
            int(match.group("day")),
        ).isoformat()
    except ValueError:
        return None


def parse_release_page(
    html: str,
    *,
    page_url: str = PUBLICATION_PAGE,
) -> ReleaseLink:
    """Resolve the latest uniquely identified official workbook release."""
    parser = _ReleasePageParser()
    parser.feed(html)
    candidates: list[ReleaseLink] = []
    for anchor in parser.anchors:
        href = str(anchor.get("href") or "").strip()
        label = re.sub(
            r"\s+",
            " ",
            f"{anchor.get('title') or ''} {anchor.get('text') or ''}",
        ).strip()
        if (
            "delinquent real property tax list" not in label.casefold()
            or not href
        ):
            continue
        tax_year = anchor.get("tax_year")
        if not isinstance(tax_year, int):
            raise DenverTaxSourceChanged(
                "Denver release link is not associated with a tax-year heading"
            )
        url = _official_url(urljoin(page_url, href), artifact=True)
        filename = Path(unquote(urlsplit(url).path)).name
        size_match = re.search(
            r"\(\s*XLSX\s*,\s*([0-9.]+\s*[KMGT]?B)\s*\)",
            str(anchor.get("text") or ""),
            re.I,
        )
        candidates.append(
            ReleaseLink(
                tax_year=tax_year,
                url=url,
                filename=filename,
                link_text=str(anchor.get("text") or ""),
                page_size_label=(
                    size_match.group(1).replace(" ", "")
                    if size_match
                    else None
                ),
                release_date=_release_date(filename),
            )
        )
    if not candidates:
        raise DenverTaxSourceChanged(
            "Denver publication page contains no recognized tax-list XLSX"
        )
    latest_year = max(item.tax_year for item in candidates)
    latest = [item for item in candidates if item.tax_year == latest_year]
    unique = {item.url: item for item in latest}
    if len(unique) != 1:
        raise DenverTaxSourceChanged(
            "Denver publication page has ambiguous latest tax-list artifacts",
            details={
                "tax_year": latest_year,
                "urls": sorted(unique),
            },
        )
    return next(iter(unique.values()))


class _RateLimitedOpener:
    def __init__(
        self,
        *,
        opener: Callable[..., Any] = urlopen,
        minimum_interval: float = 0.0,
        sleeper: Callable[[float], None] = time.sleep,
        clock: Callable[[], float] = time.monotonic,
    ) -> None:
        self.opener = opener
        self.minimum_interval = minimum_interval
        self.sleeper = sleeper
        self.clock = clock
        self.last_request_at: float | None = None

    def __call__(self, request: Request, *, timeout: float) -> Any:
        _official_url(request.full_url)
        if self.last_request_at is not None:
            remaining = self.minimum_interval - (
                self.clock() - self.last_request_at
            )
            if remaining > 0:
                self.sleeper(remaining)
        self.last_request_at = self.clock()
        response = self.opener(request, timeout=timeout)
        final_url = (
            response.geturl()
            if hasattr(response, "geturl")
            else request.full_url
        )
        try:
            _official_url(str(final_url))
        except Exception:
            if hasattr(response, "close"):
                response.close()
            raise
        return response


class DenverDelinquentTaxClient:
    """Discovery and transfer client for the official Denver release."""

    def __init__(
        self,
        *,
        timeout: float = 60.0,
        max_attempts: int = 3,
        minimum_interval: float = 0.0,
        chunk_size: int = 1024 * 1024,
        opener: Callable[..., Any] = urlopen,
        sleeper: Callable[[float], None] = time.sleep,
    ) -> None:
        if timeout <= 0:
            raise ValueError("timeout must be positive")
        if max_attempts <= 0:
            raise ValueError("max_attempts must be positive")
        if minimum_interval < 0:
            raise ValueError("minimum_interval must not be negative")
        self.timeout = timeout
        self.max_attempts = max_attempts
        self.sleeper = sleeper
        self.opener = _RateLimitedOpener(
            opener=opener,
            minimum_interval=minimum_interval,
            sleeper=sleeper,
        )
        self.transfer = BulkTransferClient(
            timeout=timeout,
            max_attempts=max_attempts,
            chunk_size=chunk_size,
            user_agent=USER_AGENT,
            opener=self.opener,
            sleeper=sleeper,
        )

    def discover(self) -> ReleaseLink:
        request = Request(
            PUBLICATION_PAGE,
            headers={
                "Accept": "text/html,application/xhtml+xml",
                "User-Agent": USER_AGENT,
            },
            method="GET",
        )
        last_error: BulkSourceError | None = None
        for attempt in range(1, self.max_attempts + 1):
            try:
                with self.opener(request, timeout=self.timeout) as response:
                    status = int(getattr(response, "status", 200))
                    if status < 200 or status >= 300:
                        raise BulkHTTPStatusError(status, PUBLICATION_PAGE)
                    content_type = str(
                        response.headers.get("Content-Type") or ""
                    )
                    if "text/html" not in content_type.casefold():
                        raise DenverTaxSourceChanged(
                            "Denver publication page is no longer HTML",
                            details={"content_type": content_type},
                        )
                    body = response.read(MAX_PUBLICATION_PAGE_BYTES + 1)
                    if len(body) > MAX_PUBLICATION_PAGE_BYTES:
                        raise DenverTaxSourceChanged(
                            "Denver publication page exceeded its parser bound",
                            details={
                                "max_page_bytes": MAX_PUBLICATION_PAGE_BYTES
                            },
                        )
                    return parse_release_page(
                        body.decode("utf-8", errors="replace")
                    )
            except HTTPError as error:
                last_error = BulkHTTPStatusError(
                    error.code,
                    PUBLICATION_PAGE,
                    error.read(500).decode("utf-8", errors="replace"),
                )
            except (URLError, TimeoutError, ConnectionError, OSError) as error:
                last_error = BulkTransportError(
                    f"Denver publication-page request failed: {error}",
                    details={"url": PUBLICATION_PAGE},
                )
            except BulkSourceError as error:
                last_error = error
            if (
                last_error is None
                or not last_error.retryable
                or attempt >= self.max_attempts
            ):
                assert last_error is not None
                raise last_error
            self.sleeper(min(0.25 * (2 ** (attempt - 1)), 5.0))
        assert last_error is not None
        raise last_error

    def probe(
        self,
        release: ReleaseLink,
        *,
        sample_bytes: int = DEFAULT_SAMPLE_BYTES,
    ) -> dict[str, Any]:
        observation = self.transfer.probe(
            release.bulk_artifact(),
            sample_bytes=sample_bytes,
        )
        if not observation.signature_hex or not observation.signature_hex.startswith(
            "504b0304"
        ):
            raise DenverTaxSourceChanged(
                "Denver artifact no longer has an OOXML ZIP signature",
                details=observation.to_dict(),
            )
        return observation.to_dict()

    def download_verified(
        self,
        release: ReleaseLink,
        destination: Path | str,
        *,
        overwrite: bool,
        max_bytes: int | None,
        archive_policy: ArchiveSafetyPolicy,
    ) -> dict[str, Any]:
        destination_path = Path(destination).expanduser()
        if destination_path.exists() and destination_path.is_dir():
            destination_path = destination_path / release.filename
        if destination_path.exists() and not overwrite:
            raise DenverTaxQueryError(
                "destination already exists; pass --overwrite to replace it",
                details={"destination": str(destination_path)},
            )
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        descriptor, staging_name = tempfile.mkstemp(
            prefix=f".{destination_path.name}.",
            suffix=".xlsx",
            dir=destination_path.parent,
        )
        os.close(descriptor)
        staging_path = Path(staging_name)
        staging_path.unlink()
        try:
            receipt = self.transfer.download(
                release.bulk_artifact(),
                staging_path,
                resume=False,
                max_bytes=max_bytes,
            )
            with staging_path.open("rb") as source:
                signature = source.read(16)
            if not signature.startswith(b"PK\x03\x04"):
                raise DenverTaxSourceChanged(
                    "downloaded Denver artifact is not an OOXML ZIP"
                )
            inspection = inspect_workbook(
                staging_path,
                archive_policy=archive_policy,
            )
            os.replace(staging_path, destination_path)
            receipt_record = receipt.to_dict()
            receipt_record["path"] = str(destination_path.resolve())
            receipt_record["signature_hex"] = signature.hex()
            return {
                "release": release,
                "artifact_receipt": receipt_record,
                "workbook_inspection": inspection,
            }
        finally:
            staging_path.unlink(missing_ok=True)
            staging_path.with_name(
                f"{staging_path.name}.part"
            ).unlink(missing_ok=True)
            staging_path.with_name(
                f"{staging_path.name}.part.json"
            ).unlink(missing_ok=True)


def _archive_policy(args: argparse.Namespace) -> ArchiveSafetyPolicy:
    return ArchiveSafetyPolicy(
        max_members=args.max_archive_members,
        max_total_uncompressed_bytes=args.max_uncompressed_bytes,
        max_member_uncompressed_bytes=args.max_member_uncompressed_bytes,
        max_compression_ratio=args.max_compression_ratio,
    )


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _json_number(value: Any, field_name: str) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise DenverTaxSourceChanged(
            f"{field_name} is no longer numeric",
            details={"value": str(value)},
        )
    if isinstance(value, float):
        if not math.isfinite(value):
            raise DenverTaxSourceChanged(
                f"{field_name} contains a non-finite number"
            )
        rounded = round(value, 2)
        return int(rounded) if rounded.is_integer() else rounded
    return value


def _marker_year(row: Sequence[Any]) -> int | None:
    owner = _clean_text(row[0])
    match = YEAR_MARKER_RE.fullmatch(owner or "")
    if match is None:
        return None
    if any(_clean_text(value) is not None for value in row[1:]):
        raise DenverTaxSourceChanged(
            "Denver tax-year marker row contains unexpected data"
        )
    return int(match.group("year"))


def _select_worksheet(workbook: Any) -> tuple[Any, tuple[str, ...]]:
    observed: dict[str, list[str | None]] = {}
    matches: list[tuple[Any, tuple[str, ...]]] = []
    for worksheet in workbook.worksheets:
        first = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            (),
        )
        headers = tuple(_clean_text(value) or "" for value in first)
        observed[worksheet.title] = list(headers)
        if headers == EXPECTED_HEADERS:
            matches.append((worksheet, headers))
    if len(matches) != 1:
        raise DenverTaxSourceChanged(
            "Denver workbook does not contain exactly one recognized data sheet",
            details={"observed_headers": observed},
        )
    return matches[0]


def inspect_workbook(
    artifact: Path | str,
    *,
    archive_policy: ArchiveSafetyPolicy | None = None,
) -> dict[str, Any]:
    """Validate and summarize an XLSX without imposing a record ceiling."""
    path = Path(artifact).expanduser()
    if not path.is_file():
        raise DenverTaxUnavailable(
            "Denver workbook artifact is not available",
            details={"artifact": str(path)},
        )
    try:
        archive = inspect_zip(path, policy=archive_policy)
    except (ArchiveSafetyError, OSError) as error:
        if isinstance(error, ArchiveSafetyError):
            raise
        raise DenverTaxUnavailable(
            f"could not read Denver workbook: {error}",
            details={"artifact": str(path)},
        ) from error
    member_paths = {
        str(member["path"])
        for member in archive.members
        if member["kind"] == "file"
    }
    required = {"[Content_Types].xml", "xl/workbook.xml"}
    if not required.issubset(member_paths) or not any(
        item.startswith("xl/worksheets/") and item.endswith(".xml")
        for item in member_paths
    ):
        raise DenverTaxSourceChanged(
            "Denver artifact lacks required OOXML workbook members",
            details={"members": sorted(member_paths)},
        )

    workbook = None
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        worksheet, headers = _select_worksheet(workbook)
        current_year: int | None = None
        marker_rows: list[dict[str, int]] = []
        rows_by_tax_year: Counter[int] = Counter()
        stable_keys: set[tuple[int, str]] = set()
        data_rows = 0
        last_row = 1
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            last_row = row_number
            if not any(_clean_text(value) is not None for value in row):
                raise DenverTaxSourceChanged(
                    "Denver workbook contains an unexpected blank row",
                    details={"row": row_number},
                )
            marker = _marker_year(row)
            if marker is not None:
                current_year = marker
                marker_rows.append({"row": row_number, "tax_year": marker})
                continue
            if current_year is None:
                raise DenverTaxSourceChanged(
                    "Denver data row appears before a tax-year marker",
                    details={"row": row_number},
                )
            parcel_id = _clean_text(row[1])
            if not parcel_id:
                raise DenverTaxSourceChanged(
                    "Denver data row lacks a parcel ID",
                    details={"row": row_number},
                )
            for index, field_name in enumerate(EXPECTED_HEADERS[2:7], start=2):
                _json_number(row[index], field_name)
            stable_key = (current_year, parcel_id)
            if stable_key in stable_keys:
                raise DenverTaxSourceChanged(
                    "Denver workbook duplicates its tax-year and parcel key",
                    details={
                        "row": row_number,
                        "tax_year": current_year,
                        "parcel_id": parcel_id,
                    },
                )
            stable_keys.add(stable_key)
            rows_by_tax_year[current_year] += 1
            data_rows += 1
        if data_rows == 0 or not marker_rows:
            raise DenverTaxSourceChanged(
                "Denver workbook contains no tax delinquency rows"
            )
        schema = {
            "format": "xlsx",
            "headers": list(headers),
            "field_count": len(headers),
            "tax_year_marker_pattern": YEAR_MARKER_RE.pattern,
            "stable_key_fields": ["tax_year", "parcel_id"],
        }
        return {
            "source_id": SOURCE_ID,
            "record_kind": "bulk_workbook_inspection",
            "artifact_path": str(path.resolve()),
            "artifact_size": path.stat().st_size,
            "artifact_sha256": archive.archive_sha256,
            "archive": archive.to_dict(),
            "worksheet": worksheet.title,
            "worksheet_count": len(workbook.worksheets),
            "last_observed_row": last_row,
            "data_row_count": data_rows,
            "tax_year_markers": marker_rows,
            "rows_by_tax_year": {
                str(year): rows_by_tax_year[year]
                for year in sorted(rows_by_tax_year)
            },
            "schema": schema,
            "schema_fingerprint": sha256_fingerprint(schema),
            "adapter_schema_fingerprint": ADAPTER_SCHEMA_FINGERPRINT,
        }
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        TypeError,
        ValueError,
        OSError,
    ) as error:
        if isinstance(error, BulkSourceError):
            raise
        raise DenverTaxSourceChanged(
            f"Denver workbook could not be parsed: {error}",
            details={"artifact": str(path)},
        ) from error
    finally:
        if workbook is not None:
            workbook.close()


@dataclass(frozen=True)
class SearchCriteria:
    query: str | None = None
    parcel: str | None = None
    owner: str | None = None
    address: str | None = None
    tax_year: int | None = None
    tax_sale_only: bool = False
    partially_paid_only: bool = False

    @classmethod
    def from_args(cls, args: argparse.Namespace) -> "SearchCriteria":
        return cls(
            query=_clean_text(args.query),
            parcel=_clean_text(args.parcel),
            owner=_clean_text(args.owner),
            address=_clean_text(args.address),
            tax_year=args.tax_year,
            tax_sale_only=args.tax_sale_only,
            partially_paid_only=args.partially_paid_only,
        )

    def to_dict(self) -> dict[str, Any]:
        return {
            "query": self.query,
            "parcel": self.parcel,
            "owner": self.owner,
            "address": self.address,
            "tax_year": self.tax_year,
            "tax_sale_only": self.tax_sale_only,
            "partially_paid_only": self.partially_paid_only,
        }

    def cursor_identity(self) -> dict[str, Any]:
        """Return criteria normalized to the adapter's matching semantics."""

        return {
            "query": self.query.casefold() if self.query else None,
            "parcel": _parcel_key(self.parcel) if self.parcel else None,
            "owner": self.owner.casefold() if self.owner else None,
            "address": self.address.casefold() if self.address else None,
            "tax_year": self.tax_year,
            "tax_sale_only": self.tax_sale_only,
            "partially_paid_only": self.partially_paid_only,
        }


def _parcel_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())


def _matches(
    row: Sequence[Any],
    tax_year: int,
    criteria: SearchCriteria,
) -> bool:
    owner_names = [
        _clean_text(row[index])
        for index in (0, 9, 10, 11)
        if _clean_text(row[index])
    ]
    address = _clean_text(row[12])
    parcel = _clean_text(row[1]) or ""
    legal = _clean_text(row[13])
    if criteria.tax_year is not None and tax_year != criteria.tax_year:
        return False
    if (
        criteria.parcel is not None
        and _parcel_key(parcel) != _parcel_key(criteria.parcel)
    ):
        return False
    if criteria.owner is not None and not any(
        criteria.owner.casefold() in name.casefold()
        for name in owner_names
    ):
        return False
    if criteria.address is not None and (
        not address
        or criteria.address.casefold() not in address.casefold()
    ):
        return False
    if criteria.tax_sale_only and _clean_text(row[7]) != "TS":
        return False
    if (
        criteria.partially_paid_only
        and _clean_text(row[8]) != "PARTIALLY PAID"
    ):
        return False
    if criteria.query is not None:
        haystack = " ".join(
            value
            for value in [parcel, *owner_names, address, legal]
            if value
        ).casefold()
        if criteria.query.casefold() not in haystack:
            return False
    return True


def _owners(row: Sequence[Any]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for index, source_field in (
        (0, "Owner Name"),
        (9, "Additional Owner1"),
        (10, "Additional Owner2"),
        (11, "Additional Owner3"),
    ):
        name = _clean_text(row[index])
        if not name:
            continue
        result.append(
            {
                "raw_name": name,
                "role": (
                    "primary_published_tax_account_owner"
                    if index == 0
                    else "additional_published_tax_account_owner"
                ),
                "assertion_type": "tax_delinquency_list",
                "source_field": source_field,
                "title_caveat": "not_a_title_chain",
            }
        )
    return result


def _normalize_record(
    row: Sequence[Any],
    *,
    row_number: int,
    tax_year: int,
    artifact_sha256: str,
    schema_fingerprint: str,
    release: ReleaseLink | None,
) -> dict[str, Any]:
    parcel_id = _clean_text(row[1])
    assert parcel_id is not None
    stable_account_key = f"{tax_year}:{parcel_id}"
    owner_records = _owners(row)
    tax_sale_raw = _clean_text(row[7])
    partially_paid_raw = _clean_text(row[8])
    raw = {
        header: (
            _json_number(value, header)
            if 2 <= index <= 6
            else _clean_text(value)
        )
        for index, (header, value) in enumerate(zip(EXPECTED_HEADERS, row))
    }
    return {
        "source_id": SOURCE_ID,
        "record_kind": "property_tax_delinquency",
        "record_scope": "annual_tax_lien_sale_input",
        "canonical_ref": canonical_property_ref(
            SOURCE_ID,
            DENVER_GEOID,
            "tax-delinquency",
            stable_account_key,
        ),
        "evidence_ref": (
            f"DENVER-TAX:{tax_year}:{parcel_id}"
        ),
        "native_parcel_id": parcel_id,
        "native_account_id": parcel_id,
        "stable_account_key": stable_account_key,
        "tax_year": tax_year,
        "owner_names": [
            str(owner["raw_name"]) for owner in owner_records
        ],
        "owners": owner_records,
        "situs_address": {
            "raw": _clean_text(row[12]),
            "city": "Denver",
            "state": "CO",
        },
        "release_date": release.release_date if release else None,
        "release_date_basis": (
            "artifact_filename"
            if release and release.release_date
            else None
        ),
        "delinquency_status": "delinquent_as_published",
        "delinquency_category": None,
        "release_scope_categories": list(RELEASE_SCOPE_CATEGORIES),
        "amounts": {
            "total_due": _json_number(row[6], "Total Owed"),
            "tax": _json_number(row[3], "Tax Owed"),
            "interest": _json_number(row[4], "Interest"),
            "fees": _json_number(row[5], "Fees And ADV"),
            "currency": "USD",
        },
        "valuation": {
            "parcel_valuation": _json_number(
                row[2],
                "Parcel Valuation",
            ),
            "currency": "USD",
            "source_field": "Parcel Valuation",
        },
        "tax_sale_indicator": {
            "raw": tax_sale_raw,
            "marked": tax_sale_raw == "TS",
            "status": (
                "prior_tax_sale_unredeemed"
                if tax_sale_raw == "TS"
                else "not_indicated"
            ),
        },
        "partial_payment_indicator": {
            "raw": partially_paid_raw,
            "marked": partially_paid_raw == "PARTIALLY PAID",
            "status": (
                "partially_paid"
                if partially_paid_raw == "PARTIALLY PAID"
                else "not_indicated"
            ),
        },
        "legal_description_raw": _clean_text(row[13]),
        "publication_page": PUBLICATION_PAGE,
        "artifact_url": release.url if release else None,
        "artifact_sha256": artifact_sha256,
        "source_row_number": row_number,
        "workbook_schema_fingerprint": schema_fingerprint,
        "adapter_schema_fingerprint": ADAPTER_SCHEMA_FINGERPRINT,
        "raw": raw,
    }


def _criteria_cursor_fingerprint(criteria: SearchCriteria) -> str:
    return sha256_fingerprint(
        {
            "source_id": SOURCE_ID,
            "cursor_version": CURSOR_VERSION,
            "criteria": criteria.cursor_identity(),
        }
    )


def _artifact_cursor_fingerprint(
    inspection: Mapping[str, Any],
    release: ReleaseLink | None,
) -> str:
    artifact_sha256 = str(inspection.get("artifact_sha256") or "").casefold()
    if not re.fullmatch(r"[0-9a-f]{64}", artifact_sha256):
        raise DenverTaxSourceChanged(
            "Denver workbook inspection lacks an immutable artifact digest"
        )
    identity: dict[str, Any] = {
        "source_id": SOURCE_ID,
        "artifact_sha256": artifact_sha256,
        "artifact_kind": "official_release" if release else "local_artifact",
    }
    if release is not None:
        identity["release"] = {
            "artifact_id": release.artifact_id,
            "tax_year": release.tax_year,
            "url": release.url,
            "filename": release.filename,
            "release_date": release.release_date,
        }
    return sha256_fingerprint(identity)


def _parse_cursor(cursor: str) -> tuple[str, str, int]:
    match = CURSOR_RE.fullmatch(cursor.strip())
    if match is None or int(match.group("row")) < 2:
        raise DenverTaxQueryError(
            (
                "cursor must be a versioned Denver delinquent-tax "
                "continuation returned by a prior search"
            ),
            code="denver_tax_cursor_invalid",
        )
    return (
        match.group("criteria"),
        match.group("artifact"),
        int(match.group("row")),
    )


def _cursor_start(
    cursor: str | None,
    *,
    criteria: SearchCriteria,
    inspection: Mapping[str, Any],
    release: ReleaseLink | None,
) -> int:
    if cursor is None:
        return 2
    cursor_criteria, cursor_artifact, row = _parse_cursor(cursor)
    criteria_fingerprint = _criteria_cursor_fingerprint(criteria)
    if cursor_criteria != criteria_fingerprint:
        raise DenverTaxQueryError(
            "cursor belongs to different Denver delinquent-tax criteria",
            code="denver_tax_cursor_criteria_mismatch",
            details={
                "cursor_criteria_fingerprint": cursor_criteria,
                "search_criteria_fingerprint": criteria_fingerprint,
            },
        )
    artifact_fingerprint = _artifact_cursor_fingerprint(
        inspection,
        release,
    )
    if cursor_artifact != artifact_fingerprint:
        raise DenverTaxQueryError(
            (
                "cursor belongs to a different Denver delinquent-tax "
                "artifact or release"
            ),
            code="denver_tax_cursor_artifact_mismatch",
            details={
                "cursor_artifact_fingerprint": cursor_artifact,
                "search_artifact_fingerprint": artifact_fingerprint,
            },
        )
    return row


def _cursor(
    row: int,
    *,
    criteria: SearchCriteria,
    inspection: Mapping[str, Any],
    release: ReleaseLink | None,
) -> str:
    return (
        f"denver-delinquent-tax:{CURSOR_VERSION}:"
        f"criteria:{_criteria_cursor_fingerprint(criteria)}:"
        f"artifact:{_artifact_cursor_fingerprint(inspection, release)}:"
        f"row:{row}"
    )


def search_workbook(
    artifact: Path | str,
    *,
    criteria: SearchCriteria,
    inspection: Mapping[str, Any],
    release: ReleaseLink | None = None,
    max_records: int | None = None,
    cursor: str | None = None,
) -> tuple[list[dict[str, Any]], str | None]:
    """Stream matching rows; an omitted max_records returns every match."""
    path = Path(artifact).expanduser()
    start_row = _cursor_start(
        cursor,
        criteria=criteria,
        inspection=inspection,
        release=release,
    )
    workbook = None
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        worksheet, _headers = _select_worksheet(workbook)
        current_year: int | None = None
        records: list[dict[str, Any]] = []
        next_cursor: str | None = None
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            marker = _marker_year(row)
            if marker is not None:
                current_year = marker
                continue
            if current_year is None:
                raise DenverTaxSourceChanged(
                    "Denver data row appears before a tax-year marker",
                    details={"row": row_number},
                )
            if row_number < start_row or not _matches(
                row,
                current_year,
                criteria,
            ):
                continue
            if max_records is not None and len(records) >= max_records:
                next_cursor = _cursor(
                    row_number,
                    criteria=criteria,
                    inspection=inspection,
                    release=release,
                )
                break
            records.append(
                _normalize_record(
                    row,
                    row_number=row_number,
                    tax_year=current_year,
                    artifact_sha256=str(inspection["artifact_sha256"]),
                    schema_fingerprint=str(
                        inspection["schema_fingerprint"]
                    ),
                    release=release,
                )
            )
        return records, next_cursor
    except (BadZipFile, InvalidFileException, KeyError, OSError) as error:
        raise DenverTaxSourceChanged(
            f"Denver workbook search failed: {error}",
            details={"artifact": str(path)},
        ) from error
    finally:
        if workbook is not None:
            workbook.close()


def _query_parameters(args: argparse.Namespace) -> dict[str, Any]:
    if args.command == "search":
        return {
            "artifact": (
                str(args.artifact)
                if args.artifact is not None
                else "official_latest"
            ),
            "criteria": SearchCriteria.from_args(args).to_dict(),
        }
    if args.command == "inspect":
        return {"artifact": str(args.artifact)}
    if args.command == "download":
        return {"destination": str(args.destination)}
    if args.command == "probe":
        return {"sample_bytes": args.sample_bytes}
    return {"publication_page": PUBLICATION_PAGE}


def build_query(
    args: argparse.Namespace,
    *,
    access_decision: Mapping[str, Any] | None = None,
) -> PublicRecordsQuery:
    return PublicRecordsQuery(
        source=SOURCE_METADATA,
        jurisdiction=JURISDICTION,
        query=QueryMetadata(
            operation=args.command,
            parameters=_query_parameters(args),
            requested_limit=(
                args.max_records if args.command == "search" else None
            ),
            cursor=(args.cursor if args.command == "search" else None),
            metadata={
                "access_decision": dict(access_decision or {}),
                "adapter_schema_fingerprint": ADAPTER_SCHEMA_FINGERPRINT,
            },
        ),
    )


def _access_contract(args: argparse.Namespace) -> dict[str, Any]:
    catalog = ensure_catalog_source(
        SOURCE_ID,
        db_path=Path(args.catalog_db).expanduser(),
        config_path=Path(args.catalog_config).expanduser(),
    )
    return catalog.require_machine_acquisition(SOURCE_ID)


def _access_failure(
    args: argparse.Namespace,
    error: Exception,
) -> PublicRecordsResult:
    if isinstance(error, AcquisitionUnavailableError):
        decision = dict(error.decision)
        status = ResultStatus(acquisition_result_status(decision))
        code = str(
            decision.get("reason_code")
            or "machine_acquisition_unavailable"
        )
        message = str(decision.get("reason") or error)
    else:
        decision = {}
        status = ResultStatus.UNAVAILABLE
        code = "catalog_unavailable"
        message = str(error)
    return PublicRecordsResult.failure(
        build_query(args, access_decision=decision),
        status,
        [
            PublicRecordsError(
                code=code,
                message=message,
                category="source_access",
                retryable=False,
                details={"access_decision": decision},
            )
        ],
        warnings=SOURCE_WARNINGS,
    )


def _source_failure(
    query: PublicRecordsQuery,
    error: BulkSourceError,
) -> PublicRecordsResult:
    return PublicRecordsResult.failure(
        query,
        ResultStatus(error.result_status),
        [
            PublicRecordsError(
                code=error.code,
                message=_sanitize_transient_artifact_text(str(error)),
                category=error.category,
                retryable=error.retryable,
                details=_sanitize_transient_artifact_values(
                    error.details
                ),
            )
        ],
        warnings=SOURCE_WARNINGS,
    )


def _client(
    args: argparse.Namespace,
    access_decision: Mapping[str, Any],
) -> DenverDelinquentTaxClient:
    limits = access_decision.get("limits")
    catalog_interval = 0.0
    if isinstance(limits, Mapping):
        raw_interval = limits.get("minimum_interval_seconds")
        if isinstance(raw_interval, (int, float)) and not isinstance(
            raw_interval,
            bool,
        ):
            catalog_interval = max(0.0, float(raw_interval))
    return DenverDelinquentTaxClient(
        timeout=args.timeout,
        max_attempts=args.retry_attempts,
        minimum_interval=max(args.minimum_interval, catalog_interval),
        chunk_size=args.chunk_size,
    )


def _log(query: PublicRecordsQuery, count: int | None) -> None:
    try:
        log_search(canonical_json(query.to_dict()), SOURCE_ID, count)
    except Exception as error:
        print(f"Warning: could not log search: {error}", file=sys.stderr)


def _sanitize_transient_artifact_text(value: str) -> str:
    return TRANSIENT_ARTIFACT_PATH_RE.sub(
        "[temporary artifact removed]",
        value,
    )


def _sanitize_transient_artifact_values(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {
            str(key): _sanitize_transient_artifact_values(item)
            for key, item in value.items()
        }
    if isinstance(value, (list, tuple)):
        return [
            _sanitize_transient_artifact_values(item)
            for item in value
        ]
    if isinstance(value, os.PathLike):
        return _sanitize_transient_artifact_text(os.fspath(value))
    if isinstance(value, str):
        return _sanitize_transient_artifact_text(value)
    return value


def _durable_probe_evidence(
    payload: Mapping[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Remove paths whose TemporaryDirectory has already been deleted."""
    receipt = dict(payload["artifact_receipt"])
    receipt.pop("path", None)
    inspection = json.loads(
        json.dumps(payload["workbook_inspection"])
    )
    inspection.pop("artifact_path", None)
    archive = inspection.get("archive")
    if isinstance(archive, dict):
        archive.pop("path", None)
    return receipt, inspection


def execute(
    args: argparse.Namespace,
    *,
    access_decision: Mapping[str, Any] | None = None,
    client: DenverDelinquentTaxClient | Any | None = None,
    log_results: bool = True,
) -> PublicRecordsResult:
    """Execute one operation through the shared public-record result envelope."""
    try:
        decision = (
            dict(access_decision)
            if access_decision is not None
            else _access_contract(args)
        )
    except (
        AcquisitionUnavailableError,
        CatalogError,
        OSError,
        ValueError,
    ) as error:
        result = _access_failure(args, error)
        if log_results and args.command != "probe":
            _log(result.query, None)
        return result
    if not decision.get("allowed", False):
        result = _access_failure(
            args,
            AcquisitionUnavailableError(decision),
        )
        if log_results and args.command != "probe":
            _log(result.query, None)
        return result

    query = build_query(args, access_decision=decision)
    source_client = client or _client(args, decision)
    raw_refs: tuple[str, ...] = ()
    try:
        if args.command == "discover":
            release = source_client.discover()
            result = PublicRecordsResult.success(
                query,
                [release.to_record()],
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "inspect":
            inspection = inspect_workbook(
                args.artifact,
                archive_policy=_archive_policy(args),
            )
            raw_refs = (str(Path(args.artifact).resolve()),)
            result = PublicRecordsResult.success(
                query,
                [inspection],
                raw_artifact_refs=raw_refs,
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "download":
            release = source_client.discover()
            payload = source_client.download_verified(
                release,
                args.destination,
                overwrite=args.overwrite,
                max_bytes=args.max_download_bytes,
                archive_policy=_archive_policy(args),
            )
            record = release.to_record()
            record["artifact_receipt"] = payload["artifact_receipt"]
            record["workbook_inspection"] = payload[
                "workbook_inspection"
            ]
            raw_refs = (payload["artifact_receipt"]["path"],)
            result = PublicRecordsResult.success(
                query,
                [record],
                raw_artifact_refs=raw_refs,
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "probe":
            release = source_client.discover()
            probe = source_client.probe(
                release,
                sample_bytes=args.sample_bytes,
            )
            with tempfile.TemporaryDirectory(
                prefix="osint-denver-tax-",
                dir="/tmp",
            ) as workdir:
                destination = Path(workdir) / release.filename
                payload = source_client.download_verified(
                    release,
                    destination,
                    overwrite=False,
                    max_bytes=args.max_download_bytes,
                    archive_policy=_archive_policy(args),
                )
            receipt, inspection = _durable_probe_evidence(payload)
            result = PublicRecordsResult.success(
                query,
                [
                    {
                        "canonical_ref": canonical_property_ref(
                            SOURCE_ID,
                            DENVER_GEOID,
                            "source-health",
                            "live-release-probe",
                        ),
                        "source_id": SOURCE_ID,
                        "record_kind": "source_health_check",
                        "native_document_id": "live-release-probe",
                        "source_url": PUBLICATION_PAGE,
                        "status": "ok",
                        "release": release.to_record(),
                        "artifact_probe": probe,
                        "artifact_receipt": receipt,
                        "workbook_inspection": inspection,
                    }
                ],
                raw_artifact_refs=(release.url,),
                warnings=SOURCE_WARNINGS,
            )
        elif args.command == "search":
            release: ReleaseLink | None = None
            if args.artifact is not None:
                artifact_path = Path(args.artifact).expanduser()
                inspection = inspect_workbook(
                    artifact_path,
                    archive_policy=_archive_policy(args),
                )
                raw_refs = (str(artifact_path.resolve()),)
                records, next_cursor = search_workbook(
                    artifact_path,
                    criteria=SearchCriteria.from_args(args),
                    inspection=inspection,
                    max_records=args.max_records,
                    cursor=args.cursor,
                )
            else:
                release = source_client.discover()
                with tempfile.TemporaryDirectory(
                    prefix="osint-denver-tax-",
                    dir="/tmp",
                ) as workdir:
                    artifact_path = Path(workdir) / release.filename
                    payload = source_client.download_verified(
                        release,
                        artifact_path,
                        overwrite=False,
                        max_bytes=args.max_download_bytes,
                        archive_policy=_archive_policy(args),
                    )
                    inspection = payload["workbook_inspection"]
                    records, next_cursor = search_workbook(
                        artifact_path,
                        criteria=SearchCriteria.from_args(args),
                        inspection=inspection,
                        release=release,
                        max_records=args.max_records,
                        cursor=args.cursor,
                    )
                raw_refs = (release.url,)
            if next_cursor is not None:
                result = PublicRecordsResult(
                    query=query,
                    status=ResultStatus.PARTIAL,
                    records=records,
                    next_cursor=next_cursor,
                    raw_artifact_refs=raw_refs,
                    warnings=SOURCE_WARNINGS,
                )
            else:
                result = PublicRecordsResult.success(
                    query,
                    records,
                    raw_artifact_refs=raw_refs,
                    warnings=SOURCE_WARNINGS,
                )
        else:
            raise DenverTaxQueryError(
                f"unsupported command: {args.command}"
            )
    except BulkSourceError as error:
        result = _source_failure(query, error)
    except OSError as error:
        result = PublicRecordsResult.failure(
            query,
            ResultStatus.UNAVAILABLE,
            [
                PublicRecordsError(
                    code="local_io_failed",
                    message=_sanitize_transient_artifact_text(str(error)),
                    category="local_io",
                    retryable=False,
                )
            ],
            warnings=SOURCE_WARNINGS,
        )

    count = (
        len(result.records)
        if result.status
        in {
            ResultStatus.OK,
            ResultStatus.NO_RESULTS,
            ResultStatus.PARTIAL,
        }
        else None
    )
    # Health probes are deliberately excluded from investigative search logs.
    if log_results and args.command != "probe":
        _log(query, count)
    return result


def _emit(result: PublicRecordsResult, args: argparse.Namespace) -> None:
    payload = result.to_dict()
    if write_output(
        payload,
        args,
        summary=(
            f"Denver delinquent tax {args.command} "
            f"({result.status.value})"
        ),
    ):
        return
    if args.json_out:
        print(json.dumps(payload, indent=2, sort_keys=True))
        return
    print(
        f"Denver delinquent tax {args.command}: {result.status.value} "
        f"({len(result.records)} records)"
    )
    if result.next_cursor:
        print(f"Next cursor: {result.next_cursor}")
    for record in result.records:
        if record.get("record_kind") == "property_tax_delinquency":
            print(
                f"- {record['tax_year']} | "
                f"{record['native_parcel_id']} | "
                f"{record['amounts']['total_due']} | "
                f"{', '.join(record['owner_names'])}"
            )
    for warning in result.warnings:
        print(f"Warning: {warning}", file=sys.stderr)
    for error in result.errors:
        print(f"ERROR [{error.code}]: {error.message}", file=sys.stderr)


def _add_archive_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--max-archive-members",
        type=int,
        default=DEFAULT_MAX_ARCHIVE_MEMBERS,
    )
    parser.add_argument(
        "--max-uncompressed-bytes",
        type=int,
        default=DEFAULT_MAX_UNCOMPRESSED_BYTES,
    )
    parser.add_argument(
        "--max-member-uncompressed-bytes",
        type=int,
        default=DEFAULT_MAX_MEMBER_UNCOMPRESSED_BYTES,
    )
    parser.add_argument(
        "--max-compression-ratio",
        type=float,
        default=DEFAULT_MAX_COMPRESSION_RATIO,
    )


def _add_runtime_args(
    parser: argparse.ArgumentParser,
    *,
    archive: bool = False,
) -> None:
    parser.add_argument("--timeout", type=float, default=60.0)
    parser.add_argument("--minimum-interval", type=float, default=0.0)
    parser.add_argument("--retry-attempts", type=int, default=3)
    parser.add_argument("--chunk-size", type=int, default=1024 * 1024)
    parser.add_argument(
        "--max-download-bytes",
        type=int,
        help="Optional caller-selected maximum artifact download size",
    )
    parser.add_argument(
        "--catalog-db",
        default=str(DEFAULT_CATALOG_DB_PATH),
    )
    parser.add_argument(
        "--catalog-config",
        default=str(DEFAULT_CATALOG_CONFIG_PATH),
    )
    if archive:
        _add_archive_args(parser)
    add_output_args(parser)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Discover, verify, and search Denver Treasury's official "
            "delinquent-real-property-tax workbook"
        )
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    discover = subparsers.add_parser(
        "discover",
        help="Discover the latest official workbook release",
    )
    _add_runtime_args(discover)

    probe = subparsers.add_parser(
        "probe",
        help="Verify release discovery, transfer, and workbook schema",
    )
    probe.add_argument(
        "--sample-bytes",
        type=int,
        default=DEFAULT_SAMPLE_BYTES,
    )
    _add_runtime_args(probe, archive=True)

    inspect_parser = subparsers.add_parser(
        "inspect",
        help="Inspect a local Denver workbook",
    )
    inspect_parser.add_argument("artifact", type=Path)
    _add_runtime_args(inspect_parser, archive=True)

    download = subparsers.add_parser(
        "download",
        help="Download and verify the latest official workbook",
    )
    download.add_argument("--destination", type=Path, required=True)
    download.add_argument("--overwrite", action="store_true")
    _add_runtime_args(download, archive=True)

    search = subparsers.add_parser(
        "search",
        help=(
            "Stream the latest official workbook or an explicit local artifact"
        ),
    )
    search.add_argument(
        "--artifact",
        type=Path,
        help=(
            "Optional local XLSX; omitted discovers and fetches the latest "
            "official workbook"
        ),
    )
    search.add_argument("--query", help="Broad text match")
    search.add_argument(
        "--parcel",
        "--account",
        dest="parcel",
        help="Exact published parcel/account identifier",
    )
    search.add_argument("--owner", help="Published owner-name substring")
    search.add_argument("--address", help="Published parcel-address substring")
    search.add_argument("--tax-year", type=int)
    search.add_argument("--tax-sale-only", action="store_true")
    search.add_argument("--partially-paid-only", action="store_true")
    search.add_argument(
        "--max-records",
        type=int,
        help="Optional caller-selected result ceiling",
    )
    search.add_argument(
        "--cursor",
        help="Continuation cursor from a previous partial result",
    )
    _add_runtime_args(search, archive=True)
    return parser


def _validate_args(
    parser: argparse.ArgumentParser,
    args: argparse.Namespace,
) -> None:
    positive = (
        "timeout",
        "retry_attempts",
        "chunk_size",
        "max_download_bytes",
        "max_archive_members",
        "max_uncompressed_bytes",
        "max_member_uncompressed_bytes",
        "max_compression_ratio",
        "sample_bytes",
        "max_records",
    )
    if args.minimum_interval < 0:
        parser.error("--minimum-interval must not be negative")
    for name in positive:
        value = getattr(args, name, None)
        if value is not None and value <= 0:
            parser.error(f"--{name.replace('_', '-')} must be positive")
    if args.command == "probe" and args.sample_bytes < 4:
        parser.error("--sample-bytes must be at least 4")
    if args.command == "search" and args.tax_year is not None:
        if args.tax_year < 1900 or args.tax_year > 2200:
            parser.error("--tax-year must be between 1900 and 2200")
    if args.command == "search" and args.cursor is not None:
        try:
            _parse_cursor(args.cursor)
        except DenverTaxQueryError as error:
            parser.error(str(error))


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    _validate_args(parser, args)
    result = execute(args)
    _emit(result, args)
    return (
        0
        if result.status
        in {
            ResultStatus.OK,
            ResultStatus.NO_RESULTS,
            ResultStatus.PARTIAL,
        }
        else 1
    )


if __name__ == "__main__":
    raise SystemExit(main())
