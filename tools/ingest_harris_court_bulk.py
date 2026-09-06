#!/usr/bin/env python3
"""Stream Harris District Clerk text extracts into the court sidecar.

The public-dataset query adapter stops at artifact discovery and verified
download receipts. This tool is the explicit byte-parsing stage for the
header-bearing civil case-summary, party, and activity files and the current
criminal filing and disposition files.

Each source row is retained as its own occurrence before any projection. The
generic court tables receive only source-visible case shells, parties,
attorneys, activities, settings, charges, and dispositions. Bulk extracts do
not establish that a filing image is available.

Examples:
    uv run python tools/ingest_harris_court_bulk.py ingest \
      /tmp/CaseSummaryMods_Daily-2026-07-30.txt \
      --native-locator 'Civil\\CaseSummaryMods_Daily-2026-07-30.txt' \
      --published-date 2026-07-30 \
      --schema-workbook /tmp/FIELD_CODES.xlsx \
      --court-db /tmp/harris-courts.db

    uv run python tools/ingest_harris_court_bulk.py ingest \
      /tmp/CrimFilingsDaily_withHeadings.txt \
      --artifact-result /tmp/harris-download.json \
      --schema-workbook /tmp/RecordLayoutsAndFieldNames.xlsx
"""

from __future__ import annotations

import argparse
import codecs
import csv
import hashlib
import json
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook

try:
    from tools.ingest_state_court_records import (
        _ensure_docket_hearing_columns,
        _project_case,
    )
    from tools.output_util import add_output_args, write_output
    from tools.public_records_contract import canonical_json, sha256_fingerprint
    from tools.public_records_store import DEFAULT_COURT_DB, connect_courts
    from tools.query_harris_court_bulk import (
        CATALOG_URL,
        SOURCE_ID,
        _dataset_cadence,
        _dataset_family,
    )
except ImportError:
    from ingest_state_court_records import (
        _ensure_docket_hearing_columns,
        _project_case,
    )
    from output_util import add_output_args, write_output
    from public_records_contract import canonical_json, sha256_fingerprint
    from public_records_store import DEFAULT_COURT_DB, connect_courts
    from query_harris_court_bulk import (
        CATALOG_URL,
        SOURCE_ID,
        _dataset_cadence,
        _dataset_family,
    )


CSV_FIELD_SIZE_LIMIT = 16 * 1024 * 1024
SHA256_RE = re.compile(r"^[0-9a-f]{64}$", re.IGNORECASE)
ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
ISO_MONTH_RE = re.compile(r"^\d{4}-\d{2}$")
ISO_YEAR_RE = re.compile(r"^\d{4}$")
COMPACT_YYYYMMDD_RE = re.compile(r"^\d{8}$")
COMPACT_MMDDYY_RE = re.compile(r"^\d{6}$")
NULL_LITERALS = frozenset({"NULL", "N/A"})
UNDEFINED_CP1252_BYTES = frozenset({0x81, 0x8D, 0x8F, 0x90, 0x9D})
CP1252_PRESERVE_ERROR_HANDLER = "harris_preserve_undefined_cp1252"

SUPPORTED_FAMILIES: dict[tuple[str, str], frozenset[str]] = {
    (
        "Civil",
        "case_summary",
    ): frozenset(
        {
            "rundate",
            "region",
            "casenbr",
            "file_crt",
            "curr_crt",
            "file_dt",
            "cst",
            "cs_typ",
            "toac",
            "plaintiff",
            "defendant",
            "pj_num",
            "judgment_dt",
            "judgment",
            "lastmoddt",
        }
    ),
    (
        "Civil",
        "party",
    ): frozenset(
        {
            "rundate",
            "region",
            "casenbr",
            "pj_num",
            "coc",
            "person_num",
            "conn_num",
            "active_ind",
            "prose_ind",
            "party_bar",
            "party_name",
            "assoc_aty_bar",
            "assoc_aty_name",
            "lastmoddt",
        }
    ),
    (
        "Civil",
        "activity",
    ): frozenset(
        {
            "rundate",
            "region",
            "casenbr",
            "pj_num",
            "file_crt",
            "code",
            "desc",
            "act/file_dt",
            "pgs",
            "seq_num",
            "fildedby",
            "lastmoddt",
            "case_type",
        }
    ),
    (
        "Criminal",
        "filings",
    ): frozenset(
        {
            "rundate",
            "cdi",
            "cas",
            "fda",
            "ins",
            "cad",
            "crt",
            "cst",
            "dst",
            "bam",
            "curr_off",
            "curr_off_lit",
            "curr_l_d",
            "nda",
            "cnc",
            "rea",
            "def_nam",
            "def_spn",
            "def_rac",
            "def_sex",
            "def_dob",
            "def_stnum",
            "def_stnam",
            "def_cty",
            "def_st",
            "def_zip",
            "aty_nam",
            "aty_spn",
            "aty_coc",
            "aty_coc_lit",
            "def_birthplace",
            "def_uscitizen",
            "bamexp",
        }
    ),
    (
        "Criminal",
        "dispositions",
    ): frozenset(
        {
            "rundate",
            "cdi",
            "cas",
            "fda",
            "ins",
            "cad",
            "crt",
            "cst",
            "dst",
            "bam",
            "curr_off",
            "curr_off_lit",
            "curr_l_d",
            "nda",
            "cnc",
            "rea",
            "def_nam",
            "def_spn",
            "def_rac",
            "def_sex",
            "def_dob",
            "def_stnum",
            "def_stnam",
            "def_cty",
            "def_st",
            "def_zip",
            "aty_nam",
            "aty_spn",
            "aty_coc",
            "aty_coc_lit",
            "dispdt",
            "disposition",
            "sentence",
            "comp_nam",
            "comp_agency",
            "off_rpt_num",
            "bamexp",
        }
    ),
}

GENERIC_COUNT_KEYS = (
    "courts",
    "related_courts",
    "cases",
    "related_cases",
    "case_relations",
    "parties",
    "attorneys",
    "representations",
    "judicial_officers",
    "assignments",
    "claims",
    "docket_entries",
    "case_events",
    "documents",
    "restriction_events",
)

HARRIS_EXTENSION_SCHEMA = """
CREATE TABLE IF NOT EXISTS harris_court_bulk_artifact (
    artifact_id TEXT PRIMARY KEY,
    source_id TEXT NOT NULL,
    snapshot_id INTEGER NOT NULL
        REFERENCES source_snapshot(snapshot_id) ON DELETE CASCADE,
    native_locator TEXT NOT NULL,
    published_date TEXT NOT NULL,
    section TEXT NOT NULL,
    dataset_family TEXT NOT NULL,
    cadence TEXT NOT NULL,
    artifact_sha256 TEXT NOT NULL,
    artifact_size_bytes INTEGER NOT NULL,
    artifact_path TEXT NOT NULL,
    source_url TEXT NOT NULL,
    retrieved_at TEXT NOT NULL,
    header_json TEXT NOT NULL,
    header_fingerprint TEXT NOT NULL,
    schema_workbook_sha256 TEXT,
    schema_workbook_path TEXT,
    UNIQUE(source_id, native_locator, artifact_sha256)
);
CREATE INDEX IF NOT EXISTS idx_harris_bulk_artifact_release
    ON harris_court_bulk_artifact(
        section, dataset_family, cadence, published_date
    );

CREATE TABLE IF NOT EXISTS harris_court_bulk_row (
    row_occurrence_id TEXT PRIMARY KEY,
    artifact_id TEXT NOT NULL
        REFERENCES harris_court_bulk_artifact(artifact_id) ON DELETE CASCADE,
    source_row_number INTEGER NOT NULL,
    row_sha256 TEXT NOT NULL,
    case_number TEXT,
    case_id INTEGER REFERENCES case_record(case_id) ON DELETE SET NULL,
    projection_status TEXT NOT NULL,
    projection_reason TEXT,
    canonical_ref TEXT,
    raw_fields_json TEXT NOT NULL,
    normalized_json TEXT NOT NULL,
    projection_refs_json TEXT NOT NULL,
    UNIQUE(artifact_id, source_row_number)
);
CREATE INDEX IF NOT EXISTS idx_harris_bulk_row_case
    ON harris_court_bulk_row(case_number, case_id);
CREATE INDEX IF NOT EXISTS idx_harris_bulk_row_sha
    ON harris_court_bulk_row(row_sha256);
"""


class HarrisCourtBulkIngestError(ValueError):
    """A downloaded Harris artifact does not match its declared source shape."""


def _preserve_undefined_cp1252(
    error: UnicodeDecodeError,
) -> tuple[str, int]:
    """Map undefined CP1252 bytes to same-valued Unicode control points."""

    if not isinstance(error, UnicodeDecodeError):
        raise error
    preserved = "".join(
        chr(value)
        for value in error.object[error.start : error.end]
    )
    return preserved, error.end


codecs.register_error(
    CP1252_PRESERVE_ERROR_HANDLER,
    _preserve_undefined_cp1252,
)


@dataclass(frozen=True)
class ArtifactContext:
    """Exact source and local-byte identity for one downloaded extract."""

    path: Path
    native_locator: str
    published_date: str
    section: str
    family: str
    cadence: str
    artifact_sha256: str
    artifact_size_bytes: int
    artifact_id: str
    source_url: str
    retrieved_at: str
    schema_workbook_path: Path | None
    schema_workbook_sha256: str | None

    def public_record(self) -> dict[str, Any]:
        return {
            "source_id": SOURCE_ID,
            "native_locator": self.native_locator,
            "published_date": self.published_date,
            "section": self.section,
            "dataset_family": self.family,
            "cadence": self.cadence,
            "artifact_sha256": self.artifact_sha256,
            "artifact_size_bytes": self.artifact_size_bytes,
            "artifact_id": self.artifact_id,
            "artifact_path": str(self.path),
            "source_url": self.source_url,
            "retrieved_at": self.retrieved_at,
            "schema_workbook_path": (
                str(self.schema_workbook_path)
                if self.schema_workbook_path is not None
                else None
            ),
            "schema_workbook_sha256": self.schema_workbook_sha256,
        }


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _undefined_cp1252_counts(path: Path) -> dict[str, int]:
    counts = {value: 0 for value in UNDEFINED_CP1252_BYTES}
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            for value in UNDEFINED_CP1252_BYTES:
                counts[value] += chunk.count(value)
    return {
        f"0x{value:02x}": count
        for value, count in sorted(counts.items())
        if count
    }


def _text(value: Any) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    if not normalized or normalized.upper() in NULL_LITERALS:
        return None
    return normalized


def _normalized_name(value: Any) -> str | None:
    text = _text(value)
    return " ".join(text.upper().split()) if text else None


def _stable_token(*values: Any, length: int = 24) -> str:
    payload = "\x1f".join("" if value is None else str(value) for value in values)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length]


def _stable_sequence(*values: Any) -> int:
    for value in values:
        text = _text(value)
        if text is not None and text.isdigit():
            return int(text)
    return int(_stable_token(*values, length=8), 16)


def _parse_date(value: Any) -> dict[str, Any]:
    raw = None if value is None else str(value)
    normalized = raw.strip() if raw is not None else ""
    observation: dict[str, Any] = {
        "raw": raw,
        "value": normalized or None,
        "iso": None,
        "precision": None,
        "source_format": None,
    }
    if not normalized:
        observation["state"] = "blank"
        return observation
    if normalized.upper() in NULL_LITERALS:
        observation["state"] = "source_null_literal"
        return observation

    candidate = normalized
    if " " in normalized and ISO_DATE_RE.fullmatch(normalized.split(" ", 1)[0]):
        candidate = normalized.split(" ", 1)[0]
        observation["source_format"] = "iso_datetime"
    elif "T" in normalized and ISO_DATE_RE.fullmatch(normalized.split("T", 1)[0]):
        candidate = normalized.split("T", 1)[0]
        observation["source_format"] = "iso_datetime"

    try:
        if ISO_DATE_RE.fullmatch(candidate):
            parsed = datetime.strptime(candidate, "%Y-%m-%d")
            observation.update(
                iso=parsed.date().isoformat(),
                precision="day",
                source_format=observation["source_format"] or "iso_date",
                state="parsed",
            )
        elif COMPACT_YYYYMMDD_RE.fullmatch(candidate):
            parsed = datetime.strptime(candidate, "%Y%m%d")
            observation.update(
                iso=parsed.date().isoformat(),
                precision="day",
                source_format="yyyymmdd",
                state="parsed",
            )
        elif COMPACT_MMDDYY_RE.fullmatch(candidate):
            parsed = datetime.strptime(candidate, "%m%d%y")
            observation.update(
                iso=parsed.date().isoformat(),
                precision="day",
                source_format="mmddyy_current_layout",
                state="parsed",
            )
        elif ISO_MONTH_RE.fullmatch(candidate):
            datetime.strptime(candidate, "%Y-%m")
            observation.update(
                iso=candidate,
                precision="month",
                source_format="iso_month",
                state="partial",
            )
        elif ISO_YEAR_RE.fullmatch(candidate):
            observation.update(
                iso=candidate,
                precision="year",
                source_format="iso_year",
                state="partial",
            )
        else:
            observation["state"] = "unparsed"
    except ValueError:
        observation["state"] = "unparsed"
    return observation


def _projectable_date(observation: Mapping[str, Any]) -> str | None:
    if observation.get("precision") != "day":
        return None
    return _text(observation.get("iso"))


def _normalize_header(fieldnames: Sequence[Any] | None) -> tuple[list[str], list[str]]:
    raw = [str(value) for value in (fieldnames or [])]
    normalized = [
        value.lstrip("\ufeff").strip().casefold()
        for value in raw
    ]
    if not normalized or any(not value for value in normalized):
        raise HarrisCourtBulkIngestError(
            "Harris text extract has an empty or blank header column"
        )
    duplicates = sorted(
        {
            field
            for field in normalized
            if normalized.count(field) > 1
        }
    )
    if duplicates:
        raise HarrisCourtBulkIngestError(
            "Harris text extract has duplicate normalized columns: "
            + ", ".join(duplicates)
        )
    return raw, normalized


def _row_by_normalized_header(
    raw_row: Mapping[Any, Any],
    raw_header: Sequence[str],
    normalized_header: Sequence[str],
) -> tuple[dict[str, str | None], list[str]]:
    row = {
        normalized: raw_row.get(raw)
        for raw, normalized in zip(raw_header, normalized_header, strict=True)
    }
    extras = [
        str(value)
        for value in (raw_row.get(None) or [])
    ]
    return row, extras


def _load_codebooks(
    path: Path | None,
) -> dict[str, dict[str, tuple[str, ...]]]:
    if path is None:
        return {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise HarrisCourtBulkIngestError(
            f"cannot read schema workbook {path}: {error}"
        ) from error
    codebooks: dict[str, dict[str, tuple[str, ...]]] = {}
    try:
        for worksheet in workbook.worksheets:
            sheet_name = worksheet.title.strip().casefold()
            rows = worksheet.iter_rows(values_only=True)
            next(rows, None)
            values: dict[str, tuple[str, ...]] = {}
            for row in rows:
                if not row:
                    continue
                code = _text(row[0])
                literals = tuple(
                    literal
                    for value in row[1:]
                    if (literal := _text(value)) is not None
                )
                if code is not None and literals:
                    values[code.casefold()] = literals
            if values:
                codebooks[sheet_name] = values
    finally:
        workbook.close()
    return codebooks


def _codebook_values(
    codebooks: Mapping[str, Mapping[str, tuple[str, ...]]],
    sheet: str,
    code: Any,
) -> tuple[str, ...]:
    value = _text(code)
    if value is None:
        return ()
    sheet_values = codebooks.get(sheet.casefold(), {})
    candidates = [value.casefold()]
    if value.isdigit():
        candidates.append(str(int(value)))
    for candidate in candidates:
        result = sheet_values.get(candidate)
        if result:
            return result
    return ()


def _code_observation(
    row: Mapping[str, Any],
    *,
    code_field: str,
    codebooks: Mapping[str, Mapping[str, tuple[str, ...]]],
    sheet: str | None = None,
    literal_field: str | None = None,
) -> dict[str, Any]:
    code_raw = row.get(code_field)
    literal_raw = row.get(literal_field) if literal_field else None
    values = _codebook_values(
        codebooks,
        sheet or code_field,
        code_raw,
    )
    return {
        "code_raw": code_raw,
        "code": _text(code_raw),
        "source_literal_raw": literal_raw,
        "source_literal": _text(literal_raw),
        "codebook_literals": list(values),
        "codebook_literal": values[-1] if values else None,
    }


def _code_label(observation: Mapping[str, Any]) -> str | None:
    return (
        _text(observation.get("source_literal"))
        or _text(observation.get("codebook_literal"))
        or _text(observation.get("code"))
    )


def _artifact_result_record(
    artifact_result_path: Path,
) -> tuple[Mapping[str, Any], Mapping[str, Any]]:
    try:
        payload = json.loads(artifact_result_path.read_text())
    except (OSError, json.JSONDecodeError) as error:
        raise HarrisCourtBulkIngestError(
            f"cannot read artifact result {artifact_result_path}: {error}"
        ) from error
    if not isinstance(payload, Mapping):
        raise HarrisCourtBulkIngestError("artifact result must be a JSON object")
    records = payload.get("records")
    if not isinstance(records, list) or len(records) != 1:
        raise HarrisCourtBulkIngestError(
            "artifact result must contain exactly one downloaded artifact record"
        )
    record = records[0]
    if not isinstance(record, Mapping):
        raise HarrisCourtBulkIngestError("artifact result record must be an object")
    receipt = record.get("artifact_receipt")
    if not isinstance(receipt, Mapping):
        raise HarrisCourtBulkIngestError(
            "artifact result record lacks an artifact_receipt"
        )
    if _text(record.get("source_id")) != SOURCE_ID:
        raise HarrisCourtBulkIngestError(
            "artifact result is not from the Harris District Clerk source"
        )
    return record, receipt


def _artifact_context(args: argparse.Namespace) -> ArtifactContext:
    path = Path(args.artifact).expanduser().resolve()
    if not path.is_file():
        raise HarrisCourtBulkIngestError(
            f"artifact path is not a readable file: {path}"
        )
    artifact_sha256 = _sha256_path(path)
    size_bytes = path.stat().st_size

    record: Mapping[str, Any] = {}
    receipt: Mapping[str, Any] = {}
    if args.artifact_result:
        record, receipt = _artifact_result_record(
            Path(args.artifact_result).expanduser().resolve()
        )
        receipt_path = _text(receipt.get("path"))
        if receipt_path and Path(receipt_path).expanduser().resolve() != path:
            raise HarrisCourtBulkIngestError(
                "artifact path conflicts with the download receipt path"
            )
        receipt_sha = _text(receipt.get("sha256"))
        if receipt_sha and receipt_sha.casefold() != artifact_sha256:
            raise HarrisCourtBulkIngestError(
                "artifact bytes do not match the download receipt SHA-256"
            )
        receipt_size = receipt.get("size")
        if receipt_size is not None and int(receipt_size) != size_bytes:
            raise HarrisCourtBulkIngestError(
                "artifact bytes do not match the download receipt size"
            )

    native_locator = (
        _text(args.native_locator)
        or _text(record.get("native_locator"))
        or _text(record.get("native_document_id"))
    )
    if native_locator is None:
        raise HarrisCourtBulkIngestError(
            "--native-locator or --artifact-result is required"
        )
    if "\\" not in native_locator:
        raise HarrisCourtBulkIngestError(
            "native locator must retain the source section prefix"
        )
    section, native_filename = native_locator.split("\\", 1)
    section = section.strip().title()
    if section not in {"Civil", "Criminal"} or not native_filename.strip():
        raise HarrisCourtBulkIngestError(
            "native locator must identify a Civil or Criminal catalog member"
        )
    family = _dataset_family(native_filename)
    cadence = _dataset_cadence(native_filename)
    if (section, family) not in SUPPORTED_FAMILIES:
        supported = ", ".join(
            f"{source_section}/{source_family}"
            for source_section, source_family in SUPPORTED_FAMILIES
        )
        raise HarrisCourtBulkIngestError(
            f"{section}/{family} is not implemented by this ingest wave; "
            f"implemented families: {supported}"
        )

    published_date = _text(args.published_date) or _text(
        record.get("published_date")
    )
    if published_date is None:
        raise HarrisCourtBulkIngestError(
            "--published-date or --artifact-result is required"
        )
    try:
        published_date = datetime.strptime(
            published_date,
            "%Y-%m-%d",
        ).date().isoformat()
    except ValueError as error:
        raise HarrisCourtBulkIngestError(
            "published date must use YYYY-MM-DD"
        ) from error

    expected_sha256 = _text(args.expected_sha256)
    if expected_sha256 and expected_sha256.casefold() != artifact_sha256:
        raise HarrisCourtBulkIngestError(
            "artifact SHA-256 does not match --expected-sha256"
        )

    schema_path = (
        Path(args.schema_workbook).expanduser().resolve()
        if args.schema_workbook
        else None
    )
    if schema_path is not None and not schema_path.is_file():
        raise HarrisCourtBulkIngestError(
            f"schema workbook is not a readable file: {schema_path}"
        )
    schema_sha256 = _sha256_path(schema_path) if schema_path else None
    source_url = (
        _text(args.source_url)
        or _text(receipt.get("source_url"))
        or _text(record.get("source_url"))
        or CATALOG_URL
    )
    retrieved_at = _text(args.retrieved_at)
    if retrieved_at is None:
        retrieved_at = datetime.fromtimestamp(
            path.stat().st_mtime,
            tz=timezone.utc,
        ).isoformat()
    artifact_id = hashlib.sha256(
        (
            f"{SOURCE_ID}\x00{native_locator}\x00{artifact_sha256}"
        ).encode("utf-8")
    ).hexdigest()
    return ArtifactContext(
        path=path,
        native_locator=native_locator,
        published_date=published_date,
        section=section,
        family=family,
        cadence=cadence,
        artifact_sha256=artifact_sha256,
        artifact_size_bytes=size_bytes,
        artifact_id=artifact_id,
        source_url=source_url,
        retrieved_at=retrieved_at,
        schema_workbook_path=schema_path,
        schema_workbook_sha256=schema_sha256,
    )


def _normalized_row(
    context: ArtifactContext,
    row: Mapping[str, Any],
    codebooks: Mapping[str, Mapping[str, tuple[str, ...]]],
) -> dict[str, Any]:
    if context.section == "Civil":
        case_number = _text(row.get("casenbr"))
        dates = {
            field: _parse_date(row.get(field))
            for field in (
                "rundate",
                "file_dt",
                "act/file_dt",
                "judgment_dt",
                "lastmoddt",
            )
            if field in row
        }
        if context.family == "case_summary":
            codes = {
                "case_status": _code_observation(
                    row,
                    code_field="cst",
                    codebooks=codebooks,
                ),
                "case_type": _code_observation(
                    row,
                    code_field="cs_typ",
                    codebooks=codebooks,
                ),
                "action_type": _code_observation(
                    row,
                    code_field="toac",
                    codebooks=codebooks,
                ),
                "judgment": _code_observation(
                    row,
                    code_field="judgment",
                    codebooks=codebooks,
                ),
            }
        elif context.family == "party":
            codes = {
                "connection": _code_observation(
                    row,
                    code_field="coc",
                    codebooks=codebooks,
                ),
                "active_indicator": {
                    "code_raw": row.get("active_ind"),
                    "code": _text(row.get("active_ind")),
                },
                "pro_se_indicator": {
                    "code_raw": row.get("prose_ind"),
                    "code": _text(row.get("prose_ind")),
                },
            }
        else:
            codes = {
                "activity": _code_observation(
                    row,
                    code_field="code",
                    literal_field="desc",
                    sheet="act",
                    codebooks=codebooks,
                )
            }
    else:
        case_number = _text(row.get("cas"))
        dates = {
            field: _parse_date(row.get(field))
            for field in (
                "rundate",
                "fda",
                "nda",
                "def_dob",
                "dispdt",
            )
            if field in row
        }
        codes = {
            "division": _code_observation(
                row,
                code_field="cdi",
                codebooks=codebooks,
            ),
            "instrument": _code_observation(
                row,
                code_field="ins",
                codebooks=codebooks,
            ),
            "case_disposition": _code_observation(
                row,
                code_field="cad",
                literal_field=(
                    "disposition"
                    if context.family == "dispositions"
                    else None
                ),
                codebooks=codebooks,
            ),
            "case_status": _code_observation(
                row,
                code_field="cst",
                codebooks=codebooks,
            ),
            "defendant_status": _code_observation(
                row,
                code_field="dst",
                codebooks=codebooks,
            ),
            "current_offense": _code_observation(
                row,
                code_field="curr_off",
                literal_field="curr_off_lit",
                codebooks=codebooks,
                sheet="curr_off",
            ),
            "offense_level": _code_observation(
                row,
                code_field="curr_l_d",
                codebooks=codebooks,
                sheet="curr_l_d",
            ),
            "docket_type": _code_observation(
                row,
                code_field="cnc",
                codebooks=codebooks,
            ),
            "appearance_reason": _code_observation(
                row,
                code_field="rea",
                codebooks=codebooks,
            ),
            "attorney_connection": _code_observation(
                row,
                code_field="aty_coc",
                literal_field="aty_coc_lit",
                codebooks=codebooks,
            ),
            "race": _code_observation(
                row,
                code_field="def_rac",
                codebooks=codebooks,
            ),
            "bond_exception": _code_observation(
                row,
                code_field="bamexp",
                codebooks=codebooks,
            ),
        }
    unparsed_dates = sorted(
        field
        for field, observation in dates.items()
        if observation.get("state") == "unparsed"
    )
    partial_dates = sorted(
        field
        for field, observation in dates.items()
        if observation.get("state") == "partial"
    )
    return {
        "source_id": SOURCE_ID,
        "native_locator": context.native_locator,
        "published_date": context.published_date,
        "section": context.section,
        "dataset_family": context.family,
        "cadence": context.cadence,
        "case_number": case_number,
        "dates": dates,
        "codes": codes,
        "normalization_issues": {
            "unparsed_dates": unparsed_dates,
            "partial_dates": partial_dates,
        },
    }


def _court_data(
    context: ArtifactContext,
    normalized: Mapping[str, Any],
) -> tuple[dict[str, Any], str]:
    if context.section == "Civil":
        division = "civil"
        court_id = "tx-harris-district-clerk-civil"
        name = "Harris County District Clerk Civil Bulk Records"
    else:
        division_code = _text(
            normalized["codes"]["division"].get("code")
        )
        division_lookup = {
            "2": "misdemeanor",
            "002": "misdemeanor",
            "3": "felony",
            "003": "felony",
        }
        division = division_lookup.get(division_code or "", "criminal_unknown")
        court_id = f"tx-harris-district-clerk-{division.replace('_', '-')}"
        name = (
            "Harris County District Clerk "
            f"{division.replace('_', ' ').title()} Bulk Records"
        )
    return (
        {
            "court_id": court_id,
            "native_court_id": division,
            "name": name,
            "state_code": "TX",
            "county_geoid": "48201",
            "court_level": "trial_court_bulk_records",
            "division": division,
            "official_url": CATALOG_URL,
        },
        division,
    )


def _party(
    *,
    role: str,
    name: Any,
    sequence_no: int,
) -> dict[str, Any] | None:
    raw_name = _text(name)
    if raw_name is None:
        return None
    return {
        "sequence_no": sequence_no,
        "role": role,
        "raw_name": raw_name,
        "normalized_name": _normalized_name(raw_name),
        "access_state": "public",
        "native_access_state": "district_clerk_public_bulk_extract",
    }


def _attorney(
    *,
    name: Any,
    source_person_number: Any,
    source_identifier_kind: str,
    connection: Mapping[str, Any],
    party_index: int | None,
    source_entry_id: str,
) -> dict[str, Any] | None:
    raw_name = _text(name)
    if raw_name is None:
        return None
    record: dict[str, Any] = {
        "attorney": {
            "raw_name": raw_name,
            "normalized_name": _normalized_name(raw_name),
            "bar_id": (
                _text(source_person_number)
                if source_identifier_kind == "bar_id"
                else ""
            ),
        },
        "source_entry_id": source_entry_id,
        "native_connection_code": connection.get("code"),
        "native_connection_literal": _code_label(connection),
        "source_person_number": _text(source_person_number),
        "source_identifier_kind": source_identifier_kind,
    }
    if party_index is not None:
        record["party_index"] = party_index
    return record


def _civil_projection(
    context: ArtifactContext,
    row: Mapping[str, Any],
    normalized: Mapping[str, Any],
    occurrence_id: str,
) -> dict[str, Any]:
    family = context.family
    dates = normalized["dates"]
    codes = normalized["codes"]
    parties: list[dict[str, Any]] = []
    attorneys: list[dict[str, Any]] = []
    docket_entries: list[dict[str, Any]] = []
    events: list[dict[str, Any]] = []
    filing_date = None
    disposition_date = None
    status = None
    case_type = None

    if family == "case_summary":
        filing_date = _projectable_date(dates["file_dt"])
        disposition_date = _projectable_date(dates["judgment_dt"])
        status = _code_label(codes["case_status"])
        case_type = _code_label(codes["case_type"])
        plaintiff = _party(
            role="plaintiff",
            name=row.get("plaintiff"),
            sequence_no=1,
        )
        defendant = _party(
            role="defendant",
            name=row.get("defendant"),
            sequence_no=2,
        )
        parties.extend(
            value for value in (plaintiff, defendant) if value is not None
        )
        judgment_code = _text(codes["judgment"].get("code"))
        if disposition_date and judgment_code:
            events.append(
                {
                    "native_event_id": (
                        "civil-judgment:"
                        + _stable_token(
                            normalized["case_number"],
                            row.get("pj_num"),
                            judgment_code,
                            disposition_date,
                        )
                    ),
                    "event_type": "judgment",
                    "event_date": disposition_date,
                    "disposition": _code_label(codes["judgment"]),
                    "assertion_kind": "judgment",
                    "native_assertion_kind": "civil_case_summary_judgment",
                    "native_code": judgment_code,
                    "native_literal": _code_label(codes["judgment"]),
                }
            )
    elif family == "party":
        connection = codes["connection"]
        role = _code_label(connection) or "source_connected_party"
        source_party = _party(
            role=role,
            name=row.get("party_name"),
            sequence_no=_stable_sequence(
                row.get("person_num"),
                row.get("conn_num"),
                row.get("coc"),
                row.get("party_name"),
            ),
        )
        role_folded = role.casefold()
        source_party_is_attorney = (
            "attorney" in role_folded
            or _text(row.get("party_bar")) is not None
        )
        represented_party_index: int | None = None
        if source_party is not None and not source_party_is_attorney:
            parties.append(source_party)
            represented_party_index = 0
        elif source_party_is_attorney:
            source_attorney = _attorney(
                name=row.get("party_name"),
                source_person_number=row.get("party_bar"),
                source_identifier_kind="bar_id",
                connection=connection,
                party_index=None,
                source_entry_id=occurrence_id,
            )
            if source_attorney is not None:
                attorneys.append(source_attorney)
        associated = _attorney(
            name=row.get("assoc_aty_name"),
            source_person_number=row.get("assoc_aty_bar"),
            source_identifier_kind="bar_id",
            connection=connection,
            party_index=represented_party_index,
            source_entry_id=occurrence_id,
        )
        if associated is not None:
            attorneys.append(associated)
    else:
        activity = codes["activity"]
        event_date = _projectable_date(dates["act/file_dt"])
        entry_id = (
            "civil-activity:"
            + _stable_token(
                normalized["case_number"],
                row.get("pj_num"),
                row.get("seq_num"),
                activity.get("code"),
                event_date,
                activity.get("source_literal"),
            )
        )
        docket_entries.append(
            {
                "native_entry_id": entry_id,
                "sequence_no": _text(row.get("seq_num")),
                "subsequence_no": _text(row.get("pj_num")),
                "event_code": _text(activity.get("code")),
                "event_type": "civil_activity",
                "raw_text": _text(activity.get("source_literal")),
                "filed_date": event_date,
                "event_date": event_date,
                "filer_raw": _text(row.get("fildedby")),
                "document_available": None,
                "access_state": "public",
                "native_access_state": "district_clerk_bulk_activity",
                "native_code": _text(activity.get("code")),
                "native_literal": _code_label(activity),
                "source_occurrence_id": occurrence_id,
            }
        )
        case_type = _text(row.get("case_type"))

    return {
        "filing_date": filing_date,
        "disposition_date": disposition_date,
        "status": status,
        "case_type": case_type,
        "parties": parties,
        "attorneys": attorneys,
        "docket_entries": docket_entries,
        "case_events": events,
    }


def _criminal_projection(
    context: ArtifactContext,
    row: Mapping[str, Any],
    normalized: Mapping[str, Any],
    occurrence_id: str,
) -> dict[str, Any]:
    dates = normalized["dates"]
    codes = normalized["codes"]
    filing_date = _projectable_date(dates["fda"])
    next_appearance = _projectable_date(dates["nda"])
    disposition_date = (
        _projectable_date(dates["dispdt"])
        if "dispdt" in dates
        else None
    )
    defendant = _party(
        role="defendant",
        name=row.get("def_nam"),
        sequence_no=_stable_sequence(
            row.get("def_spn"),
            row.get("def_nam"),
        ),
    )
    parties = [defendant] if defendant is not None else []
    if context.family == "dispositions":
        complainant = _party(
            role="complainant",
            name=row.get("comp_nam"),
            sequence_no=_stable_sequence(
                row.get("off_rpt_num"),
                row.get("comp_nam"),
                "complainant",
            ),
        )
        if complainant is not None:
            parties.append(complainant)

    attorney = _attorney(
        name=row.get("aty_nam"),
        source_person_number=row.get("aty_spn"),
        source_identifier_kind="harris_person_number",
        connection=codes["attorney_connection"],
        party_index=0 if defendant is not None else None,
        source_entry_id=occurrence_id,
    )
    attorneys = [attorney] if attorney is not None else []

    instrument = codes["instrument"]
    offense = codes["current_offense"]
    filing_entry_id = (
        "criminal-filing:"
        + _stable_token(
            normalized["case_number"],
            instrument.get("code"),
            filing_date,
            offense.get("code"),
        )
    )
    docket_entries = [
        {
            "native_entry_id": filing_entry_id,
            "event_code": _text(instrument.get("code")),
            "event_type": "criminal_filing_metadata",
            "raw_text": _code_label(instrument),
            "filed_date": filing_date,
            "event_date": filing_date,
            "document_available": None,
            "access_state": "public",
            "native_access_state": "district_clerk_bulk_filing_metadata",
            "native_code": _text(instrument.get("code")),
            "native_literal": _code_label(instrument),
            "source_occurrence_id": occurrence_id,
        }
    ]

    events: list[dict[str, Any]] = []
    offense_code = _text(offense.get("code"))
    offense_literal = _code_label(offense)
    if filing_date and (offense_code or offense_literal):
        events.append(
            {
                "native_event_id": (
                    "criminal-charge:"
                    + _stable_token(
                        normalized["case_number"],
                        offense_code,
                        offense_literal,
                        filing_date,
                    )
                ),
                "event_type": "charge",
                "event_date": filing_date,
                "filed_date": filing_date,
                "assertion_kind": "charge",
                "native_assertion_kind": "publisher_current_offense",
                "disposition": None,
                "native_code": offense_code,
                "native_literal": offense_literal,
                "offense_level": _code_label(codes["offense_level"]),
                "source_entry_native_id": filing_entry_id,
            }
        )
    if next_appearance:
        events.append(
            {
                "native_event_id": (
                    "criminal-setting:"
                    + _stable_token(
                        normalized["case_number"],
                        next_appearance,
                        codes["docket_type"].get("code"),
                        codes["appearance_reason"].get("code"),
                    )
                ),
                "event_type": "future_setting",
                "event_date": next_appearance,
                "assertion_kind": "docket_metadata",
                "native_assertion_kind": "publisher_next_appearance",
                "native_docket_code": _text(
                    codes["docket_type"].get("code")
                ),
                "native_docket_literal": _code_label(codes["docket_type"]),
                "native_reason_code": _text(
                    codes["appearance_reason"].get("code")
                ),
                "native_reason_literal": _code_label(
                    codes["appearance_reason"]
                ),
            }
        )
    disposition = codes["case_disposition"]
    disposition_code = _text(disposition.get("code"))
    disposition_literal = _code_label(disposition)
    if context.family == "dispositions" and disposition_date and (
        disposition_code or disposition_literal
    ):
        events.append(
            {
                "native_event_id": (
                    "criminal-disposition:"
                    + _stable_token(
                        normalized["case_number"],
                        disposition_code,
                        disposition_literal,
                        disposition_date,
                    )
                ),
                "event_type": "disposition",
                "event_date": disposition_date,
                "disposition": disposition_literal,
                "assertion_kind": "docket_metadata",
                "native_assertion_kind": "publisher_case_disposition",
                "native_code": disposition_code,
                "native_literal": disposition_literal,
                "sentence_raw": _text(row.get("sentence")),
            }
        )

    return {
        "filing_date": filing_date,
        "disposition_date": disposition_date,
        "status": _code_label(codes["case_status"]),
        "case_type": _code_label(codes["division"]) or "criminal",
        "parties": parties,
        "attorneys": attorneys,
        "docket_entries": docket_entries,
        "case_events": events,
    }


def _projection_record(
    context: ArtifactContext,
    row: Mapping[str, Any],
    normalized: Mapping[str, Any],
    occurrence_id: str,
) -> tuple[dict[str, Any] | None, str | None]:
    case_number = _text(normalized.get("case_number"))
    if case_number is None:
        return None, "source row has no case number"
    court, division = _court_data(context, normalized)
    components = (
        _civil_projection(context, row, normalized, occurrence_id)
        if context.section == "Civil"
        else _criminal_projection(context, row, normalized, occurrence_id)
    )
    source_internal_id = f"{division}:{case_number}"
    return (
        {
            "source_id": SOURCE_ID,
            "record_kind": f"harris_bulk_{context.family}_row",
            "source_result_id": occurrence_id,
            "canonical_ref": (
                f"HARRIS-COURT-BULK:{context.artifact_id}:{occurrence_id}"
            ),
            "case": {
                "source_id": SOURCE_ID,
                "court": court,
                "raw_case_number": case_number,
                "display_case_number": case_number,
                "source_internal_id": source_internal_id,
                "case_type": components["case_type"],
                "filing_date": components["filing_date"],
                "disposition_date": components["disposition_date"],
                "status": components["status"],
                "access_state": "public",
                "native_access_state": "district_clerk_public_bulk_extract",
                "certified_record": False,
                "source_url": CATALOG_URL,
                "preserve_existing_case_fields": (
                    context.family not in {"case_summary", "dispositions"}
                ),
                "source_artifact": {
                    "artifact_id": context.artifact_id,
                    "native_locator": context.native_locator,
                    "artifact_sha256": context.artifact_sha256,
                    "published_date": context.published_date,
                    "dataset_family": context.family,
                    "cadence": context.cadence,
                    "source_occurrence_id": occurrence_id,
                },
                "parties": components["parties"],
                "attorneys": components["attorneys"],
                "docket_entries": components["docket_entries"],
                "case_events": components["case_events"],
                "documents": [],
            },
        },
        None,
    )


def _ensure_artifact_snapshot(
    db: sqlite3.Connection,
    context: ArtifactContext,
    *,
    raw_header: Sequence[str],
    normalized_header: Sequence[str],
) -> tuple[int, bool, str]:
    header_fingerprint = sha256_fingerprint(
        {
            "source_id": SOURCE_ID,
            "section": context.section,
            "dataset_family": context.family,
            "raw_header": list(raw_header),
            "normalized_header": list(normalized_header),
        }
    )
    existing = db.execute(
        """
        SELECT *
        FROM harris_court_bulk_artifact
        WHERE artifact_id=?
        """,
        (context.artifact_id,),
    ).fetchone()
    if existing is not None:
        expected = {
            "native_locator": context.native_locator,
            "published_date": context.published_date,
            "section": context.section,
            "dataset_family": context.family,
            "cadence": context.cadence,
            "artifact_sha256": context.artifact_sha256,
            "artifact_size_bytes": context.artifact_size_bytes,
            "header_fingerprint": header_fingerprint,
        }
        mismatches = {
            field: {
                "stored": existing[field],
                "observed": value,
            }
            for field, value in expected.items()
            if existing[field] != value
        }
        if mismatches:
            raise HarrisCourtBulkIngestError(
                "existing artifact identity conflicts with observed metadata: "
                + canonical_json(mismatches)
            )
        db.execute(
            """
            UPDATE harris_court_bulk_artifact
            SET artifact_path=?, source_url=?, retrieved_at=?,
                schema_workbook_sha256=?, schema_workbook_path=?
            WHERE artifact_id=?
            """,
            (
                str(context.path),
                context.source_url,
                context.retrieved_at,
                context.schema_workbook_sha256,
                (
                    str(context.schema_workbook_path)
                    if context.schema_workbook_path
                    else None
                ),
                context.artifact_id,
            ),
        )
        return int(existing["snapshot_id"]), False, header_fingerprint

    query_fingerprint = sha256_fingerprint(
        {
            "source_id": SOURCE_ID,
            "operation": "bulk_artifact_ingest",
            "native_locator": context.native_locator,
            "artifact_sha256": context.artifact_sha256,
        }
    )
    snapshot_cursor = db.execute(
        """
        INSERT INTO source_snapshot(
            source_id, query_fingerprint, source_url, retrieved_at,
            access_status, coverage_json, schema_fingerprint,
            raw_artifact_sha256, raw_artifact_path, raw_json, warning_json
        ) VALUES (?, ?, ?, ?, 'ok', ?, ?, ?, ?, ?, ?)
        """,
        (
            SOURCE_ID,
            query_fingerprint,
            context.source_url,
            context.retrieved_at,
            canonical_json(
                {
                    "jurisdiction_geoid": "48201",
                    "published_date": context.published_date,
                    "section": context.section,
                    "dataset_family": context.family,
                    "cadence": context.cadence,
                }
            ),
            header_fingerprint,
            context.artifact_sha256,
            str(context.path),
            canonical_json(
                {
                    "record_kind": "court_bulk_artifact_snapshot",
                    "artifact": context.public_record(),
                    "header": {
                        "raw": list(raw_header),
                        "normalized": list(normalized_header),
                    },
                }
            ),
            canonical_json(
                [
                    (
                        "The extract contains record metadata, not a claim "
                        "that filing-document images are available."
                    ),
                    (
                        "Source rows remain distinct observations even when "
                        "their normalized case components coincide."
                    ),
                ]
            ),
        ),
    )
    snapshot_id = int(snapshot_cursor.lastrowid)
    db.execute(
        """
        INSERT INTO harris_court_bulk_artifact(
            artifact_id, source_id, snapshot_id, native_locator,
            published_date, section, dataset_family, cadence,
            artifact_sha256, artifact_size_bytes, artifact_path,
            source_url, retrieved_at, header_json, header_fingerprint,
            schema_workbook_sha256, schema_workbook_path
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            context.artifact_id,
            SOURCE_ID,
            snapshot_id,
            context.native_locator,
            context.published_date,
            context.section,
            context.family,
            context.cadence,
            context.artifact_sha256,
            context.artifact_size_bytes,
            str(context.path),
            context.source_url,
            context.retrieved_at,
            canonical_json(
                {
                    "raw": list(raw_header),
                    "normalized": list(normalized_header),
                }
            ),
            header_fingerprint,
            context.schema_workbook_sha256,
            (
                str(context.schema_workbook_path)
                if context.schema_workbook_path
                else None
            ),
        ),
    )
    return snapshot_id, True, header_fingerprint


def _case_id_for_projection(
    db: sqlite3.Connection,
    projection: Mapping[str, Any],
) -> int:
    case = projection["case"]
    row = db.execute(
        """
        SELECT case_id
        FROM case_record
        WHERE source_id=? AND court_id=? AND case_identity_key=?
        """,
        (
            SOURCE_ID,
            case["court"]["court_id"],
            f"native:{case['source_internal_id']}",
        ),
    ).fetchone()
    if row is None:
        raise HarrisCourtBulkIngestError(
            "generic case projection did not produce a case record"
        )
    return int(row["case_id"])


def _upsert_row_observation(
    db: sqlite3.Connection,
    *,
    context: ArtifactContext,
    source_row_number: int,
    raw_row: Mapping[str, Any],
    normalized: Mapping[str, Any],
    projection_status: str,
    projection_reason: str | None,
    case_id: int | None,
    canonical_ref: str | None,
    projection_refs: Mapping[str, Any],
) -> bool:
    raw_json = canonical_json(raw_row)
    row_sha256 = hashlib.sha256(
        canonical_json(
            {
                "fields": raw_row.get("fields"),
                "extra_fields": raw_row.get("extra_fields"),
            }
        ).encode("utf-8")
    ).hexdigest()
    occurrence_id = hashlib.sha256(
        (
            f"{context.artifact_id}\x00{source_row_number}\x00{row_sha256}"
        ).encode("utf-8")
    ).hexdigest()
    existing = db.execute(
        """
        SELECT row_occurrence_id
        FROM harris_court_bulk_row
        WHERE artifact_id=? AND source_row_number=?
        """,
        (context.artifact_id, source_row_number),
    ).fetchone()
    if existing is not None and existing["row_occurrence_id"] != occurrence_id:
        raise HarrisCourtBulkIngestError(
            "source row identity changed without an artifact SHA change"
        )
    db.execute(
        """
        INSERT INTO harris_court_bulk_row(
            row_occurrence_id, artifact_id, source_row_number, row_sha256,
            case_number, case_id, projection_status, projection_reason,
            canonical_ref, raw_fields_json, normalized_json,
            projection_refs_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(row_occurrence_id) DO UPDATE SET
            case_number=excluded.case_number,
            case_id=excluded.case_id,
            projection_status=excluded.projection_status,
            projection_reason=excluded.projection_reason,
            canonical_ref=excluded.canonical_ref,
            raw_fields_json=excluded.raw_fields_json,
            normalized_json=excluded.normalized_json,
            projection_refs_json=excluded.projection_refs_json
        """,
        (
            occurrence_id,
            context.artifact_id,
            source_row_number,
            row_sha256,
            _text(normalized.get("case_number")),
            case_id,
            projection_status,
            projection_reason,
            canonical_ref,
            raw_json,
            canonical_json(normalized),
            canonical_json(projection_refs),
        ),
    )
    return existing is None


def _ingest(args: argparse.Namespace, context: ArtifactContext) -> dict[str, Any]:
    csv.field_size_limit(CSV_FIELD_SIZE_LIMIT)
    codebooks = _load_codebooks(context.schema_workbook_path)
    undefined_cp1252_bytes = _undefined_cp1252_counts(context.path)
    generic_counts = {key: 0 for key in GENERIC_COUNT_KEYS}
    counts = {
        "rows_scanned": 0,
        "rows_skipped_before_checkpoint": 0,
        "rows_processed": 0,
        "row_occurrences_inserted": 0,
        "row_occurrences_reused": 0,
        "rows_projected": 0,
        "rows_projected_with_issues": 0,
        "rows_unresolved": 0,
    }
    exhausted = True
    raw_header: list[str]
    normalized_header: list[str]
    db = connect_courts(args.court_db)
    try:
        db.executescript(HARRIS_EXTENSION_SCHEMA)
        _ensure_docket_hearing_columns(db)
        db.commit()
        with context.path.open(
            "r",
            encoding="cp1252",
            errors=CP1252_PRESERVE_ERROR_HANDLER,
            newline="",
        ) as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            raw_header, normalized_header = _normalize_header(reader.fieldnames)
            required = SUPPORTED_FAMILIES[(context.section, context.family)]
            missing = sorted(required - set(normalized_header))
            if missing:
                raise HarrisCourtBulkIngestError(
                    f"{context.section}/{context.family} header lacks required "
                    "source columns: " + ", ".join(missing)
                )
            db.execute("BEGIN IMMEDIATE")
            snapshot_id, artifact_inserted, header_fingerprint = (
                _ensure_artifact_snapshot(
                    db,
                    context,
                    raw_header=raw_header,
                    normalized_header=normalized_header,
                )
            )
            db.commit()

            for source_row_number, source_row in enumerate(reader, start=1):
                counts["rows_scanned"] += 1
                zero_based_row = source_row_number - 1
                if zero_based_row < args.start_row:
                    counts["rows_skipped_before_checkpoint"] += 1
                    continue
                if (
                    args.limit is not None
                    and counts["rows_processed"] >= args.limit
                ):
                    exhausted = False
                    break

                row, extras = _row_by_normalized_header(
                    source_row,
                    raw_header,
                    normalized_header,
                )
                raw_observation = {
                    "source_row_number": source_row_number,
                    "fields": {
                        raw: source_row.get(raw)
                        for raw in raw_header
                    },
                    "extra_fields": extras,
                }
                normalized = _normalized_row(context, row, codebooks)
                row_sha256 = hashlib.sha256(
                    canonical_json(
                        {
                            "fields": raw_observation["fields"],
                            "extra_fields": extras,
                        }
                    ).encode("utf-8")
                ).hexdigest()
                occurrence_id = hashlib.sha256(
                    (
                        f"{context.artifact_id}\x00{source_row_number}"
                        f"\x00{row_sha256}"
                    ).encode("utf-8")
                ).hexdigest()

                projection: dict[str, Any] | None
                projection_reason: str | None
                if extras:
                    projection = None
                    projection_reason = (
                        "source row has more values than the declared header"
                    )
                else:
                    projection, projection_reason = _projection_record(
                        context,
                        row,
                        normalized,
                        occurrence_id,
                    )

                canonical_ref = None
                case_id = None
                projection_refs: dict[str, Any] = {}
                normalization_issues = normalized["normalization_issues"]
                has_issues = any(normalization_issues.values())
                if projection is None:
                    projection_status = "unresolved"
                    counts["rows_unresolved"] += 1
                else:
                    canonical_ref = _project_case(
                        db,
                        projection,
                        source_id=SOURCE_ID,
                        snapshot_id=snapshot_id,
                        counts=generic_counts,
                    )
                    case_id = _case_id_for_projection(db, projection)
                    projection_refs = {
                        "case_id": case_id,
                        "canonical_ref": canonical_ref,
                        "party_count": len(projection["case"]["parties"]),
                        "attorney_count": len(projection["case"]["attorneys"]),
                        "docket_entry_ids": [
                            entry["native_entry_id"]
                            for entry in projection["case"]["docket_entries"]
                        ],
                        "case_event_ids": [
                            event["native_event_id"]
                            for event in projection["case"]["case_events"]
                        ],
                        "document_count": 0,
                    }
                    if has_issues:
                        projection_status = "projected_with_unparsed_fields"
                        projection_reason = canonical_json(
                            normalization_issues
                        )
                        counts["rows_projected_with_issues"] += 1
                    else:
                        projection_status = "projected"
                    counts["rows_projected"] += 1

                inserted = _upsert_row_observation(
                    db,
                    context=context,
                    source_row_number=source_row_number,
                    raw_row=raw_observation,
                    normalized=normalized,
                    projection_status=projection_status,
                    projection_reason=projection_reason,
                    case_id=case_id,
                    canonical_ref=canonical_ref,
                    projection_refs=projection_refs,
                )
                if inserted:
                    counts["row_occurrences_inserted"] += 1
                else:
                    counts["row_occurrences_reused"] += 1
                counts["rows_processed"] += 1
                if counts["rows_processed"] % args.batch_size == 0:
                    db.commit()
            db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    next_checkpoint = (
        args.start_row + counts["rows_processed"]
        if not exhausted
        else None
    )
    return {
        "source_id": SOURCE_ID,
        "status": "ok",
        "artifact": context.public_record(),
        "schema": {
            "encoding": "cp1252",
            "decode_error_policy": (
                "undefined_cp1252_bytes_preserved_as_same-valued_C1_controls"
            ),
            "undefined_cp1252_byte_counts": undefined_cp1252_bytes,
            "delimiter": "tab",
            "raw_header": raw_header,
            "normalized_header": normalized_header,
            "header_fingerprint": header_fingerprint,
            "required_projection_columns": sorted(
                SUPPORTED_FAMILIES[(context.section, context.family)]
            ),
            "schema_workbook_sha256": context.schema_workbook_sha256,
            "codebook_sheets_loaded": sorted(codebooks),
        },
        "artifact_snapshot_inserted": artifact_inserted,
        "snapshot_id": snapshot_id,
        "counts": counts,
        "generic_projection_attempts": generic_counts,
        "exhausted": exhausted,
        "next_checkpoint_row": next_checkpoint,
        "projection_scope": {
            "case_shells": True,
            "parties_and_attorneys": context.family
            in {"case_summary", "party", "filings", "dispositions"},
            "docket_or_activity": context.family
            in {"activity", "filings", "dispositions"},
            "charges_and_dispositions": context.section == "Criminal",
            "filing_document_images": False,
            "unresolved_rows_preserved": True,
        },
        "follow_on_families": [
            "Civil/case_setting",
            "Civil/service",
            "Civil/historical_daily",
            "Criminal/future_settings",
            "Criminal/historical_snapshot",
        ],
    }


def execute(args: argparse.Namespace) -> dict[str, Any]:
    """Validate and ingest one caller-selected local Harris extract."""

    context = _artifact_context(args)
    return _ingest(args, context)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Stream a downloaded Harris District Clerk text extract into "
            "the state/local court sidecar"
        )
    )
    subparsers = parser.add_subparsers(dest="command", required=True)
    ingest = subparsers.add_parser(
        "ingest",
        help="Validate and ingest one supported header-bearing text artifact",
    )
    ingest.add_argument("artifact")
    ingest.add_argument(
        "--artifact-result",
        help=(
            "JSON output from query_harris_court_bulk.py download; supplies "
            "and verifies exact catalog and receipt provenance"
        ),
    )
    ingest.add_argument(
        "--native-locator",
        help="Exact live-catalog locator when no artifact result is supplied",
    )
    ingest.add_argument(
        "--published-date",
        help="Exact live-catalog publication date when no result is supplied",
    )
    ingest.add_argument("--expected-sha256")
    ingest.add_argument("--source-url")
    ingest.add_argument("--retrieved-at")
    ingest.add_argument(
        "--schema-workbook",
        help=(
            "Optional official FIELD_CODES or criminal layout workbook used "
            "to retain published code literals alongside row values"
        ),
    )
    ingest.add_argument("--court-db", default=str(DEFAULT_COURT_DB))
    ingest.add_argument(
        "--start-row",
        type=int,
        default=0,
        help="Caller-selected zero-based data-row checkpoint",
    )
    ingest.add_argument(
        "--limit",
        type=int,
        help="Optional caller-selected data-row ceiling",
    )
    ingest.add_argument(
        "--batch-size",
        type=int,
        default=1_000,
        help="Rows committed per transaction; this is not a record ceiling",
    )
    add_output_args(ingest)
    return parser


def _validate_args(
    parser: argparse.ArgumentParser,
    args: argparse.Namespace,
) -> None:
    if args.start_row < 0:
        parser.error("--start-row must not be negative")
    if args.limit is not None and args.limit <= 0:
        parser.error("--limit must be positive")
    if args.batch_size <= 0:
        parser.error("--batch-size must be positive")
    if args.expected_sha256 and not SHA256_RE.fullmatch(args.expected_sha256):
        parser.error("--expected-sha256 must be a SHA-256 hex digest")


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    _validate_args(parser, args)
    try:
        payload = execute(args)
    except (
        HarrisCourtBulkIngestError,
        OSError,
        UnicodeError,
        csv.Error,
        sqlite3.Error,
    ) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(2) from error
    if write_output(
        payload,
        args,
        summary=(
            "Harris court bulk ingest "
            f"({payload['counts']['rows_processed']} rows)"
        ),
    ):
        return
    print(json.dumps(payload, indent=2 if args.json_out else None, sort_keys=True))


if __name__ == "__main__":
    main()
