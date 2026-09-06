from __future__ import annotations

import hashlib
import zipfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from tools import query_ohio_franklin_auditor_bulk as franklin
from tools.public_records_bulk import ArtifactProbe, DownloadResult
from tools.public_records_contract import ResultStatus


FIXTURES = (
    Path(__file__).parent
    / "fixtures"
    / "public_records"
    / "ohio_franklin_auditor_bulk"
)


def _text(name: str) -> str:
    return (FIXTURES / name).read_text(encoding="utf-8")


class FixtureDirectoryClient:
    request_count = 0

    _fixtures = {
        franklin.DIRECTORY_ROOT: "root.html",
        franklin.OUTSIDE_USER_ROOT: "outside.html",
        f"{franklin.OUTSIDE_USER_ROOT}2026/": "outside-2026.html",
        (
            f"{franklin.OUTSIDE_USER_ROOT}2026/"
            "2026-07-15%20Appraisal/"
        ): "appraisal-current.html",
        (
            f"{franklin.OUTSIDE_USER_ROOT}2026/"
            "2026-07-15%20Tax%20Accounting/"
        ): "tax-current.html",
        franklin.DAILY_ROOT: "daily.html",
        franklin.GIS_ROOT: "gis.html",
        f"{franklin.GIS_ROOT}2026/": "gis-2026.html",
        franklin.GIS_CURRENT_ROOT: "gis-current.html",
        franklin.PARCEL_CSV_ROOT: "parcel.html",
        f"{franklin.PARCEL_CSV_ROOT}1997/": "parcel-1997.html",
        f"{franklin.PARCEL_CSV_ROOT}1997/01/": "parcel-1997-01.html",
        f"{franklin.PARCEL_CSV_ROOT}2025/": "parcel-2025.html",
        f"{franklin.PARCEL_CSV_ROOT}2025/07/": "parcel-current.html",
    }

    def listing(self, url: str, *, refresh: bool = False):
        del refresh
        try:
            fixture = self._fixtures[url]
        except KeyError as error:
            raise AssertionError(f"unexpected fixture URL: {url}") from error
        self.request_count += 1
        return franklin.parse_iis_listing(_text(fixture), source_url=url)


class FixtureBulkClient:
    def __init__(
        self,
        *,
        content_length: int | None = None,
        signature_hex: str = "504b030414000000",
        format_hint: str = "zip",
    ) -> None:
        self.content_length = content_length
        self.signature_hex = signature_hex
        self.format_hint = format_hint
        self.probes = []
        self.downloads = []

    def probe(self, artifact, *, sample_bytes: int) -> ArtifactProbe:
        self.probes.append((artifact, sample_bytes))
        return ArtifactProbe(
            url=artifact.url,
            http_status=206,
            content_length=(
                self.content_length
                if self.content_length is not None
                else artifact.expected_size
            ),
            media_type=artifact.media_type,
            etag='"fixture"',
            last_modified=artifact.last_modified,
            accept_ranges=True,
            source_sha256=None,
            sample_size=sample_bytes,
            sample_sha256="0" * 64,
            signature_hex=self.signature_hex,
            format_hint=self.format_hint,
            headers={"accept-ranges": "bytes"},
        )

    def download(
        self,
        artifact,
        destination,
        *,
        resume: bool,
        max_bytes: int | None,
    ) -> DownloadResult:
        destination = Path(destination)
        self.downloads.append((artifact, destination, resume, max_bytes))
        return DownloadResult(
            path=str(destination),
            url=artifact.url,
            size=artifact.expected_size or 1,
            sha256=artifact.expected_sha256 or "a" * 64,
            expected_sha256=artifact.expected_sha256,
            etag='"fixture"',
            last_modified=artifact.last_modified,
            resumed_from=17 if resume else 0,
            reused_existing=False,
        )


def _args(*values: str):
    return franklin.build_parser().parse_args(list(values))


def _write_payment_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Parcel Id", "EffectiveDate", "TaxYear", "BillType", "Amount"])
    sheet.append(["010-000001-00", "2026-07-01", 2024, "REAL", 100.25])
    sheet.append(["010-000002-00", "2026-07-02", 2024, "REAL", 200.50])
    sheet.append(["010-000003-00", "2026-07-03", 2024, "REAL", 300.75])
    workbook.save(path)


def _write_daily_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AUDR - atENT - Daily Conveyance"
    sheet.append(["2 Daily Sales of 2 parcels between 07/23/2026 and 07/23/2026"])
    sheet.append(
        [
            "ISEXEMPT",
            "CONVEYNUMBER",
            "PARCELNUMBER",
            "SALEDATE",
            "SALETYPE",
            "SALESPRICE",
            "PARCELCOUNT",
            "OWNERNAME1",
            "OWNERNAME2",
            "OWNERADDRESS1",
            "OWNERADDRESS2",
            "PRIOROWNERNAME1",
            "PRIOROWNERNAME2",
            "LUC",
            "LANDUSE",
            "SITEADDRESS",
            "INSTRUMENTTYPE",
        ]
    )
    sheet.append(
        [
            "NON-EXEMPT",
            "C-100",
            "010-000001-00",
            "07/23/2026",
            "LAND AND BUILDING",
            450000,
            1,
            "NEW OWNER",
            None,
            "1 MAIL ST",
            "COLUMBUS OH",
            "PRIOR OWNER",
            None,
            "510",
            "SINGLE FAMILY",
            "1 SITE ST",
            "WD",
        ]
    )
    workbook.save(path)


def _write_appraisal_sales_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales010"
    sheet.append(
        [
            "PARCEL ID",
            "MAP ROUTING",
            "SALEDT",
            "NOPAR",
            "INSTRUMENT",
            "INSTRUNO",
            "VALID",
            "SALETYPE",
            "PRICE",
            "ADJAMT",
            "ADJPRICE",
            "CONDSALE_GRANTORRELATIVE",
            "CONDSALE_LIFESTATE",
            "CONDSALE_PARTINTEREST",
            "CONDSALE_LEASEDFEE",
            "CONDSALE_GRANTORMORTGAGEE",
            "CONDSALE_LANDCONTRACT",
            "CONDSALE_LEASEHOLD",
            "CONDSALE_TRADE",
            "CONDSALE_MINERALRIGHTSRESERVED",
            "CONDSALE_OTHER",
        ]
    )
    sheet.append(
        [
            "010-000001-00",
            "010",
            datetime(1986, 5, 1),
            1,
            "WD",
            "202607010000001",
            "0 - VALID",
            "1 - LAND & BUILDING",
            450000,
        ]
    )
    sheet.append(
        [
            "010-000002-00",
            "010",
            datetime(1995, 6, 2),
            1,
            "WD",
            "202607020000002",
            "99 - RMS INVALID",
            "1 - LAND & BUILDING",
            1,
        ]
    )
    sheet.append(
        [
            "010-000003-00",
            "010",
            datetime(2006, 7, 3),
            1,
            "WD",
            "202607030000003",
            None,
            "1 - LAND & BUILDING",
            400000,
        ]
    )
    workbook.save(path)


def test_iis_listing_preserves_native_metadata_and_rejects_identity_drift() -> None:
    listing = franklin.parse_iis_listing(
        _text("parcel-current.html"),
        source_url=f"{franklin.PARCEL_CSV_ROOT}2025/07/",
    )

    assert listing.path == "/Parcel_CSV/2025/07/"
    assert listing.entries[0].name == "Parcel.csv"
    assert listing.entries[0].size == 262_500_000
    assert listing.entries[0].modified_raw == "7/17/2026 3:00 PM"
    assert listing.entries[0].modified_at == "2026-07-17T15:00-04:00"

    with pytest.raises(franklin.FranklinSourceChanged):
        franklin.parse_iis_listing(
            _text("parcel-current.html").replace(
                "apps.franklincountyauditor.com - /Parcel_CSV/2025/07/",
                "Unexpected Portal",
            ),
            source_url=f"{franklin.PARCEL_CSV_ROOT}2025/07/",
        )
    with pytest.raises(franklin.FranklinSourceChanged):
        franklin.parse_iis_listing(
            _text("root.html"),
            source_url="https://example.org/",
        )


def test_release_discovery_keeps_source_date_semantics_separate() -> None:
    client = FixtureDirectoryClient()

    appraisal = franklin.discover_releases(client, "appraisal")[0]
    tax = franklin.discover_releases(client, "tax-accounting")[0]
    daily = franklin.discover_releases(client, "daily-conveyances")[0]
    gis = franklin.discover_releases(client, "gis-shapefiles")[0]
    parcel = franklin.discover_releases(client, "parcel-csv")[0]

    assert appraisal.release_id == "appraisal-2026-07-15"
    assert appraisal.release_date_basis == "source_directory_name"
    assert tax.release_id == "tax-accounting-2026-07-15"
    assert tax.directory_modified_at == "2026-07-30T13:30-04:00"
    assert daily.release_id == "daily-conveyance-2026-07-30"
    assert daily.release_date_basis == "artifact_filename_end_date"
    assert gis.release_id == "gis-current-2026-07-29"
    assert gis.release_date_basis == "artifact_filename_prefix"
    assert parcel.release_id == "parcel-csv-2025-07"
    assert parcel.release_date is None
    assert parcel.path_period == "2025-07"


def test_archive_selection_supports_1990s_years_and_exact_release_ids() -> None:
    client = FixtureDirectoryClient()
    parcel_root = client.listing(franklin.PARCEL_CSV_ROOT)
    assert [entry.name for entry in franklin._year_directories(parcel_root)] == [
        "1997",
        "2025",
    ]

    appraisal = franklin.resolve_release(
        client,
        "appraisal",
        "appraisal-2026-06-15",
    )
    parcel = franklin.resolve_release(
        client,
        "parcel-csv",
        "parcel-csv-2025-06",
    )
    historical_parcel = franklin.resolve_release(
        client,
        "parcel-csv",
        "parcel-csv-1997-01",
    )
    gis = franklin.resolve_release(
        client,
        "gis-shapefiles",
        "gis-archive-2026-07",
    )
    gis_current_id = franklin.discover_releases(
        client, "gis-shapefiles"
    )[0].release_id
    gis_current = franklin.resolve_release(
        client,
        "gis-shapefiles",
        gis_current_id,
    )

    assert appraisal.release_date == "2026-06-15"
    assert parcel.path_period == "2025-06"
    assert historical_parcel.path_period == "1997-01"
    assert gis.path_period == "2026-07"
    assert gis_current.release_id == gis_current_id
    assert franklin.artifacts_for_release(client, gis_current)


def test_artifact_identity_is_path_scoped_and_listing_cursor_detects_change() -> None:
    client = FixtureDirectoryClient()
    release = franklin.resolve_release(client, "appraisal", "current")
    records = franklin.artifacts_for_release(client, release)

    assert len(records) == 6
    parcel = next(record for record in records if record["filename"] == "Parcel.xlsx")
    assert parcel["native_document_id"].endswith(
        "2026-07-15 Appraisal/Parcel.xlsx"
    )
    assert parcel["artifact_id"] == hashlib.sha256(
        parcel["relative_path"].encode("utf-8")
    ).hexdigest()[:24]

    first, cursor = franklin.paginate_records(
        records,
        selection={"family": "appraisal", "release": release.release_id},
        limit=2,
        cursor=None,
    )
    assert len(first) == 2
    assert cursor
    second, _ = franklin.paginate_records(
        records,
        selection={"family": "appraisal", "release": release.release_id},
        limit=2,
        cursor=cursor,
    )
    assert {item["artifact_id"] for item in first}.isdisjoint(
        item["artifact_id"] for item in second
    )

    changed = [dict(record) for record in records]
    changed[0]["directory_size"] += 1
    with pytest.raises(franklin.FranklinSourceChanged):
        franklin.paginate_records(
            changed,
            selection={"family": "appraisal", "release": release.release_id},
            limit=2,
            cursor=cursor,
        )
    with pytest.raises(franklin.FranklinSelectionError):
        franklin.paginate_records(
            records,
            selection={"family": "tax-accounting"},
            limit=2,
            cursor=cursor,
        )


def test_artifact_probe_detects_directory_size_drift() -> None:
    client = FixtureDirectoryClient()
    bulk = FixtureBulkClient(content_length=36_543)
    args = _args(
        "artifact-probe",
        "daily-conveyances",
        "DailyConveyances_20260723.xlsx",
        "--release",
        "daily-conveyance-2026-07-23",
    )

    result = franklin.execute(
        args,
        directory_client=client,
        bulk_client=bulk,
        log_results=False,
    )

    assert result.status == ResultStatus.SOURCE_CHANGED
    assert result.errors[0].code == "franklin_auditor_source_changed"


def test_download_passes_explicit_destination_resume_and_hash(tmp_path: Path) -> None:
    client = FixtureDirectoryClient()
    bulk = FixtureBulkClient()
    destination = tmp_path / "daily.xlsx"
    expected = "b" * 64
    args = _args(
        "download",
        "daily-conveyances",
        "DailyConveyances_20260723.xlsx",
        "--release",
        "daily-conveyance-2026-07-23",
        "--destination",
        str(destination),
        "--expected-sha256",
        expected,
        "--max-download-bytes",
        "40000",
    )

    result = franklin.execute(
        args,
        directory_client=client,
        bulk_client=bulk,
        log_results=False,
    )

    assert result.status == ResultStatus.OK
    artifact, observed_destination, resume, max_bytes = bulk.downloads[0]
    assert observed_destination == destination
    assert resume is True
    assert max_bytes == 40_000
    assert artifact.expected_sha256 == expected
    assert result.records[0]["download"]["sha256"] == expected


def test_payment_inspection_and_rows_preserve_raw_grain_and_continuation(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "Payment2025.xlsx"
    _write_payment_workbook(workbook)

    inspection = franklin.inspect_local_artifact(
        workbook,
        record_family="payment",
        member=None,
        sheet_name=None,
        sample_rows=1,
        header_scan_rows=25,
        encoding="utf-8-sig",
    )
    assert inspection["raw_headers"] == [
        "Parcel Id",
        "EffectiveDate",
        "TaxYear",
        "BillType",
        "Amount",
    ]
    assert inspection["sample_rows"][0]["source_row_number"] == 2

    first, cursor, metadata = franklin.stream_local_rows(
        workbook,
        record_family="payment",
        release_id="tax-accounting-2026-07-15",
        release_date="2026-07-15",
        source_url="https://example.test/Payment2025.xlsx",
        member=None,
        sheet_name=None,
        query=None,
        parcel=None,
        from_date=None,
        to_date=None,
        limit=1,
        cursor=None,
        header_scan_rows=25,
        encoding="utf-8-sig",
    )
    assert cursor
    assert metadata["continuation_available"] is True
    assert first[0]["source_row_number"] == 2
    assert first[0]["parsed_fields"]["tax_year"] == 2024
    assert first[0]["join_candidates"]["normalized_parcel_id"] == "01000000100"
    assert first[0]["raw_headers"] == inspection["raw_headers"]
    assert '"worksheet":"Sheet1"' in first[0]["native_occurrence"]

    rest, final_cursor, _ = franklin.stream_local_rows(
        workbook,
        record_family="payment",
        release_id="tax-accounting-2026-07-15",
        release_date="2026-07-15",
        source_url="https://example.test/Payment2025.xlsx",
        member=None,
        sheet_name=None,
        query=None,
        parcel=None,
        from_date=None,
        to_date=None,
        limit=10,
        cursor=cursor,
        header_scan_rows=25,
        encoding="utf-8-sig",
    )
    assert [record["source_row_number"] for record in rest] == [3, 4]
    assert final_cursor is None
    assert len({first[0]["native_document_id"], *(r["native_document_id"] for r in rest)}) == 3


def test_row_cursor_binds_query_and_artifact_version(tmp_path: Path) -> None:
    workbook = tmp_path / "Payment2025.xlsx"
    _write_payment_workbook(workbook)
    common = {
        "record_family": "payment",
        "release_id": "tax-accounting-2026-07-15",
        "release_date": "2026-07-15",
        "source_url": None,
        "member": None,
        "sheet_name": None,
        "query": None,
        "parcel": None,
        "from_date": None,
        "to_date": None,
        "limit": 1,
        "cursor": None,
        "header_scan_rows": 25,
        "encoding": "utf-8-sig",
    }
    _, cursor, _ = franklin.stream_local_rows(workbook, **common)
    assert cursor

    with pytest.raises(franklin.FranklinSelectionError):
        franklin.stream_local_rows(
            workbook,
            **{**common, "parcel": "010-000002-00", "cursor": cursor},
        )

    book = Workbook()
    sheet = book.active
    sheet.append(["Parcel Id", "EffectiveDate", "TaxYear", "BillType", "Amount"])
    sheet.append(["010-000009-00", "2026-07-09", 2024, "REAL", 999])
    book.save(workbook)
    with pytest.raises(franklin.FranklinSourceChanged):
        franklin.stream_local_rows(
            workbook,
            **{**common, "cursor": cursor},
        )


def test_daily_conveyance_title_row_and_component_assertions_are_preserved(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "DailyConveyances_20260723.xlsx"
    _write_daily_workbook(workbook)

    rows, cursor, _ = franklin.stream_local_rows(
        workbook,
        record_family="daily-conveyance",
        release_id="daily-conveyance-2026-07-23",
        release_date="2026-07-23",
        source_url=None,
        member=None,
        sheet_name=None,
        query=None,
        parcel="01000000100",
        from_date="2026-07-23",
        to_date="2026-07-23",
        limit=None,
        cursor=None,
        header_scan_rows=25,
        encoding="utf-8-sig",
    )

    assert cursor is None
    assert len(rows) == 1
    row = rows[0]
    assert row["header_row_number"] == 2
    assert row["source_row_number"] == 3
    assert row["parsed_fields"]["amount"] == 450000
    assert row["parsed_fields"]["owner_names"] == ["NEW OWNER"]
    assert row["parsed_fields"]["prior_owner_names"] == ["PRIOR OWNER"]
    assert row["parsed_fields"]["instrument"] == "WD"
    assert row["parsed_fields"]["conveyance_number"] == "C-100"
    assert row["parsed_fields"]["instrument_number"] is None
    assert row["parsed_fields"]["is_exempt"] == "NON-EXEMPT"
    assert row["parsed_fields"]["sale_type"] == "LAND AND BUILDING"
    assert row["parsed_fields"]["source_sale_flags"] == {
        "is_exempt": "NON-EXEMPT",
        "sale_type": "LAND AND BUILDING",
        "sale_validity": None,
    }
    assert row["same_authority_lineage"] == (
        "us-oh-franklin-county-auditor-property"
    )


def test_live_appraisal_sales_header_and_validity_codes_are_preserved(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "Sales010.xlsx"
    _write_appraisal_sales_workbook(workbook)

    rows, cursor, _ = franklin.stream_local_rows(
        workbook,
        record_family="sales",
        release_id="appraisal-2026-07-15",
        release_date="2026-07-15",
        source_url=None,
        member=None,
        sheet_name=None,
        query=None,
        parcel=None,
        from_date=None,
        to_date=None,
        limit=None,
        cursor=None,
        header_scan_rows=25,
        encoding="utf-8-sig",
    )

    assert cursor is None
    assert len(rows) == 3
    assert rows[0]["raw_headers"] == [
        "PARCEL ID",
        "MAP ROUTING",
        "SALEDT",
        "NOPAR",
        "INSTRUMENT",
        "INSTRUNO",
        "VALID",
        "SALETYPE",
        "PRICE",
        "ADJAMT",
        "ADJPRICE",
        "CONDSALE_GRANTORRELATIVE",
        "CONDSALE_LIFESTATE",
        "CONDSALE_PARTINTEREST",
        "CONDSALE_LEASEDFEE",
        "CONDSALE_GRANTORMORTGAGEE",
        "CONDSALE_LANDCONTRACT",
        "CONDSALE_LEASEHOLD",
        "CONDSALE_TRADE",
        "CONDSALE_MINERALRIGHTSRESERVED",
        "CONDSALE_OTHER",
    ]
    assert [row["parsed_fields"]["event_date"] for row in rows] == [
        "1986-05-01",
        "1995-06-02",
        "2006-07-03",
    ]
    assert [row["parsed_fields"]["sale_validity"] for row in rows] == [
        "0 - VALID",
        "99 - RMS INVALID",
        None,
    ]
    assert [row["parsed_fields"]["instrument_number"] for row in rows] == [
        "202607010000001",
        "202607020000002",
        "202607030000003",
    ]
    assert all(
        row["parsed_fields"]["conveyance_number"] is None for row in rows
    )


def test_zip_member_is_explicit_and_part_of_row_identity(tmp_path: Path) -> None:
    workbook = tmp_path / "Payment2025.xlsx"
    archive_path = tmp_path / "Excel.zip"
    _write_payment_workbook(workbook)
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(workbook, arcname="Tax Accounting/Payment2025.xlsx")

    inventory = franklin.inspect_local_artifact(
        archive_path,
        record_family=None,
        member=None,
        sheet_name=None,
        sample_rows=0,
        header_scan_rows=25,
        encoding="utf-8-sig",
    )
    assert inventory["schema_state"] == "member_selection_required"
    assert inventory["archive_members"][0]["name"] == (
        "Tax Accounting/Payment2025.xlsx"
    )

    rows, cursor, _ = franklin.stream_local_rows(
        archive_path,
        record_family="payment",
        release_id="tax-accounting-2026-07-15",
        release_date="2026-07-15",
        source_url=None,
        member="Tax Accounting/Payment2025.xlsx",
        sheet_name=None,
        query=None,
        parcel="010-000001-00",
        from_date=None,
        to_date=None,
        limit=None,
        cursor=None,
        header_scan_rows=25,
        encoding="utf-8-sig",
    )
    assert cursor is None
    assert rows[0]["artifact_filename"] == "Excel.zip"
    assert rows[0]["archive_member"]["archive_member"] == (
        "Tax Accounting/Payment2025.xlsx"
    )
    assert '"archive_member":"Tax Accounting/Payment2025.xlsx"' in (
        rows[0]["native_occurrence"]
    )
    assert '"worksheet":"Sheet1"' in rows[0]["native_occurrence"]


def test_zip_xlsx_occurrence_identity_includes_member_and_worksheet(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "multi-sheet.xlsx"
    archive_path = tmp_path / "Excel.zip"
    workbook = Workbook()
    for index, title in enumerate(("Current", "Archive")):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = title
        sheet.append(
            ["Parcel Id", "EffectiveDate", "TaxYear", "BillType", "Amount"]
        )
        sheet.append(["010-000001-00", "2026-07-01", 2024, "REAL", 100.25])
    workbook.save(workbook_path)
    member = "Tax Accounting/Payment2025.xlsx"
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(workbook_path, arcname=member)

    records = []
    for sheet_name in ("Current", "Archive"):
        rows, cursor, _ = franklin.stream_local_rows(
            archive_path,
            record_family="payment",
            release_id="tax-accounting-2026-07-15",
            release_date="2026-07-15",
            source_url=None,
            member=member,
            sheet_name=sheet_name,
            query=None,
            parcel=None,
            from_date=None,
            to_date=None,
            limit=None,
            cursor=None,
            header_scan_rows=25,
            encoding="utf-8-sig",
        )
        assert cursor is None
        records.append(rows[0])

    assert records[0]["native_occurrence"] != records[1]["native_occurrence"]
    assert records[0]["native_document_id"] != records[1]["native_document_id"]
    for record, sheet_name in zip(records, ("Current", "Archive"), strict=True):
        assert f'"worksheet":"{sheet_name}"' in record["native_occurrence"]
        assert f'"archive_member":"{member}"' in record["native_occurrence"]


def test_unrecognized_row_schema_is_source_changed(tmp_path: Path) -> None:
    workbook = tmp_path / "changed.xlsx"
    book = Workbook()
    book.active.append(["Mystery", "Columns"])
    book.active.append(["a", "b"])
    book.save(workbook)

    with pytest.raises(franklin.FranklinSourceChanged):
        franklin.inspect_local_artifact(
            workbook,
            record_family="payment",
            member=None,
            sheet_name=None,
            sample_rows=1,
            header_scan_rows=10,
            encoding="utf-8-sig",
        )


def test_bounded_probe_covers_each_current_family_without_large_downloads() -> None:
    client = FixtureDirectoryClient()
    bulk = FixtureBulkClient()
    args = _args("probe", "--sample-bytes", "32")

    result = franklin.execute(
        args,
        directory_client=client,
        bulk_client=bulk,
        log_results=False,
    )

    assert result.status == ResultStatus.OK
    record = result.records[0]
    assert set(record["current_releases"]) == set(franklin.FAMILY_CHOICES)
    assert record["large_artifacts_downloaded"] is False
    assert record["sample_bytes_requested"] == 32
    assert len(bulk.probes) == 1


@pytest.mark.parametrize(
    ("bulk", "error_fragment"),
    [
        (FixtureBulkClient(content_length=36_543), "size differs"),
        (
            FixtureBulkClient(
                signature_hex="3c68746d6c3e",
                format_hint="html",
            ),
            "ZIP container signature",
        ),
    ],
)
def test_bounded_source_probe_rejects_artifact_transport_drift(
    bulk: FixtureBulkClient,
    error_fragment: str,
) -> None:
    result = franklin.execute(
        _args("probe", "--sample-bytes", "32"),
        directory_client=FixtureDirectoryClient(),
        bulk_client=bulk,
        log_results=False,
    )

    assert result.status == ResultStatus.SOURCE_CHANGED
    assert error_fragment in result.errors[0].message
