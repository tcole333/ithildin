from __future__ import annotations

import hashlib
import json
import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook

from tools import ingest_harris_court_bulk


FIXTURE_ROOT = (
    Path(__file__).parent
    / "fixtures"
    / "public_records"
    / "harris_court_bulk"
    / "ingest"
)
FIXTURES = (
    (
        "civil_case_summary.tsv",
        r"Civil\CaseSummaryMods_Daily-2026-07-30.txt",
        "civil",
    ),
    (
        "civil_party.tsv",
        r"Civil\PartyMods_Daily-2026-07-30.txt",
        "civil",
    ),
    (
        "civil_activity.tsv",
        r"Civil\ActivityMods_Daily-2026-07-30.txt",
        "civil",
    ),
    (
        "criminal_filings.tsv",
        r"Criminal\2026-07-30 CrimFilingsDaily_withHeadings.txt",
        "criminal",
    ),
    (
        "criminal_dispositions.tsv",
        r"Criminal\2026-07-30 CrimDisposDaily_withHeadings.txt",
        "criminal",
    ),
)


def _worksheet(workbook: Workbook, name: str, rows: list[tuple[str, ...]]) -> None:
    worksheet = workbook.create_sheet(name)
    worksheet.append(("code", "code literal"))
    for row in rows:
        worksheet.append(row)


def _civil_codebook(path: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _worksheet(workbook, "cst", [("A", "ACTIVE"), ("D", "DISPOSED")])
    _worksheet(workbook, "cs_typ", [("CV", "OTHER CIVIL")])
    _worksheet(workbook, "toac", [("MAR", "MASS ACTION")])
    _worksheet(
        workbook,
        "coc",
        [
            ("DEF", "DEFENDANT - Civil"),
            ("ATD", "Attorney For Defendant - CIVIL"),
        ],
    )
    _worksheet(workbook, "act", [("ORD", "ORDER SIGNED")])
    _worksheet(workbook, "judgment", [("DISM", "DISMISSED")])
    workbook.save(path)
    return path


def _criminal_codebook(path: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _worksheet(
        workbook,
        "CDI",
        [("2", "MISDEMEANOR"), ("3", "FELONY")],
    )
    _worksheet(workbook, "INS", [("MIN", "MISDEMEANOR INFORMATION")])
    _worksheet(workbook, "CAD", [("DISM", "DISMISSED")])
    _worksheet(
        workbook,
        "CST",
        [("A", "ACTIVE"), ("D", "DISMISSED")],
    )
    _worksheet(workbook, "DST", [("B", "BOND MADE"), ("D", "DISPOSED")])
    _worksheet(workbook, "CURR_L_D", [("M", "MISDEMEANOR")])
    _worksheet(workbook, "CNC", [("MO", "MOTIONS DOCKET")])
    _worksheet(workbook, "REA", [("ARRG", "ARRAIGNMENT")])
    workbook.save(path)
    return path


def _args(
    fixture: str,
    locator: str,
    codebook: Path,
    court_db: Path,
    *extra: str,
):
    return ingest_harris_court_bulk.build_parser().parse_args(
        [
            "ingest",
            str(FIXTURE_ROOT / fixture),
            "--native-locator",
            locator,
            "--published-date",
            "2026-07-30",
            "--schema-workbook",
            str(codebook),
            "--retrieved-at",
            "2026-07-30T12:00:00Z",
            "--court-db",
            str(court_db),
            "--batch-size",
            "1",
            *extra,
        ]
    )


def _table_counts(path: Path) -> dict[str, int]:
    db = sqlite3.connect(path)
    try:
        return {
            table: db.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            for table in (
                "source_snapshot",
                "harris_court_bulk_artifact",
                "harris_court_bulk_row",
                "court",
                "case_record",
                "case_party",
                "attorney",
                "case_representation",
                "docket_entry",
                "case_event",
                "document_artifact",
            )
        }
    finally:
        db.close()


def test_priority_families_stream_to_generic_sidecar_and_repeat_idempotently(
    tmp_path: Path,
) -> None:
    civil = _civil_codebook(tmp_path / "civil.xlsx")
    criminal = _criminal_codebook(tmp_path / "criminal.xlsx")
    court_db = tmp_path / "courts.db"
    first_results = []
    for fixture, locator, codebook_kind in FIXTURES:
        codebook = civil if codebook_kind == "civil" else criminal
        first_results.append(
            ingest_harris_court_bulk.execute(
                _args(fixture, locator, codebook, court_db)
            )
        )
    first_counts = _table_counts(court_db)
    second_results = []
    for fixture, locator, codebook_kind in FIXTURES:
        codebook = civil if codebook_kind == "civil" else criminal
        second_results.append(
            ingest_harris_court_bulk.execute(
                _args(fixture, locator, codebook, court_db)
            )
        )
    second_counts = _table_counts(court_db)

    assert first_results[0]["artifact"]["dataset_family"] == "case_summary"
    assert first_results[2]["artifact"]["dataset_family"] == "activity"
    assert first_results[3]["artifact"]["dataset_family"] == "filings"
    assert first_results[4]["artifact"]["dataset_family"] == "dispositions"
    assert all(result["exhausted"] is True for result in first_results)
    assert all(
        result["artifact_snapshot_inserted"] is False
        for result in second_results
    )
    assert sum(
        result["counts"]["row_occurrences_inserted"]
        for result in first_results
    ) == 6
    assert sum(
        result["counts"]["row_occurrences_reused"]
        for result in second_results
    ) == 6
    assert second_counts == first_counts
    assert first_counts == {
        "source_snapshot": 5,
        "harris_court_bulk_artifact": 5,
        "harris_court_bulk_row": 6,
        "court": 2,
        "case_record": 2,
        "case_party": 5,
        "attorney": 2,
        "case_representation": 2,
        "docket_entry": 2,
        "case_event": 3,
        "document_artifact": 0,
    }

    db = sqlite3.connect(court_db)
    db.row_factory = sqlite3.Row
    try:
        cases = {
            row["raw_case_number"]: dict(row)
            for row in db.execute("SELECT * FROM case_record")
        }
        assert cases["202600001"]["filing_date"] == "2026-07-25"
        assert cases["202600001"]["status"] == "ACTIVE"
        assert cases["263164001010"]["filing_date"] == "2026-07-25"
        assert cases["263164001010"]["disposition_date"] == "2026-07-30"
        assert cases["263164001010"]["status"] == "DISMISSED"
        attorneys = {
            (row["raw_name"], row["bar_id"])
            for row in db.execute("SELECT raw_name, bar_id FROM attorney")
        }
        assert attorneys == {
            ("COUNSEL, CASEY", ""),
            ("COUNSEL, CASEY", "24000001"),
        }

        compact_date_row = db.execute(
            """
            SELECT normalized_json
            FROM harris_court_bulk_row
            WHERE artifact_id IN (
                SELECT artifact_id FROM harris_court_bulk_artifact
                WHERE dataset_family='filings'
            )
            """
        ).fetchone()
        compact = json.loads(compact_date_row["normalized_json"])
        assert compact["dates"]["fda"] == {
            "iso": "2026-07-25",
            "precision": "day",
            "raw": "072526",
            "source_format": "mmddyy_current_layout",
            "state": "parsed",
            "value": "072526",
        }
        assert compact["codes"]["current_offense"]["code"] == "540471"
        assert (
            compact["codes"]["current_offense"]["source_literal"]
            == "DWI 1ST OFFENDER BAC>=0.15"
        )

        disposition_row = db.execute(
            """
            SELECT raw_fields_json, normalized_json
            FROM harris_court_bulk_row
            WHERE artifact_id IN (
                SELECT artifact_id FROM harris_court_bulk_artifact
                WHERE dataset_family='dispositions'
            )
            """
        ).fetchone()
        raw_disposition = json.loads(disposition_row["raw_fields_json"])
        normalized_disposition = json.loads(disposition_row["normalized_json"])
        assert raw_disposition["fields"]["nda"] == "NULL"
        assert normalized_disposition["dates"]["nda"]["state"] == (
            "source_null_literal"
        )

        activity_rows = list(
            db.execute(
                """
                SELECT row_occurrence_id, row_sha256
                FROM harris_court_bulk_row
                WHERE artifact_id IN (
                    SELECT artifact_id FROM harris_court_bulk_artifact
                    WHERE dataset_family='activity'
                )
                ORDER BY source_row_number
                """
            )
        )
        assert len(activity_rows) == 2
        assert activity_rows[0]["row_sha256"] == activity_rows[1]["row_sha256"]
        assert (
            activity_rows[0]["row_occurrence_id"]
            != activity_rows[1]["row_occurrence_id"]
        )
    finally:
        db.close()


def test_caller_checkpoint_preserves_duplicate_occurrences_without_duplicate_docket(
    tmp_path: Path,
) -> None:
    civil = _civil_codebook(tmp_path / "civil.xlsx")
    court_db = tmp_path / "courts.db"
    fixture, locator, _kind = FIXTURES[2]
    first = ingest_harris_court_bulk.execute(
        _args(fixture, locator, civil, court_db, "--limit", "1")
    )
    second = ingest_harris_court_bulk.execute(
        _args(fixture, locator, civil, court_db, "--start-row", "1")
    )

    assert first["exhausted"] is False
    assert first["next_checkpoint_row"] == 1
    assert first["counts"]["rows_processed"] == 1
    assert second["exhausted"] is True
    assert second["next_checkpoint_row"] is None
    assert second["counts"]["rows_skipped_before_checkpoint"] == 1
    counts = _table_counts(court_db)
    assert counts["harris_court_bulk_row"] == 2
    assert counts["docket_entry"] == 1


def test_artifact_result_supplies_and_verifies_exact_download_provenance(
    tmp_path: Path,
) -> None:
    source = FIXTURE_ROOT / "civil_case_summary.tsv"
    artifact = tmp_path / "CaseSummaryMods_Daily-2026-07-30.txt"
    artifact.write_bytes(source.read_bytes())
    digest = hashlib.sha256(artifact.read_bytes()).hexdigest()
    result_path = tmp_path / "download.json"
    result_path.write_text(
        json.dumps(
            {
                "status": "ok",
                "records": [
                    {
                        "source_id": ingest_harris_court_bulk.SOURCE_ID,
                        "native_locator": (
                            r"Civil\CaseSummaryMods_Daily-2026-07-30.txt"
                        ),
                        "published_date": "2026-07-30",
                        "source_url": ingest_harris_court_bulk.CATALOG_URL,
                        "artifact_receipt": {
                            "path": str(artifact),
                            "size": artifact.stat().st_size,
                            "sha256": digest,
                            "source_url": ingest_harris_court_bulk.CATALOG_URL,
                        },
                    }
                ],
            }
        )
    )
    args = ingest_harris_court_bulk.build_parser().parse_args(
        [
            "ingest",
            str(artifact),
            "--artifact-result",
            str(result_path),
            "--court-db",
            str(tmp_path / "courts.db"),
        ]
    )

    payload = ingest_harris_court_bulk.execute(args)

    assert payload["artifact"]["artifact_sha256"] == digest
    assert payload["artifact"]["native_locator"] == (
        r"Civil\CaseSummaryMods_Daily-2026-07-30.txt"
    )
    assert payload["counts"]["rows_processed"] == 1


def test_missing_projection_column_fails_before_snapshot_or_case_projection(
    tmp_path: Path,
) -> None:
    malformed = tmp_path / "activity.txt"
    malformed.write_text("casenbr\tcode\tdesc\n202600001\tORD\tORDER SIGNED\n")
    court_db = tmp_path / "courts.db"
    args = ingest_harris_court_bulk.build_parser().parse_args(
        [
            "ingest",
            str(malformed),
            "--native-locator",
            r"Civil\ActivityMods_Daily-2026-07-30.txt",
            "--published-date",
            "2026-07-30",
            "--court-db",
            str(court_db),
        ]
    )

    with pytest.raises(
        ingest_harris_court_bulk.HarrisCourtBulkIngestError,
        match="header lacks required source columns",
    ):
        ingest_harris_court_bulk.execute(args)

    db = sqlite3.connect(court_db)
    try:
        assert db.execute("SELECT COUNT(*) FROM source_snapshot").fetchone()[0] == 0
        assert db.execute("SELECT COUNT(*) FROM case_record").fetchone()[0] == 0
    finally:
        db.close()


def test_undefined_cp1252_source_byte_is_preserved_without_replacing_row(
    tmp_path: Path,
) -> None:
    artifact = tmp_path / "ActivityMods_Daily-2026-07-30.txt"
    source = (FIXTURE_ROOT / "civil_activity.tsv").read_bytes()
    artifact.write_bytes(
        source.replace(b"ORDER SIGNED", b"ORDER\x90SIGNED", 1)
    )
    args = ingest_harris_court_bulk.build_parser().parse_args(
        [
            "ingest",
            str(artifact),
            "--native-locator",
            r"Civil\ActivityMods_Daily-2026-07-30.txt",
            "--published-date",
            "2026-07-30",
            "--court-db",
            str(tmp_path / "courts.db"),
        ]
    )

    payload = ingest_harris_court_bulk.execute(args)

    assert payload["counts"]["rows_processed"] == 2
    assert payload["schema"]["undefined_cp1252_byte_counts"] == {"0x90": 1}
    assert payload["schema"]["decode_error_policy"].startswith(
        "undefined_cp1252_bytes_preserved"
    )
    db = sqlite3.connect(tmp_path / "courts.db")
    try:
        raw_json = db.execute(
            """
            SELECT raw_fields_json
            FROM harris_court_bulk_row
            ORDER BY source_row_number
            LIMIT 1
            """
        ).fetchone()[0]
    finally:
        db.close()
    assert json.loads(raw_json)["fields"]["desc"] == "ORDER\x90SIGNED"
